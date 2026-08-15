from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from .schemas import SendPlan


class ExecutionPolicyDenied(RuntimeError):
    """Raised before approval consumption when a supervised execution is unsafe."""


@dataclass(frozen=True)
class ApprovedTextTask:
    workbook: str
    row_number: int
    target: str
    message: str


@dataclass(frozen=True)
class ApprovedBatchTask:
    workbook: str
    tasks: tuple[ApprovedTextTask, ...]

    @property
    def count(self) -> int:
        return len(self.tasks)


class SupervisedExecutionPolicy:
    """Hard application boundary for the first supervised-send milestone."""

    def __init__(self, allowed_targets: Iterable[str]) -> None:
        self.allowed_targets = frozenset(value.strip() for value in allowed_targets if value.strip())

    def validate(self, plan: SendPlan) -> ApprovedTextTask:
        batch = _validate_batch(plan, self.allowed_targets, max_rows=1, allow_resume=False)
        return batch.tasks[0]


class SupervisedBatchExecutionPolicy:
    """Safety boundary for a confirmed batch of up to ten text contacts."""

    def __init__(self, allowed_targets: Iterable[str], max_rows: int = 10) -> None:
        self.allowed_targets = frozenset(value.strip() for value in allowed_targets if value.strip())
        self.max_rows = max_rows

    def validate(self, plan: SendPlan) -> ApprovedBatchTask:
        return _validate_batch(plan, self.allowed_targets, max_rows=self.max_rows, allow_resume=True)


def _validate_batch(
    plan: SendPlan,
    allowed_targets: frozenset[str],
    *,
    max_rows: int,
    allow_resume: bool,
) -> ApprovedBatchTask:
    if not allowed_targets:
        raise ExecutionPolicyDenied("WECOM_ALLOWED_TARGETS must contain at least one exact test target")
    if not plan.workbook or plan.lesson_workbook or plan.target_workbook or plan.lesson:
        raise ExecutionPolicyDenied("Supervised execution requires one direct workbook")
    if plan.schedule_mode != "immediate" or plan.run_at:
        raise ExecutionPolicyDenied("Supervised execution only permits immediate delivery")
    if plan.resend:
        raise ExecutionPolicyDenied("Supervised execution does not permit resend")

    workbook = load_workbook(Path(plan.workbook), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    if not rows:
        raise ExecutionPolicyDenied("Workbook is empty")
    headers = {_normalize(value): index for index, value in enumerate(rows[0]) if _normalize(value)}
    data_rows = [
        (row_number, row)
        for row_number, row in enumerate(rows[1:], start=2)
        if any(_text(value) for value in row)
    ]
    if not allow_resume and len(data_rows) != 1:
        raise ExecutionPolicyDenied("Supervised single execution requires exactly one non-empty data row")
    selected = [item for item in data_rows if _row_selected(item[0], plan)]
    if not selected:
        raise ExecutionPolicyDenied("The workbook has no selected data rows")

    approved: list[ApprovedTextTask] = []
    for row_number, row in selected:
        status = _field(row, headers, ("发送状态",))
        if status:
            if not allow_resume:
                raise ExecutionPolicyDenied("Selected row must not have a prior send status")
            if status in {"已发送", "已重发", "发送待核对", "部分发送", "超时已发送"}:
                continue
            if status not in {"失败", "发送失败"}:
                raise ExecutionPolicyDenied(f"Unsupported prior send status at row {row_number}: {status}")
        send_flag = _field(row, headers, ("是否发送", "发送", "执行", "是否执行"))
        if send_flag.lower() not in {"是", "true", "1", "yes", "y"}:
            if allow_resume:
                continue
            raise ExecutionPolicyDenied("Selected row must be explicitly enabled for sending")
        approved.append(_validate_text_row(plan, allowed_targets, headers, row_number, row))

    if not approved:
        raise ExecutionPolicyDenied("The batch has no pending rows to send")
    if len(approved) > max_rows:
        raise ExecutionPolicyDenied(f"Supervised batch permits at most {max_rows} pending rows")
    targets = [task.target for task in approved]
    if len(set(targets)) != len(targets):
        raise ExecutionPolicyDenied("Supervised batch does not permit duplicate targets")
    return ApprovedBatchTask(str(plan.workbook), tuple(approved))


def _validate_text_row(
    plan: SendPlan,
    allowed_targets: frozenset[str],
    headers: dict[str, int],
    row_number: int,
    row: tuple[Any, ...],
) -> ApprovedTextTask:
    channel = _field(row, headers, ("渠道",))
    if channel.lower() not in {"企业微信", "企微", "wecom", "wxwork"}:
        raise ExecutionPolicyDenied("The selected row must use the WeCom channel")
    object_type = _field(row, headers, ("对象类型",))
    if object_type not in {"个人", "联系人"}:
        raise ExecutionPolicyDenied("Supervised execution only permits an individual contact")
    target = _field(row, headers, ("发送对象", "对象", "联系人", "联系人姓名", "姓名"))
    if target not in allowed_targets:
        raise ExecutionPolicyDenied("Target is not in the exact supervised-send allowlist")
    message_type = _field(row, headers, ("消息类型", "信息类型"))
    if message_type.lower() not in {"文字", "文本", "text"}:
        raise ExecutionPolicyDenied("Supervised execution only permits plain text")
    message = _field(row, headers, ("发送内容", "文本内容", "消息", "消息内容", "文案", "内容", "正文"))
    if not message:
        raise ExecutionPolicyDenied("Plain-text message must not be empty")
    if _field(row, headers, ("图片", "图片路径", "图片文件")):
        raise ExecutionPolicyDenied("Image attachments are not permitted")
    if _field(row, headers, ("文档", "文件", "附件", "文件路径")):
        raise ExecutionPolicyDenied("Document attachments are not permitted")
    if _field(row, headers, ("计划发送时间", "发送时间", "定时发送时间")):
        raise ExecutionPolicyDenied("Workbook scheduling is not permitted")
    return ApprovedTextTask(str(plan.workbook), row_number, target, message)


def parse_allowed_targets(value: str | None) -> tuple[str, ...]:
    return tuple(part.strip() for part in (value or "").split(",") if part.strip())


def _row_selected(row_number: int, plan: SendPlan) -> bool:
    selection = plan.selection
    if selection.rows:
        return row_number in selection.rows
    start = selection.row_from or 2
    return row_number >= start and (selection.row_to is None or row_number <= selection.row_to)


def _field(row: tuple[Any, ...], headers: dict[str, int], aliases: tuple[str, ...]) -> str:
    for alias in aliases:
        index = headers.get(_normalize(alias))
        if index is not None and index < len(row):
            return _text(row[index])
    return ""


def _normalize(value: Any) -> str:
    return _text(value).replace(" ", "").lower()


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()
