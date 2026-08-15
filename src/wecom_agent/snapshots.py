from __future__ import annotations

from dataclasses import asdict, dataclass
import hashlib
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .schemas import SendPlan


class SnapshotMismatch(RuntimeError):
    """Raised when a workbook changed after preview and approval."""


@dataclass(frozen=True)
class FileSnapshot:
    path: str
    sha256: str
    size: int
    modified_ns: int
    selected_rows_sha256: str


@dataclass(frozen=True)
class PlanSnapshot:
    files: tuple[FileSnapshot, ...]
    digest: str

    def as_dict(self) -> dict[str, Any]:
        return {"files": [asdict(item) for item in self.files], "digest": self.digest}

    @classmethod
    def from_dict(cls, value: dict[str, Any]) -> "PlanSnapshot":
        return cls(files=tuple(FileSnapshot(**item) for item in value["files"]), digest=value["digest"])


def capture_plan_snapshot(plan: SendPlan) -> PlanSnapshot:
    paths = [Path(value) for value in (plan.workbook, plan.lesson_workbook, plan.target_workbook) if value]
    snapshots: list[FileSnapshot] = []
    selected_path = Path(plan.workbook or plan.target_workbook or "").resolve()
    for path in paths:
        resolved = path.resolve()
        stat = resolved.stat()
        snapshots.append(
            FileSnapshot(
                path=str(resolved),
                sha256=_file_sha256(resolved),
                size=stat.st_size,
                modified_ns=stat.st_mtime_ns,
                selected_rows_sha256=_selected_rows_sha256(resolved, plan) if resolved == selected_path else "",
            )
        )
    canonical = json.dumps([asdict(item) for item in snapshots], ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    digest = hashlib.sha256(canonical.encode("utf-8")).hexdigest()
    return PlanSnapshot(files=tuple(snapshots), digest=digest)


def assert_snapshot_unchanged(plan: SendPlan, expected: PlanSnapshot) -> PlanSnapshot:
    current = capture_plan_snapshot(plan)
    if current.digest != expected.digest:
        raise SnapshotMismatch("One or more approved workbooks changed after preview")
    return current


def delivery_key(
    plan_hash: str,
    snapshot_digest: str,
    row_number: int,
    target: str,
    message_hash: str,
) -> str:
    payload = "|".join([plan_hash, snapshot_digest, str(row_number), target.strip(), message_hash])
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def execution_idempotency_key(task_id: str, plan_hash: str, snapshot_digest: str) -> str:
    payload = "|".join([task_id, plan_hash, snapshot_digest])
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _selected_rows_sha256(path: Path, plan: SendPlan) -> str:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        sheet = workbook.active
        explicit_rows = set(plan.selection.rows)
        start = plan.selection.row_from or 2
        end = plan.selection.row_to
        values = []
        for row_number, row in enumerate(sheet.iter_rows(), start=1):
            selected = row_number in explicit_rows if explicit_rows else row_number >= start and (end is None or row_number <= end)
            if selected:
                values.append([_json_value(cell.value) for cell in row])
        canonical = json.dumps(values, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
        return hashlib.sha256(canonical.encode("utf-8")).hexdigest()
    finally:
        workbook.close()


def _json_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)
