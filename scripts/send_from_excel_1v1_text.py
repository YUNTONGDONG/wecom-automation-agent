#!/usr/bin/env python3
"""Fast Excel-driven Enterprise WeChat sender.

The hot path is intentionally one-shot:
1. Load the workbook once.
2. Build eligible rows from Excel columns.
3. Open/focus WeCom once, search targets only when needed.
4. Send text/images/documents for each row.
5. Save result columns once at the end and append compact logs.

Use --execute --yes for a real send. Without --execute the script only previews.
"""

from __future__ import annotations

import argparse
import ctypes
import difflib
import importlib.util
import json
import platform
import re
import shutil
import subprocess
import sys
import time
import uuid
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook

SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

from excel_utils import (
    col_value,
    ensure_columns,
    find_workbook,
    first_col,
    parse_parts,
    resolve_paths,
    split_paths,
    worksheet_headers,
)
from log_utils import (
    log_line,
    save_evidence_manifest,
    save_run_result,
)

LOG_ROOT_RELATIVE_PATH = Path("wechat-work-message-validation/logs")
OCR_SCRIPT_PATH = Path(__file__).with_name("ocr_windows_image.ps1")
MACOS_VISION_OCR_SCRIPT_PATH = Path(__file__).with_name("macos_vision_ocr.py")
DEFAULT_SCHEDULE_ON_TIME_TOLERANCE_SECONDS = 60.0
DEFAULT_SINGLE_SCHEDULE_PREPARE_LEAD_SECONDS = 12.0
DEFAULT_BATCH_SCHEDULE_PREPARE_MIN_SECONDS = 30.0
DEFAULT_BATCH_SCHEDULE_PREPARE_PER_TARGET_SECONDS = 10.0
DEFAULT_BATCH_SCHEDULE_PREPARE_MAX_SECONDS = 300.0
DEFAULT_SCHEDULE_FINAL_TARGET_CHECK_LEAD_SECONDS = 4.0
DEFAULT_BATCH_FAST_DISPATCH_THRESHOLD = 2
DEFAULT_BATCH_DISPATCH_ESTIMATE_PER_TARGET_SECONDS = 4.0
DEFAULT_PASTE_METHOD_ORDER = "ctrl-v,ctrl-v,ctrl-v"
DEFAULT_TEST_DISCLAIMER = "以上为测试，可忽略谢谢。"
GENERATED_SEND_WORKBOOK_PREFIXES = ("send_lesson", "send_第")


class OcrUnavailableError(RuntimeError):
    """Raised when all configured OCR engines are unavailable."""


class OcrNoTextError(RuntimeError):
    """Raised when OCR engines ran but produced no usable text."""


class GuiUnavailableError(RuntimeError):
    """Raised when the messaging app window is not available for automation."""


class WeComBlockedError(RuntimeError):
    """Enterprise WeChat is not in a sendable state, so the batch must stop."""


CHANNEL_HEADERS = ("渠道",)
OBJECT_TYPE_HEADERS = ("对象类型",)
CONTACT_HEADERS = ("联系人", "联系人姓名", "姓名")
GROUP_HEADERS = ("群聊名称", "群名称", "群聊", "群")
TARGET_HEADERS = ("发送对象", "对象")
TARGET_ALIAS_HEADERS = ("目标别名", "核对关键词", "会话关键词", "英文名", "别名")
TEXT_HEADERS = ("发送内容", "文本内容", "消息", "消息内容", "文案", "内容", "正文")
MESSAGE_TYPE_HEADERS = ("消息类型", "信息类型")
IMAGE_HEADERS = ("图片", "图片路径", "图片文件")
DOCUMENT_HEADERS = ("文档", "文件", "附件", "文件路径")
SEND_FLAG_HEADERS = ("是否发送", "发送", "执行", "是否执行")
STATUS_HEADERS = ("发送状态",)
ERROR_HEADERS = ("错误原因",)
TIME_HEADERS = ("发送时间",)
SCHEDULE_HEADERS = ("计划发送时间", "定时发送时间", "预定发送时间", "发送时间要求")

STATUS_COLUMNS = (
    "发送状态",
    "错误原因",
    "发送时间",
    "目标会话核对",
    "发送后核对",
    "执行批次",
    "执行模式",
    "发送完成时间",
    "发送完成耗时秒",
    "验证完成时间",
    "验证完成耗时秒",
)

CHANNEL_ALIASES = {"企业微信", "企微", "WXWork", "WeCom", "wecom", "wxwork"}
PERSON_TYPES = {"个人", "1v1", "联系人", "单聊", ""}
GROUP_TYPES = {"群聊", "群", "群组"}
NO_SEND_FLAGS = {"否", "不发送", "false", "0", "no", "n"}
SENT_STATUSES = {"已发送", "已重发", "发送待核对", "部分发送", "超时已发送"}
SEND_ERROR_MARKERS = (
    "发送失败",
    "不能发送",
    "无法发送",
    "空白消息",
    "设备环境异常",
    "安全验证",
    "发送异常",
)
SEND_PENDING_MARKERS = ("发送中", "正在发送", "上传", "正在上传", "等待发送", "解析中", "正在解析")
WECOM_GLOBAL_BLOCK_MARKERS = (
    "设备环境异常",
    "安全验证",
    "扫码",
    "退出登录",
)
TITLE_ROLE_MARKERS = (
    "经理",
    "负责人",
    "总监",
    "主管",
    "助理",
    "老师",
    "顾问",
    "助教",
    "讲师",
    "导师",
    "班主任",
    "运营",
    "销售",
    "商务",
    "客服",
    "教务",
    "咨询",
    "专员",
    "组长",
    "主任",
    "创始人",
    "合伙人",
    "CEO",
    "COO",
    "CTO",
    "CFO",
    "hr",
    "HR",
)
TITLE_DEPARTMENT_MARKERS = (
    "部门",
    "事业部",
    "中心",
    "团队",
    "小组",
    "项目组",
    "办公室",
    "公司",
    "学校",
    "学院",
    "校区",
    "门店",
    "分店",
    "总部",
    "分部",
    "区域",
    "华北",
    "华东",
    "华南",
    "华中",
    "西南",
    "东北",
    "北京",
    "上海",
    "广州",
    "深圳",
    "成都",
    "杭州",
    "武汉",
    "西安",
    "南京",
    "苏州",
)
TITLE_NOTE_MARKERS = (
    "备注",
    "昵称",
    "实名",
    "联系人",
    "外部联系人",
    "企业微信",
    "微信",
    "@",
    "-",
    "_",
    "｜",
    "|",
    "/",
    "（",
    "(",
    "java",
    "Java",
    "JAVA",
    "前端",
    "后端",
    "开发",
    "测试",
    "产品",
)
TITLE_CONTEXT_MARKERS = (
    *TITLE_ROLE_MARKERS,
    *TITLE_DEPARTMENT_MARKERS,
    *TITLE_NOTE_MARKERS,
)
TITLE_STATUS_MARKERS = (
    "上班中",
    "离线",
    "在线",
    "忙碌",
    "外出",
    "请假",
)
TITLE_METADATA_PREFIXES = (
    "老师",
    "负责人",
    "经理",
    "主管",
    "总监",
    "运营",
    "顾问",
    "助教",
    "教务",
    "客服",
    "销售",
    "商务",
    "前端工程师",
    "后端工程师",
    "测试工程师",
    "开发工程师",
    "软件工程师",
    "工程师",
    "项目经理",
    "产品经理",
    "产品",
)
SEARCH_GLOBAL_MARKERS = (
    "进入全局搜索",
    "查找所有聊天",
    "ctrlaltf",
    "ctrl+alt+f",
    "go to global search",
    "search for all chats",
    "all chats",
)
SEARCH_NON_CHAT_RESULT_MARKERS = (
    "mp.weixin.qq.com",
    "weixin.qq.com",
    "http://",
    "https://",
    "www.",
    "公众号",
    "订阅号",
    "服务号",
    "文章",
    "网页",
    "小程序",
)
OCR_CONFUSIONS = str.maketrans(
    {
        "芸": "云",
        "靶": "云",
        "郊": "郑",
        "消": "洋",
    }
)


def now_text() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def timestamp_slug() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def input_locator_cache_path_for_folder(folder: Path) -> Path:
    return folder / LOG_ROOT_RELATIVE_PATH.parent / "input_locator_cache.json"


def run_minute_slug() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M")


def cell_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def parse_scheduled_datetime(raw: str, base: datetime | None = None) -> datetime | None:
    text = cell_text(raw)
    if not text:
        return None
    base = base or datetime.now()
    normalized = (
        text.replace("：", ":")
        .replace("年", "-")
        .replace("月", "-")
        .replace("日", " ")
        .replace("/", "-")
        .strip()
    )
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            parsed = datetime.strptime(normalized, fmt)
            if fmt == "%Y-%m-%d":
                return parsed.replace(hour=0, minute=0, second=0)
            return parsed
        except ValueError:
            pass

    marker = ""
    for candidate in ("凌晨", "早上", "上午", "中午", "下午", "晚上"):
        if candidate in normalized:
            marker = candidate
            normalized = normalized.replace(candidate, "")
            break
    match = re.search(r"([0-2]?\d)\s*:\s*([0-5]\d)", normalized)
    if not match:
        match = re.search(r"([0-2]?\d)\s*点(?:\s*([0-5]\d)\s*分?)?", normalized)
    if not match:
        return None
    hour = int(match.group(1))
    minute = int(match.group(2) or 0)
    if marker in {"下午", "晚上"} and hour < 12:
        hour += 12
    if marker == "中午" and hour < 11:
        hour += 12
    if not (0 <= hour <= 23 and 0 <= minute <= 59):
        return None
    return base.replace(hour=hour, minute=minute, second=0, microsecond=0)


def compact_text(value: str) -> str:
    text = cell_text(value).lower().translate(OCR_CONFUSIONS)
    return "".join(re.findall(r"[0-9a-z\u4e00-\u9fff]+", text))


def compact_text_strict(value: str) -> str:
    text = cell_text(value).lower()
    return "".join(re.findall(r"[0-9a-z\u4e00-\u9fff]+", text))


def is_cjk_char(ch: str) -> bool:
    return "\u4e00" <= ch <= "\u9fff"


def is_name_char(ch: str) -> bool:
    return is_cjk_char(ch) or (ch.isascii() and ch.isalnum())


def is_short_cjk_name(value: str) -> bool:
    text = compact_text(value)
    return 1 <= len(text) <= 4 and all(is_cjk_char(ch) for ch in text)


def ocr_raw_lines(value: str) -> list[str]:
    return [line.strip() for line in re.split(r"[\r\n]+", cell_text(value)) if line.strip()]


def compact_ocr_lines(value: str) -> list[str]:
    return [compact_text(line) for line in ocr_raw_lines(value) if compact_text(line)]


def compact_ocr_lines_strict(value: str) -> list[str]:
    return [compact_text_strict(line) for line in ocr_raw_lines(value) if compact_text_strict(line)]


def strip_ocr_spaces(value: str) -> str:
    return re.sub(r"\s+", "", cell_text(value)).lower().translate(OCR_CONFUSIONS)


def strip_ocr_spaces_strict(value: str) -> str:
    return re.sub(r"\s+", "", cell_text(value)).lower()


def search_non_chat_result_marker(value: str) -> str:
    raw = cell_text(value).lower()
    compact_raw = compact_text_strict(raw)
    stripped_raw = strip_ocr_spaces_strict(raw)
    for marker in SEARCH_NON_CHAT_RESULT_MARKERS:
        marker_raw = marker.lower()
        if marker_raw in raw or marker_raw in stripped_raw:
            return marker
        marker_key = compact_text_strict(marker_raw)
        if marker_key and marker_key in compact_raw:
            return marker
    return ""


def short_cjk_name_matches(expected: str, ocr_text: str) -> bool:
    """Match short Chinese names only as an exact token, never as a prefix or folded repeat."""
    expected = compact_text(expected)
    for compact_line in compact_ocr_lines(ocr_text):
        if compact_line == expected:
            return True

    spaced_name = r"\s*".join(re.escape(ch) for ch in expected)
    token_pattern = re.compile(rf"(?<![\u4e00-\u9fff]){spaced_name}(?!\s*[\u4e00-\u9fffa-zA-Z0-9])")
    for raw_line in ocr_raw_lines(ocr_text):
        normalized_line = raw_line.lower().translate(OCR_CONFUSIONS)
        if token_pattern.search(normalized_line):
            return True
        folded = strip_ocr_spaces(raw_line)
        start = folded.find(expected)
        while start >= 0:
            before = folded[start - 1] if start > 0 else ""
            end = start + len(expected)
            after = folded[end] if end < len(folded) else ""
            before_ok = not before or not is_cjk_char(before)
            after_ok = not after or not is_name_char(after)
            if before_ok and after_ok:
                return True
            start = folded.find(expected, start + 1)
    return False


def ocr_has_short_name_near_miss(target: str, ocr_text: str) -> bool:
    expected = compact_text_strict(target)
    if not is_short_cjk_name(expected):
        return False
    actual = compact_text_strict(ocr_text)
    return bool(expected and expected in actual and not ocr_contains_target(target, ocr_text))


def ocr_near_miss_target(targets: Iterable[str], ocr_text: str) -> str:
    for target in targets:
        if ocr_has_short_name_near_miss(target, ocr_text):
            return cell_text(target)
    return ""


def ocr_search_result_short_tail_match(target: str, ocr_text: str) -> bool:
    expected = compact_text_strict(target)
    if len(expected) < 3 or len(expected) > 4 or not all(is_cjk_char(ch) for ch in expected):
        return False
    if wecom_global_block_marker(ocr_text):
        return False
    tail = expected[-2:]
    for line in compact_ocr_lines_strict(ocr_text):
        if tail in line:
            return True
    return False


def ocr_search_result_short_tail_target(targets: Iterable[str], ocr_text: str) -> str:
    for target in targets:
        if ocr_search_result_short_tail_match(target, ocr_text):
            return cell_text(target)
    return ""


def ocr_short_cjk_shared_chars_count(target: str, ocr_text: str) -> int:
    expected = compact_text_strict(target)
    if len(expected) < 3 or len(expected) > 4 or not all(is_cjk_char(ch) for ch in expected):
        return 0
    actual = compact_text_strict(ocr_text)
    if not actual or wecom_global_block_marker(ocr_text):
        return 0
    return sum(1 for ch in set(expected) if ch in actual)


def ocr_short_cjk_title_near_match(target: str, ocr_text: str) -> bool:
    return ocr_short_cjk_shared_chars_count(target, ocr_text) >= 2


def ocr_short_cjk_title_near_target(targets: Iterable[str], ocr_text: str) -> str:
    for target in targets:
        if ocr_short_cjk_title_near_match(target, ocr_text):
            return cell_text(target)
    return ""


def stable_open_failure_reason(error: str) -> bool:
    return "未找到本地精确目标" in error or "搜索结果为空" in error


def near_search_result_fallback_allowed(target: str, title_ocr_text: str) -> bool:
    expected = compact_text(target)
    if len(expected) < 3 or not all(is_cjk_char(ch) for ch in expected):
        return False
    actual = compact_text_strict(title_ocr_text)
    if not actual or wecom_global_block_marker(title_ocr_text):
        return False
    shared = sum(1 for ch in set(expected) if ch in actual)
    return shared >= 1


def opened_title_has_context_marker(ocr_text: str) -> str:
    actual = compact_text_strict(ocr_text)
    if not actual or wecom_global_block_marker(ocr_text):
        return ""
    markers = (*TITLE_ROLE_MARKERS, *TITLE_DEPARTMENT_MARKERS, *TITLE_NOTE_MARKERS, *TITLE_STATUS_MARKERS)
    for marker in markers:
        marker_key = compact_text_strict(marker)
        if marker_key and marker_key in actual:
            return cell_text(marker)
    return ""


def cjk_shared_chars_count(expected: str, actual_text: str) -> int:
    expected_key = compact_text_strict(expected)
    actual_key = compact_text_strict(actual_text)
    if not expected_key or not actual_key:
        return 0
    expected_chars = {ch for ch in expected_key if is_cjk_char(ch)}
    if not expected_chars:
        return 0
    return sum(1 for ch in expected_chars if ch in actual_key)


def ocr_near_matches_any_target(targets: Iterable[str], ocr_text: str, min_shared: int = 2) -> str:
    for target in targets:
        value = cell_text(target)
        if cjk_shared_chars_count(value, ocr_text) >= min_shared:
            return value
    return ""


def safe_filename(value: str, fallback: str = "target") -> str:
    text = re.sub(r'[\\/:*?"<>|\s]+', "_", cell_text(value)).strip("._")
    return text[:40] or fallback


def ocr_contains_target(target: str, ocr_text: str) -> bool:
    expected = compact_text(target)
    actual = compact_text(ocr_text)
    if not expected:
        return False
    if is_short_cjk_name(expected):
        return short_cjk_name_matches(expected, ocr_text)
    if expected in actual:
        return True
    latin_tokens = re.findall(r"[a-z0-9]+", cell_text(target).lower())
    if latin_tokens and any(len(token) >= 4 and token in actual for token in latin_tokens):
        return True
    return False


def ocr_contains_any_target(targets: Iterable[str], ocr_text: str) -> str:
    for target in targets:
        if ocr_contains_target(target, ocr_text):
            return cell_text(target)
    return ""


def short_cjk_title_name_matches(expected: str, ocr_text: str) -> bool:
    expected = compact_text_strict(expected)
    if not is_short_cjk_name(expected):
        return False
    metadata_prefixes = tuple(compact_text_strict(prefix) for prefix in TITLE_METADATA_PREFIXES)
    context_markers = tuple(compact_text_strict(marker) for marker in TITLE_CONTEXT_MARKERS)
    for compact_line in compact_ocr_lines_strict(ocr_text):
        start = compact_line.find(expected)
        while start >= 0:
            before = compact_line[start - 1] if start > 0 else ""
            prefix = compact_line[max(0, start - 12) : start]
            suffix = compact_line[start + len(expected) :]
            before_ok = not before or not is_name_char(before)
            if before_ok and suffix and any(suffix.startswith(prefix) for prefix in metadata_prefixes if prefix):
                return True
            if before_ok and any(marker and marker in suffix[:14] for marker in context_markers):
                return True
            if suffix and any(marker and marker in prefix for marker in context_markers):
                after = suffix[0]
                if not after or not is_name_char(after) or any(marker and suffix.startswith(marker) for marker in context_markers):
                    return True
            start = compact_line.find(expected, start + 1)
    return False


def latin_fuzzy_contains_target(target: str, ocr_text: str, threshold: float = 0.78) -> bool:
    expected = compact_text(target)
    if not expected or not re.search(r"[a-z0-9]", expected):
        return False
    actual = compact_text(ocr_text)
    if expected in actual:
        return True
    if len(expected) < 4:
        return False
    tokens = re.findall(r"[a-z0-9]{4,}", cell_text(ocr_text).lower())
    if actual:
        width = len(expected)
        for size in {max(4, width - 1), width, width + 1}:
            tokens.extend(actual[start : start + size] for start in range(0, max(0, len(actual) - size + 1)))
    for token in tokens:
        if difflib.SequenceMatcher(None, expected, token).ratio() >= threshold:
            return True
    return False


def cjk_long_title_anchor_match(expected: str, actual: str) -> bool:
    if len(expected) < 5 or not all(is_cjk_char(ch) for ch in expected):
        return False
    if not actual:
        return False
    head = expected[:2]
    tail = expected[-2:]
    head_index = actual.find(head)
    tail_index = actual.rfind(tail)
    if head_index < 0 or tail_index < 0 or head_index >= tail_index:
        return False
    matched_chars = sum(1 for ch in expected if ch in actual)
    return matched_chars >= max(4, len(expected) - 2)


def ocr_contains_title_candidate(target: str, ocr_text: str) -> bool:
    if ocr_contains_target(target, ocr_text):
        return True
    if short_cjk_title_name_matches(target, ocr_text):
        return True
    expected = compact_text(target)
    actual = compact_text(ocr_text)
    if not expected or not actual:
        return False
    if latin_fuzzy_contains_target(target, ocr_text):
        return True
    if is_short_cjk_name(expected):
        for raw_line in ocr_raw_lines(ocr_text):
            line_key = compact_text(raw_line)
            if len(expected) >= 3 and line_key.startswith(expected):
                after = line_key[len(expected)] if len(line_key) > len(expected) else ""
                if not after or not is_cjk_char(after):
                    return True
            if expected in line_key and any((marker_key := compact_text(marker)) and marker_key in line_key for marker in TITLE_CONTEXT_MARKERS):
                return True
        return False
    if len(expected) < 3 or not all(is_cjk_char(ch) for ch in expected):
        return False
    if len(expected) >= 3 and all(is_cjk_char(ch) for ch in expected) and expected in actual:
        return True
    return cjk_long_title_anchor_match(expected, actual)


def ocr_contains_any_title_candidate(targets: Iterable[str], ocr_text: str) -> str:
    for target in targets:
        if ocr_contains_title_candidate(target, ocr_text):
            return cell_text(target)
    return ""


def ocr_contains_relaxed_opened_title_candidate(target: str, ocr_text: str) -> bool:
    expected = compact_text_strict(target)
    if not is_short_cjk_name(expected):
        return False
    compact_lines = compact_ocr_lines_strict(ocr_text)
    if not compact_lines:
        return False
    context_markers = tuple(compact_text_strict(marker) for marker in TITLE_CONTEXT_MARKERS)
    metadata_prefixes = tuple(compact_text_strict(prefix) for prefix in TITLE_METADATA_PREFIXES)
    for line in compact_lines:
        if expected not in line:
            continue
        if sum(1 for _ in re.finditer(re.escape(expected), line)) >= 2:
            return True
        prefix = line[max(0, line.find(expected) - 12) : line.find(expected)]
        suffix = line[line.find(expected) + len(expected) :]
        if any(marker and marker in suffix[:14] for marker in context_markers):
            return True
        if any(marker and marker in prefix for marker in context_markers):
            after = suffix[0] if suffix else ""
            if not after or not is_name_char(after) or any(marker and suffix.startswith(marker) for marker in context_markers):
                return True
        if any(prefix and prefix in suffix[:8] for prefix in metadata_prefixes):
            return True
    return False


def ocr_contains_any_relaxed_opened_title_candidate(targets: Iterable[str], ocr_text: str) -> str:
    for target in targets:
        if ocr_contains_relaxed_opened_title_candidate(target, ocr_text):
            return cell_text(target)
    return ""


def ocr_contains_search_result_target(target: str, ocr_text: str) -> bool:
    if ocr_contains_target(target, ocr_text):
        return True
    expected = compact_text(target)
    actual = compact_text(ocr_text)
    if not expected or not actual:
        return False
    if latin_fuzzy_contains_target(target, ocr_text, threshold=0.80):
        return True
    return cjk_long_title_anchor_match(expected, actual)


def ocr_contains_any_search_result_target(targets: Iterable[str], ocr_text: str) -> str:
    for target in targets:
        if ocr_contains_search_result_target(target, ocr_text):
            return cell_text(target)
    return ""


def target_ocr_variants(target: str) -> list[str]:
    text = cell_text(target)
    if not text:
        return []
    substitutions = {
        "妤": ("好",),
        "芸": ("云", "靶"),
        "珅": ("呻", "砷", "坤"),
        "喆": ("吉", "品"),
        "宁": ("丁", "了"),
    }
    variants = {text}
    for index, char in enumerate(text):
        replacements = substitutions.get(char, ())
        if not replacements:
            continue
        current_variants = list(variants)
        for variant in current_variants:
            if index >= len(variant) or variant[index] != char:
                continue
            for replacement in replacements:
                variants.add(variant[:index] + replacement + variant[index + 1 :])
    return [variant for variant in variants if variant]


def target_name_parts(target: str) -> list[str]:
    text = cell_text(target)
    if not text:
        return []
    parts: list[str] = []

    def add(value: str) -> None:
        value = cell_text(value).strip()
        if value and value != text and value not in parts:
            parts.append(value)

    for match in re.findall(r"[（(]([^（）()]+)[）)]", text):
        add(match)
    without_parentheses = re.sub(r"[（(][^（）()]+[）)]", "", text).strip()
    add(without_parentheses)
    for piece in re.split(r"[-－—–_·/｜|]", without_parentheses or text):
        add(piece)
    return parts


def target_ocr_candidates(target: str, aliases: Iterable[str], include_variants: bool = True) -> list[str]:
    candidates: list[str] = []
    for value in [target, *aliases]:
        base_values = [cell_text(value), *target_name_parts(value)]
        values = []
        for base_value in base_values:
            values.extend(target_ocr_variants(base_value) if include_variants else [base_value])
        for candidate in values:
            if candidate and candidate not in candidates:
                candidates.append(candidate)
    return candidates


def marker_in_ocr(ocr_text: str, markers: Iterable[str]) -> str:
    raw = cell_text(ocr_text)
    compact_raw = compact_text(raw)
    for marker in markers:
        if marker in raw:
            return marker
        compact_marker = compact_text(marker)
        if compact_marker and compact_marker in compact_raw:
            return marker
    return ""


def send_error_marker(ocr_text: str) -> str:
    return marker_in_ocr(ocr_text, SEND_ERROR_MARKERS)


def wecom_global_block_marker(ocr_text: str) -> str:
    return marker_in_ocr(ocr_text, WECOM_GLOBAL_BLOCK_MARKERS)


def send_pending_marker(ocr_text: str) -> str:
    raw = cell_text(ocr_text)
    percent_matches = re.findall(r"(?<![0-9])([0-9]{1,3})\s*%", raw)
    for value in percent_matches:
        if int(value) < 100:
            return f"{value}%"
    return marker_in_ocr(raw, SEND_PENDING_MARKERS)


def is_runtime_skip_error(error: str) -> bool:
    text = cell_text(error)
    return any(
        marker in text
        for marker in (
            "已跳过发送",
            "跳过发送",
            "搜索结果未命中目标名",
            "搜索结果 OCR 为空",
            "搜索结果未找到本地精确目标",
        )
    )


def latest_outgoing_bubble_bbox(image: Any) -> tuple[int, int, int, int] | None:
    width, height = image.size
    pixels = image.load()
    blue_pixels: set[tuple[int, int]] = set()
    x_start = max(0, int(width * 0.34))
    y_stop = max(0, height - 115)
    for y in range(max(0, int(height * 0.08)), y_stop):
        for x in range(x_start, width):
            r, g, b = pixels[x, y]
            if 170 <= r <= 225 and 200 <= g <= 245 and 220 <= b <= 255 and b - r >= 25 and g - r >= 5:
                blue_pixels.add((x, y))

    seen: set[tuple[int, int]] = set()
    candidates: list[tuple[int, int, int, int, int]] = []
    for point in list(blue_pixels):
        if point in seen:
            continue
        stack = [point]
        seen.add(point)
        xs: list[int] = []
        ys: list[int] = []
        while stack:
            x, y = stack.pop()
            xs.append(x)
            ys.append(y)
            for nx in (x - 1, x, x + 1):
                for ny in (y - 1, y, y + 1):
                    neighbor = (nx, ny)
                    if neighbor in blue_pixels and neighbor not in seen:
                        seen.add(neighbor)
                        stack.append(neighbor)

        count = len(xs)
        left, right = min(xs), max(xs)
        top, bottom = min(ys), max(ys)
        component_width = right - left + 1
        component_height = bottom - top + 1
        if count >= 2500 and component_width >= 220 and component_height >= 35:
            candidates.append((count, left, top, right, bottom))

    if not candidates:
        return None
    _, left, top, right, bottom = max(candidates, key=lambda item: (item[4], item[0]))
    return left, top, right, bottom


def crop_latest_outgoing_message(image_path: Path, output_path: Path) -> Path:
    try:
        from PIL import Image
    except Exception:
        return image_path
    try:
        image = Image.open(image_path).convert("RGB")
    except Exception:
        return image_path
    bubble = latest_outgoing_bubble_bbox(image)
    if bubble is None:
        return image_path
    width, height = image.size
    left, top, right, bottom = bubble
    crop_box = (
        max(0, left - 75),
        max(0, top - 28),
        min(width, right + 12),
        min(height, bottom + 75),
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    image.crop(crop_box).save(output_path)
    return output_path


def screenshot_red_failure_indicator(image_path: Path) -> str:
    try:
        from PIL import Image
    except Exception:
        return ""
    try:
        image = Image.open(image_path).convert("RGB")
    except Exception:
        return ""

    width, height = image.size
    pixels = image.load()
    bubble = latest_outgoing_bubble_bbox(image)
    if bubble is None:
        return ""
    bubble_left, bubble_top, _bubble_right, bubble_bottom = bubble
    red_pixels: set[tuple[int, int]] = set()
    scan_left = max(0, bubble_left - 70)
    scan_right = max(scan_left, bubble_left - 3)
    scan_top = max(0, bubble_top - 18)
    scan_bottom = min(height, min(bubble_bottom + 18, bubble_top + 110))
    for y in range(scan_top, scan_bottom):
        for x in range(scan_left, scan_right):
            r, g, b = pixels[x, y]
            if r >= 190 and g <= 110 and b <= 110 and r - g >= 80 and r - b >= 80:
                red_pixels.add((x, y))

    seen: set[tuple[int, int]] = set()
    for point in list(red_pixels):
        if point in seen:
            continue
        stack = [point]
        seen.add(point)
        xs: list[int] = []
        ys: list[int] = []
        while stack:
            x, y = stack.pop()
            xs.append(x)
            ys.append(y)
            for nx in (x - 1, x, x + 1):
                for ny in (y - 1, y, y + 1):
                    neighbor = (nx, ny)
                    if neighbor in red_pixels and neighbor not in seen:
                        seen.add(neighbor)
                        stack.append(neighbor)

        count = len(xs)
        left, right = min(xs), max(xs)
        top, bottom = min(ys), max(ys)
        component_width = right - left + 1
        component_height = bottom - top + 1
        if count < 45 or not (8 <= component_width <= 32 and 8 <= component_height <= 32):
            continue

        sample_points = [
            (max(left - 3, 0), (top + bottom) // 2),
            (min(right + 3, width - 1), (top + bottom) // 2),
            ((left + right) // 2, max(top - 3, 0)),
            ((left + right) // 2, min(bottom + 3, height - 1)),
        ]
        blue_neighbors = 0
        for sx, sy in sample_points:
            sr, sg, sb = pixels[sx, sy]
            if sb >= 185 and sg >= 175 and sr <= 215:
                blue_neighbors += 1
        if blue_neighbors >= 2:
            continue
        return f"最新消息旁疑似出现红色失败感叹号：区域=({left},{top})-({right},{bottom})"
    return ""


def latest_outgoing_bubble_from_path(image_path: Path) -> tuple[int, int, int, int] | None:
    try:
        from PIL import Image
    except Exception:
        return None
    try:
        image = Image.open(image_path).convert("RGB")
    except Exception:
        return None
    return latest_outgoing_bubble_bbox(image)


def same_bubble_area(before: tuple[int, int, int, int] | None, after: tuple[int, int, int, int] | None) -> bool:
    if before is None or after is None:
        return False
    return all(abs(a - b) <= 8 for a, b in zip(before, after))


def latest_bubble_confirmed_new(
    before: tuple[int, int, int, int] | None,
    after: tuple[int, int, int, int] | None,
    ocr_text: str,
    send_action_time: str,
) -> bool:
    if after is None:
        return False
    if before is None:
        return True
    if not same_bubble_area(before, after):
        return True
    return ocr_contains_time_marker(ocr_text, send_action_time)


def detect_input_panel_bbox_from_image(image: Any) -> tuple[int, int, int, int] | None:
    width, height = image.size
    if width < 520 or height < 320:
        return None
    pixels = image.load()
    min_run_width = max(280, int(width * 0.32))
    y_start = max(0, height - 155)
    y_stop = max(y_start + 1, height - 24)
    best: tuple[int, int, int] | None = None
    for y in range(y_start, y_stop, 5):
        start = -1
        for x in range(0, width):
            red, green, blue = pixels[x, y]
            whiteish = red >= 248 and green >= 248 and blue >= 248
            if whiteish and start < 0:
                start = x
            if start >= 0 and (not whiteish or x == width - 1):
                end = x - 1 if not whiteish else x
                run_width = end - start + 1
                if run_width >= min_run_width and width * 0.25 <= start <= width * 0.45:
                    score = run_width - abs((height - 38) - y) - max(0, start - int(width * 0.32))
                    if best is None or score > best[0]:
                        best = (score, start, end)
                start = -1
    if best is None:
        return None
    _score, left, right = best
    return left, max(0, height - 160), right, max(0, height - 16)


def screenshot_has_input_panel(image_path: Path) -> bool:
    try:
        from PIL import Image
    except Exception:
        return True
    try:
        with Image.open(image_path) as image:
            return detect_input_panel_bbox_from_image(image.convert("RGB")) is not None
    except Exception:
        return True


def screenshot_has_visible_search_candidate(image_path: Path) -> bool:
    try:
        from PIL import Image
    except Exception:
        return True
    try:
        with Image.open(image_path) as image:
            rgb = image.convert("RGB")
            width, height = rgb.size
            pixels = rgb.getdata()
            visible_pixels = 0
            for red, green, blue in pixels:
                if min(red, green, blue) < 235 or max(red, green, blue) - min(red, green, blue) > 30:
                    visible_pixels += 1
            return visible_pixels >= max(40, int(width * height * 0.03))
    except Exception:
        return True


def image_difference_ratio(before_path: Path, after_path: Path) -> float:
    try:
        from PIL import Image
    except Exception:
        return 1.0
    try:
        with Image.open(before_path) as before_image, Image.open(after_path) as after_image:
            before = before_image.convert("RGB")
            after = after_image.convert("RGB")
            if before.size != after.size:
                after = after.resize(before.size)
            before_pixels = before.load()
            after_pixels = after.load()
            width, height = before.size
            if width <= 0 or height <= 0:
                return 1.0
            changed = 0
            for y in range(height):
                for x in range(width):
                    br, bg, bb = before_pixels[x, y]
                    ar, ag, ab = after_pixels[x, y]
                    if abs(br - ar) + abs(bg - ag) + abs(bb - ab) >= 45:
                        changed += 1
            return changed / float(width * height)
    except Exception:
        return 1.0


def ocr_contains_time_marker(ocr_text: str, time_text: str) -> bool:
    if not time_text:
        return False
    try:
        sent_at = datetime.strptime(time_text, "%Y-%m-%d %H:%M:%S")
    except ValueError:
        return False
    raw_head = cell_text(ocr_text)[:120]
    compact_digits = re.sub(r"\D+", "", raw_head)
    for minute in (sent_at, sent_at + timedelta(minutes=1)):
        time_pattern = re.compile(rf"(?<!\d){minute.hour:02d}\s*[:：]?\s*{minute.minute:02d}(?!\d)")
        loose_time_pattern = re.compile(rf"(?<!\d){minute.hour}\s*[:：]?\s*{minute.minute:02d}(?!\d)")
        if time_pattern.search(raw_head) or loose_time_pattern.search(raw_head):
            return True
        marker = f"{minute.hour:02d}{minute.minute:02d}"
        loose_marker = f"{minute.hour}{minute.minute:02d}"
        if marker in compact_digits or loose_marker in compact_digits:
            return True
    return False


def normalize_object_type(raw: str) -> str:
    text = cell_text(raw)
    if text in GROUP_TYPES:
        return "群聊"
    if text in PERSON_TYPES:
        return "个人"
    return text


def is_wecom_channel(raw: str) -> bool:
    text = cell_text(raw)
    return text in CHANNEL_ALIASES or text.lower() in {"wecom", "wxwork"}


def tesseract_command() -> str | None:
    found = shutil.which("tesseract")
    if found:
        return found
    for raw in ("/usr/local/bin/tesseract", "/opt/homebrew/bin/tesseract", "/opt/local/bin/tesseract"):
        path = Path(raw)
        if path.exists() and path.is_file():
            return str(path)
    return None


@dataclass
class SendRow:
    row: int
    channel: str
    object_type: str
    target: str
    target_aliases: list[str]
    scheduled_at: str
    message: str
    message_type: str
    image_paths: list[Path]
    document_paths: list[Path]
    parts: dict[str, bool]
    prior_status: str = ""
    should_send: bool = True
    reason: str = ""
    search_seconds: float = 0.0
    send_seconds: float = 0.0
    total_seconds: float = 0.0
    schedule_status: str = ""
    schedule_detail: str = ""
    precheck_status: str = ""
    precheck_time: str = ""
    precheck_detail: str = ""
    targetcheck_status: str = ""
    targetcheck_time: str = ""
    targetcheck_detail: str = ""
    target_screenshot: str = ""
    postcheck_status: str = ""
    postcheck_time: str = ""
    postcheck_detail: str = ""
    before_screenshot: str = ""
    after_screenshot: str = ""
    after_ocr_text: str = ""
    send_action_time: str = ""
    sent_steps: list[str] = field(default_factory=list)
    step_events: list[dict[str, Any]] = field(default_factory=list)
    evidence_files: list[dict[str, Any]] = field(default_factory=list)
    low_evidence_deleted_files: int = 0

    @property
    def cache_key(self) -> str:
        return f"{self.object_type}::{self.target}"

    @property
    def file_paths(self) -> list[Path]:
        files: list[Path] = []
        if self.parts["image"]:
            files.extend(self.image_paths)
        if self.parts["file"]:
            files.extend(self.document_paths)
        return files

    def as_preview(self) -> dict[str, Any]:
        return {
            "row": self.row,
            "channel": self.channel,
            "object_type": self.object_type,
            "target": self.target,
            "target_aliases": self.target_aliases,
            "scheduled_at": self.scheduled_at,
            "message": self.message,
            "message_type": self.message_type,
            "image_paths": [str(path) for path in self.image_paths],
            "document_paths": [str(path) for path in self.document_paths],
            "should_send": self.should_send,
            "reason": self.reason,
            "prior_status": self.prior_status,
            "precheck_detail": build_precheck_detail(self),
        }


@dataclass
class RunStats:
    batch_id: str = field(default_factory=lambda: uuid.uuid4().hex[:8])
    started_at: str = field(default_factory=now_text)
    ended_at: str = ""
    run_dir: Path | None = None
    log_path: Path | None = None
    live_log: bool = True
    total_seconds: float = 0.0
    dispatch_completed_at: str = ""
    dispatch_seconds: float = 0.0
    verification_completed_at: str = ""
    verification_seconds: float = 0.0
    success: int = 0
    late_success: int = 0
    failed: int = 0
    skipped: int = 0


def create_run_log_folder(folder: Path) -> tuple[Path, Path]:
    root = folder / LOG_ROOT_RELATIVE_PATH
    root.mkdir(parents=True, exist_ok=True)
    base_name = run_minute_slug()
    run_dir = root / base_name
    if run_dir.exists():
        index = 2
        while True:
            candidate = root / f"{base_name}_{index:02d}"
            if not candidate.exists():
                run_dir = candidate
                break
            index += 1
    run_dir.mkdir()
    return run_dir, run_dir / "run_log.txt"


def resolve_run_log_folder(folder: Path, requested_run_dir: str = "") -> tuple[Path, Path]:
    if not requested_run_dir:
        return create_run_log_folder(folder)
    run_dir = Path(requested_run_dir).expanduser()
    if not run_dir.is_absolute():
        run_dir = folder / run_dir
    run_dir.mkdir(parents=True, exist_ok=True)
    return run_dir, run_dir / "run_log.txt"


def init_run_logging(stats: RunStats, folder: Path, workbook: Path, execute: bool, requested_run_dir: str = "") -> None:
    stats.run_dir, stats.log_path = resolve_run_log_folder(folder, requested_run_dir)
    log_line(
        stats,
        (
            f"[{stats.started_at}] 程序启动 | 批次={stats.batch_id} | "
            f"模式={'真实发送' if execute else '预览'} | Excel={workbook.name}"
        ),
    )


def runtime_dependency_status() -> str:
    modules = ["openpyxl", "PIL", "pyautogui", "Quartz", "Vision", "AppKit"]
    statuses = []
    for module in modules:
        statuses.append(f"{module}={'OK' if importlib.util.find_spec(module) else 'MISSING'}")
    return (
        f"Python={sys.executable} | version={platform.python_version()} | "
        f"platform={platform.platform()} | deps={', '.join(statuses)}"
    )


def expected_steps(item: SendRow) -> list[str]:
    if item.parts["text"] and item.file_paths:
        return ["文字+附件"]
    steps: list[str] = []
    if item.parts["text"]:
        steps.append("文字")
    if item.parts["image"]:
        steps.append(f"图片({len(item.image_paths)})")
    if item.parts["file"]:
        steps.append(f"文档({len(item.document_paths)})")
    return steps


def append_test_disclaimer(message: str, disclaimer: str) -> str:
    text = cell_text(message)
    note = cell_text(disclaimer)
    if not note:
        return text
    if note in text:
        return text
    if not text:
        return note
    return f"{text}\n\n{note}"


def apply_test_disclaimer(rows: list[SendRow], disclaimer: str) -> int:
    changed = 0
    for item in rows:
        if not item.parts.get("text"):
            continue
        updated = append_test_disclaimer(item.message, disclaimer)
        if updated != item.message:
            item.message = updated
            changed += 1
    return changed


def message_ocr_fragments(message: str) -> list[str]:
    fragments: list[str] = []
    for url in re.findall(r"https?://\S+", cell_text(message).lower()):
        path_match = re.search(r"/s/([a-z0-9_-]{8,})", url, flags=re.IGNORECASE)
        if path_match:
            slug = compact_text(path_match.group(1))
            if len(slug) >= 8:
                fragments.append(slug[:18])
                fragments.append(slug[-18:])
    for phrase in re.findall(r"[\u4e00-\u9fff]{4,}", cell_text(message)):
        compact = compact_text(phrase)
        if len(compact) >= 6:
            fragments.append(compact[:12])
    for token in re.findall(r"[a-z0-9]{6,}", cell_text(message).lower()):
        fragments.append(token[:16])
    compact_message = compact_text(message)
    if len(compact_message) >= 8:
        fragments.append(compact_message[:12])

    unique: list[str] = []
    for fragment in fragments:
        if fragment and fragment not in unique:
            unique.append(fragment)
    return unique[:12]


def message_short_fragments(message: str) -> list[str]:
    fragments: list[str] = []
    for phrase in re.findall(r"[\u4e00-\u9fff]{2,}", cell_text(message)):
        compact = compact_text(phrase)
        if 2 <= len(compact) <= 5:
            fragments.append(compact)
        elif len(compact) > 5:
            fragments.extend(compact[index : index + 4] for index in range(0, len(compact) - 3, 3))
        if len(compact) >= 4:
            fragments.extend(compact[index : index + 2] for index in range(0, len(compact) - 1))
    unique: list[str] = []
    for fragment in fragments:
        if fragment and fragment not in unique:
            unique.append(fragment)
    return unique[:80]


def message_visible_fragments(message: str) -> list[str]:
    fragments: list[str] = []
    for phrase in re.findall(r"[\u4e00-\u9fff]{4,}", cell_text(message)):
        compact = compact_text(phrase)
        if len(compact) >= 8:
            fragments.extend([compact[:8], compact[len(compact) // 2 : len(compact) // 2 + 8], compact[-8:]])
        if len(compact) >= 4:
            fragments.extend(compact[index : index + 4] for index in range(0, len(compact) - 3, 2))
    for token in re.findall(r"[a-z0-9]{6,}", cell_text(message).lower()):
        compact = compact_text(token)
        if len(compact) >= 8:
            fragments.extend([compact[:8], compact[-8:]])
    unique: list[str] = []
    for fragment in fragments:
        if fragment and fragment not in unique:
            unique.append(fragment)
    return unique[:220]


def ocr_matches_message(message: str, ocr_text: str) -> bool:
    actual = compact_text(ocr_text)
    if not actual:
        return False
    if any(fragment in actual for fragment in message_ocr_fragments(message)):
        return True
    short_hits = [fragment for fragment in message_short_fragments(message) if fragment in actual]
    return len(short_hits) >= 2


def ocr_matches_input_message(message: str, ocr_text: str) -> bool:
    if ocr_matches_message(message, ocr_text):
        return True
    actual = compact_text(ocr_text)
    if not actual:
        return False
    visible_hits = [fragment for fragment in message_visible_fragments(message) if fragment in actual]
    long_hits = [fragment for fragment in visible_hits if len(fragment) >= 8]
    medium_hits = [fragment for fragment in visible_hits if len(fragment) >= 4]
    return bool(long_hits) or len(medium_hits) >= 2


def input_message_presence(message: str, ocr_text: str) -> str:
    if ocr_matches_input_message(message, ocr_text):
        return "matched"
    actual = compact_text(ocr_text)
    if len(actual) < 2:
        return "empty"
    hint_fragments = [
        fragment
        for fragment in message_visible_fragments(message) + message_short_fragments(message)
        if len(fragment) >= 3 and fragment in actual
    ]
    return "partial" if hint_fragments else "unrelated"


def input_image_has_visible_text(image_path: Path) -> bool:
    try:
        from PIL import Image
    except Exception:
        return False
    try:
        image = Image.open(image_path).convert("RGB")
    except Exception:
        return False
    width, height = image.size
    if width < 20 or height < 20:
        return False
    left = max(0, int(width * 0.02))
    top = max(0, int(height * 0.04))
    right = min(width, int(width * 0.98))
    bottom = min(height, int(height * 0.88))
    dark_pixels = 0
    colored_pixels = 0
    pixel_bytes = image.crop((left, top, right, bottom)).tobytes()
    for index in range(0, len(pixel_bytes), 3):
        r, g, b = pixel_bytes[index], pixel_bytes[index + 1], pixel_bytes[index + 2]
        if r < 150 and g < 150 and b < 150:
            dark_pixels += 1
        elif max(r, g, b) - min(r, g, b) > 50 and min(r, g, b) < 210:
            colored_pixels += 1
    return dark_pixels + colored_pixels >= 80


def input_image_has_left_aligned_content(image_path: Path) -> bool:
    try:
        from PIL import Image
    except Exception:
        return False
    try:
        image = Image.open(image_path).convert("RGB")
    except Exception:
        return False
    width, height = image.size
    if width < 300 or height < 80:
        return False
    scores: list[int] = []
    for start, end in ((0.0, 0.25), (0.25, 0.5), (0.5, 0.75), (0.75, 1.0)):
        crop = image.crop((int(width * start), 0, int(width * end), height))
        pixel_bytes = crop.tobytes()
        score = 0
        for index in range(0, len(pixel_bytes), 3):
            r, g, b = pixel_bytes[index], pixel_bytes[index + 1], pixel_bytes[index + 2]
            if r < 150 and g < 150 and b < 150:
                score += 1
            elif max(r, g, b) - min(r, g, b) > 50 and min(r, g, b) < 210:
                score += 1
        scores.append(score)
    left_score = scores[0]
    middle_score = scores[1] + scores[2]
    total_score = sum(scores)
    if left_score < 1000 or total_score < 4500:
        return False
    return left_score / max(1, middle_score) >= 0.35


def build_precheck_detail(item: SendRow) -> str:
    image_names = ", ".join(path.name for path in item.image_paths) or "无"
    document_names = ", ".join(path.name for path in item.document_paths) or "无"
    text_state = "有文本" if item.message else "无文本"
    return (
        f"行{item.row}；渠道={item.channel}；对象类型={item.object_type}；发送对象={item.target}；"
        f"核对关键词={','.join(item.target_aliases) or '无'}；"
        f"计划发送时间={item.scheduled_at or '无'}；"
        f"消息类型={item.message_type}；{text_state}；图片={image_names}；文档={document_names}；"
        f"预期步骤={' -> '.join(expected_steps(item)) or '无'}"
    )


def add_step_event(item: SendRow, step: str, status: str, detail: str = "", screenshot: str = "") -> None:
    item.step_events.append(
        {
            "time": now_text(),
            "step": step,
            "status": status,
            "detail": brief_log_text(detail, 240) if detail else "",
            "screenshot": screenshot,
        }
    )


def verify_before_send(item: SendRow) -> None:
    failures: list[str] = []
    if item.reason:
        failures.append(item.reason)
    if not is_wecom_channel(item.channel):
        failures.append(f"渠道核对失败：{item.channel or '空'}")
    if item.object_type not in {"个人", "群聊"}:
        failures.append(f"对象类型核对失败：{item.object_type or '空'}")
    if not item.target:
        failures.append("发送对象核对失败：空")
    if item.parts["text"] and not item.message:
        failures.append("文本核对失败：空")
    for path in item.image_paths + item.document_paths:
        if not path.exists():
            failures.append(f"附件核对失败：{path}")
    if failures:
        item.precheck_status = "未通过"
        item.precheck_time = now_text()
        item.precheck_detail = "；".join(failures)
        add_step_event(item, "发送前核对", "未通过", item.precheck_detail)
        raise RuntimeError(item.precheck_detail)
    item.precheck_status = "已通过"
    item.precheck_time = now_text()
    item.precheck_detail = build_precheck_detail(item)
    add_step_event(item, "发送前核对", "已通过", item.precheck_detail)


def scheduled_datetime_for_item(item: SendRow, base: datetime | None = None) -> datetime | None:
    if not item.scheduled_at:
        return None
    planned = parse_scheduled_datetime(item.scheduled_at, base)
    if planned is None:
        raise RuntimeError(f"计划发送时间无法识别：{item.scheduled_at}")
    return planned


def wait_for_schedule_prepare(
    planned: datetime | None,
    max_wait_minutes: float,
    prepare_lead_seconds: float,
    now_value: datetime | None = None,
    sleep_func: Any = time.sleep,
) -> None:
    if planned is None:
        return
    now = now_value or datetime.now()
    seconds_until_plan = (planned - now).total_seconds()
    if seconds_until_plan > max_wait_minutes * 60:
        raise RuntimeError(
            f"计划发送时间未到且超过最大等待：计划={planned.strftime('%Y-%m-%d %H:%M:%S')}；"
            f"当前={now.strftime('%Y-%m-%d %H:%M:%S')}"
        )
    prepare_at = planned - timedelta(seconds=max(0.0, prepare_lead_seconds))
    seconds_until_prepare = (prepare_at - now).total_seconds()
    if seconds_until_prepare > 0:
        sleep_func(seconds_until_prepare)


def schedule_prepare_lead_seconds_for_rows(rows: list[SendRow], *, batch_fast_dispatch: bool) -> float:
    scheduled_rows = [row for row in rows if row.should_send and row.scheduled_at]
    if not scheduled_rows:
        return 0.0
    target_count = len({(row.channel, row.object_type, row.target) for row in scheduled_rows})
    if target_count <= 1:
        return DEFAULT_SINGLE_SCHEDULE_PREPARE_LEAD_SECONDS
    if not batch_fast_dispatch:
        return max(DEFAULT_BATCH_SCHEDULE_PREPARE_MIN_SECONDS, DEFAULT_SINGLE_SCHEDULE_PREPARE_LEAD_SECONDS)
    estimated = DEFAULT_SINGLE_SCHEDULE_PREPARE_LEAD_SECONDS + target_count * DEFAULT_BATCH_SCHEDULE_PREPARE_PER_TARGET_SECONDS
    return min(
        DEFAULT_BATCH_SCHEDULE_PREPARE_MAX_SECONDS,
        max(DEFAULT_BATCH_SCHEDULE_PREPARE_MIN_SECONDS, estimated),
    )


def scheduled_groups_for_rows(rows: list[SendRow]) -> dict[datetime, list[SendRow]]:
    groups: dict[datetime, list[SendRow]] = {}
    for row in rows:
        if not row.should_send or not row.scheduled_at:
            continue
        planned = scheduled_datetime_for_item(row)
        if planned is None:
            continue
        groups.setdefault(planned, []).append(row)
    return groups


def should_auto_batch_fast_dispatch(rows: list[SendRow], threshold: int = DEFAULT_BATCH_FAST_DISPATCH_THRESHOLD) -> bool:
    for group_rows in scheduled_groups_for_rows(rows).values():
        target_count = len({(row.channel, row.object_type, row.target) for row in group_rows})
        if target_count >= threshold:
            return True
    return False


def batch_dispatch_estimate(rows: list[SendRow], per_target_seconds: float) -> tuple[int, float]:
    target_count = len({(row.channel, row.object_type, row.target) for row in rows if row.should_send})
    return target_count, max(0.0, target_count * max(0.0, per_target_seconds))


def common_planned_send_time(rows: list[SendRow]) -> datetime | None:
    planned_values = set(scheduled_groups_for_rows(rows).keys())
    if len(planned_values) == 1:
        return next(iter(planned_values))
    return None


def wait_until_scheduled_time(
    planned: datetime | None,
    now_value: datetime | None = None,
    sleep_func: Any = time.sleep,
) -> None:
    if planned is None:
        return
    now = now_value or datetime.now()
    seconds_until_plan = (planned - now).total_seconds()
    if seconds_until_plan > 0:
        sleep_func(seconds_until_plan)


def wait_until_before_scheduled_time(
    planned: datetime | None,
    lead_seconds: float,
    now_value: datetime | None = None,
    sleep_func: Any = time.sleep,
) -> None:
    if planned is None:
        return
    target_time = planned - timedelta(seconds=max(0.0, lead_seconds))
    wait_until_scheduled_time(target_time, now_value, sleep_func)


def classify_scheduled_send_time(
    planned: datetime | None,
    on_time_tolerance_seconds: float,
    now_value: datetime | None = None,
) -> tuple[str, str]:
    if planned is None:
        return "", ""
    now = now_value or datetime.now()
    delay_seconds = max(0.0, (now - planned).total_seconds())
    planned_text = planned.strftime("%Y-%m-%d %H:%M:%S")
    actual_text = now.strftime("%Y-%m-%d %H:%M:%S")
    if delay_seconds <= max(0.0, on_time_tolerance_seconds):
        return "准时发送", f"计划={planned_text}；实际开始发送={actual_text}；延迟={delay_seconds:.1f}s"
    return "超时已发送", f"计划={planned_text}；实际开始发送={actual_text}；延迟={delay_seconds:.1f}s"


def classify_scheduled_send_action_time(
    planned: datetime | None,
    on_time_tolerance_seconds: float,
    send_action_time: str,
) -> tuple[str, str]:
    if planned is None:
        return "", ""
    try:
        actual = datetime.strptime(send_action_time, "%Y-%m-%d %H:%M:%S")
    except ValueError:
        actual = datetime.now()
    status, detail = classify_scheduled_send_time(planned, on_time_tolerance_seconds, actual)
    return status, detail.replace("实际开始发送=", "实际点击发送=")


def wait_for_scheduled_time(
    planned: datetime | None,
    on_time_tolerance_seconds: float,
    now_value: datetime | None = None,
    sleep_func: Any = time.sleep,
) -> tuple[str, str]:
    if planned is None:
        return "", ""
    wait_until_scheduled_time(planned, now_value, sleep_func)
    actual = planned if now_value is not None and planned > now_value else now_value
    return classify_scheduled_send_time(planned, on_time_tolerance_seconds, actual)


def verify_after_send(item: SendRow) -> None:
    expected = expected_steps(item)
    if item.targetcheck_status not in {"已通过", "已跳过"}:
        item.postcheck_status = "未通过"
        item.postcheck_time = now_text()
        item.postcheck_detail = f"目标会话未确认：{item.targetcheck_status or '未执行'}；{item.targetcheck_detail}"
        add_step_event(item, "发送后汇总核对", "未通过", item.postcheck_detail, item.after_screenshot)
        raise RuntimeError(item.postcheck_detail)
    if expected != item.sent_steps:
        item.postcheck_status = "未通过"
        item.postcheck_time = now_text()
        item.postcheck_detail = f"预期步骤={expected}；实际步骤={item.sent_steps}"
        add_step_event(item, "发送后汇总核对", "未通过", item.postcheck_detail, item.after_screenshot)
        raise RuntimeError(item.postcheck_detail)
    error = send_error_marker(item.after_ocr_text)
    pending = send_pending_marker(item.after_ocr_text)
    if error or pending:
        item.postcheck_status = "未通过"
        item.postcheck_time = now_text()
        marker = error or pending
        item.postcheck_detail = f"发送后界面仍有异常/待完成标记：{marker}；截图={item.after_screenshot}"
        add_step_event(item, "发送后汇总核对", "未通过", item.postcheck_detail, item.after_screenshot)
        raise RuntimeError(item.postcheck_detail)
    item.postcheck_status = "已通过"
    item.postcheck_time = now_text()
    detail = (
        f"发送后核对通过；目标={item.target}；"
        f"发送步骤：{' -> '.join(item.sent_steps)}"
    )
    if item.after_screenshot:
        detail += f"；发送后截图={item.after_screenshot}"
    if item.schedule_status:
        detail += f"；计划发送核对={item.schedule_status}；{item.schedule_detail}"
    item.postcheck_detail = detail
    add_step_event(item, "发送后汇总核对", "已通过", item.postcheck_detail, item.after_screenshot)


def find_lesson_header_row(ws: Any) -> tuple[int, dict[str, int]]:
    for row in range(1, min(ws.max_row, 20) + 1):
        headers = worksheet_headers(ws, row)
        if "课程链接" in headers and ("发送课程提醒文案" in headers or "发送课后总结" in headers):
            return row, headers
    raise ValueError("课程表未找到表头：需要包含 课程链接 以及 发送课程提醒文案/发送课后总结")


def normalize_lesson_key(value: str) -> str:
    text = compact_text(value)
    match = re.search(r"第?([0-9]+)课", text)
    if match:
        return f"第{int(match.group(1))}课"
    if text.isdigit():
        return f"第{int(text)}课"
    return cell_text(value)


def normalize_lesson_content(value: str) -> tuple[str, str]:
    text = compact_text(value)
    if text in {"课后总结", "发送课后总结", "课后文案", "总结", "summary", "aftersummary"}:
        return "发送课后总结", "课后总结"
    return "发送课程提醒文案", "课程提醒"


def lesson_payload(lesson_workbook: Path, lesson: str, lesson_content: str = "课程提醒") -> tuple[str, str, str, str, str]:
    wb = load_workbook(lesson_workbook, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_row, headers = find_lesson_header_row(ws)
    lesson_key = normalize_lesson_key(lesson)
    message_header, content_label = normalize_lesson_content(lesson_content)
    if message_header not in headers:
        raise ValueError(f"课程表未找到表头：{message_header}")
    message_col = headers[message_header]
    link_col = headers["课程链接"]
    schedule_col = headers.get("发送时间")
    lesson_col = 1
    for row in range(header_row + 1, ws.max_row + 1):
        current = normalize_lesson_key(cell_text(ws.cell(row, lesson_col).value))
        if current != lesson_key:
            continue
        message = cell_text(ws.cell(row, message_col).value)
        link = cell_text(ws.cell(row, link_col).value)
        scheduled_at = cell_text(ws.cell(row, schedule_col).value) if schedule_col else ""
        if not message:
            raise ValueError(f"{lesson_key} 的{message_header}为空")
        if not link:
            raise ValueError(f"{lesson_key} 的课程链接为空")
        if scheduled_at and parse_scheduled_datetime(scheduled_at) is None:
            raise ValueError(f"{lesson_key} 的发送时间无法识别：{scheduled_at}")
        text = f"{message}\n\n课程链接：{link}"
        return lesson_key, text, link, scheduled_at, content_label
    raise ValueError(f"课程表未找到 {lesson_key}")


def normalized_schedule_text(raw: str) -> str:
    planned = parse_scheduled_datetime(raw)
    if planned:
        return planned.strftime("%Y-%m-%d %H:%M:%S")
    return cell_text(raw)


def target_rows_from_workbook(target_workbook: Path) -> list[dict[str, str]]:
    wb = load_workbook(target_workbook, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = worksheet_headers(ws, 1)
    if first_col(headers, CHANNEL_HEADERS) is None:
        raise ValueError("目标表缺少渠道列")
    if first_col(headers, TARGET_HEADERS + CONTACT_HEADERS + GROUP_HEADERS) is None:
        raise ValueError("目标表缺少发送对象/联系人/群聊名称列")

    targets: list[dict[str, str]] = []
    for row in range(2, ws.max_row + 1):
        channel = col_value(ws, row, headers, CHANNEL_HEADERS)
        object_type = normalize_object_type(col_value(ws, row, headers, OBJECT_TYPE_HEADERS))
        contact = col_value(ws, row, headers, CONTACT_HEADERS)
        group_name = col_value(ws, row, headers, GROUP_HEADERS)
        fallback_target = col_value(ws, row, headers, TARGET_HEADERS)
        if object_type == "群聊":
            target = group_name or fallback_target or contact
        else:
            target = contact or fallback_target or group_name
        if not any((channel, object_type, target)):
            continue
        targets.append(
            {
                "source_row": str(row),
                "channel": channel or "企业微信",
                "object_type": object_type or "个人",
                "target": target,
                "aliases": col_value(ws, row, headers, TARGET_ALIAS_HEADERS) or target,
            }
        )
    return targets


def build_lesson_target_workbook(
    lesson_workbook: Path,
    target_workbook: Path,
    lesson: str,
    output_path: Path,
    dedupe_targets: bool = True,
    scheduled_at_override: str = "",
    lesson_content: str = "课程提醒",
) -> dict[str, Any]:
    lesson_key, text, _link, scheduled_at, content_label = lesson_payload(lesson_workbook, lesson, lesson_content)
    if scheduled_at_override:
        parsed_override = parse_scheduled_datetime(scheduled_at_override)
        if parsed_override is None:
            raise ValueError(f"计划发送时间无法识别：{scheduled_at_override}")
        scheduled_at = scheduled_at_override
    planned_text = normalized_schedule_text(scheduled_at)
    targets = target_rows_from_workbook(target_workbook)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    headers = ["渠道", "对象类型", "发送对象", "目标别名", "计划发送时间", "文本内容", "消息类型"]
    ws.append(headers)
    seen: set[tuple[str, str, str, str]] = set()
    duplicates: list[dict[str, str]] = []
    written = 0
    for target in targets:
        key = (
            compact_text(target["channel"]),
            target["object_type"],
            compact_text(target["target"]),
            compact_text(text)[:80],
        )
        if dedupe_targets and key in seen:
            duplicates.append({"source_row": target["source_row"], "target": target["target"]})
            continue
        seen.add(key)
        ws.append(
            [
                target["channel"],
                target["object_type"],
                target["target"],
                target["aliases"],
                planned_text,
                text,
                "文字",
            ]
        )
        written += 1
    widths = [16, 12, 18, 20, 24, 90, 12]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + index)].width = width
    wb.save(output_path)
    return {
        "lesson": lesson_key,
        "content": content_label,
        "scheduled_at": planned_text,
        "target_rows": len(targets),
        "send_rows": written,
        "duplicates": duplicates,
        "output_path": str(output_path),
    }


def build_lesson_single_target_workbook(
    lesson_workbook: Path,
    lesson: str,
    output_path: Path,
    target: str,
    object_type: str = "个人",
    channel: str = "企业微信",
    aliases: str = "",
    scheduled_at_override: str = "",
    lesson_content: str = "课程提醒",
) -> dict[str, Any]:
    lesson_key, text, _link, scheduled_at, content_label = lesson_payload(lesson_workbook, lesson, lesson_content)
    if scheduled_at_override:
        parsed_override = parse_scheduled_datetime(scheduled_at_override)
        if parsed_override is None:
            raise ValueError(f"计划发送时间无法识别：{scheduled_at_override}")
        scheduled_at = scheduled_at_override
    planned_text = normalized_schedule_text(scheduled_at)
    target_text = cell_text(target)
    if not target_text:
        raise ValueError("单个发送对象不能为空")
    object_type_text = normalize_object_type(object_type)
    if object_type_text not in {"个人", "群聊"}:
        raise ValueError(f"对象类型无法识别：{object_type}")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    headers = ["渠道", "对象类型", "发送对象", "目标别名", "计划发送时间", "文本内容", "消息类型"]
    ws.append(headers)
    ws.append(
        [
            cell_text(channel) or "企业微信",
            object_type_text,
            target_text,
            cell_text(aliases) or target_text,
            planned_text,
            text,
            "文字",
        ]
    )
    widths = [16, 12, 18, 20, 24, 90, 12]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + index)].width = width
    wb.save(output_path)
    return {
        "lesson": lesson_key,
        "content": content_label,
        "scheduled_at": planned_text,
        "target_rows": 1,
        "send_rows": 1,
        "duplicates": [],
        "output_path": str(output_path),
    }


class WorkbookStore:
    def __init__(self, workbook_path: Path, folder: Path, ensure_status: bool = False) -> None:
        self.path = workbook_path
        self.folder = folder
        self.saved_path = workbook_path
        self.writeback_mode = "original"
        self.wb = load_workbook(workbook_path)
        self.ws = self.wb[self.wb.sheetnames[0]]
        self.headers = {
            str(self.ws.cell(1, col).value).strip(): col
            for col in range(1, self.ws.max_column + 1)
            if self.ws.cell(1, col).value is not None and str(self.ws.cell(1, col).value).strip()
        }
        if ensure_status:
            ensure_columns(self.ws, self.headers, STATUS_COLUMNS)
            self.prepare_writeback()

    def recovery_path(self) -> Path:
        return self.path.with_name(f"{self.path.stem}_发送记录_{timestamp_slug()}{self.path.suffix}")

    def prepare_writeback(self) -> None:
        try:
            with self.path.open("a+b"):
                pass
            self.saved_path = self.path
            self.writeback_mode = "original"
        except OSError:
            self.saved_path = self.recovery_path()
            self.writeback_mode = "recovery_copy"

    def build_rows(
        self,
        allow_sent: bool = False,
        only_rows: set[int] | None = None,
    ) -> list[SendRow]:
        missing = []
        if first_col(self.headers, CHANNEL_HEADERS) is None:
            missing.append("渠道")
        if first_col(self.headers, TARGET_HEADERS + CONTACT_HEADERS + GROUP_HEADERS) is None:
            missing.append("发送对象/联系人/群聊名称")
        if missing:
            raise ValueError(f"无法识别必要字段：{', '.join(missing)}")

        rows: list[SendRow] = []
        for row in range(2, self.ws.max_row + 1):
            if only_rows and row not in only_rows:
                continue

            channel = col_value(self.ws, row, self.headers, CHANNEL_HEADERS)
            object_type_raw = col_value(self.ws, row, self.headers, OBJECT_TYPE_HEADERS)
            object_type = normalize_object_type(object_type_raw)
            contact = col_value(self.ws, row, self.headers, CONTACT_HEADERS)
            group_name = col_value(self.ws, row, self.headers, GROUP_HEADERS)
            fallback_target = col_value(self.ws, row, self.headers, TARGET_HEADERS)
            if object_type == "群聊":
                target = group_name or fallback_target or contact
            else:
                target = contact or fallback_target or group_name
            target_aliases = split_paths(col_value(self.ws, row, self.headers, TARGET_ALIAS_HEADERS))
            scheduled_at = col_value(self.ws, row, self.headers, SCHEDULE_HEADERS)
            message = col_value(self.ws, row, self.headers, TEXT_HEADERS)
            message_type = col_value(self.ws, row, self.headers, MESSAGE_TYPE_HEADERS)
            image_paths = resolve_paths(self.folder, col_value(self.ws, row, self.headers, IMAGE_HEADERS))
            document_paths = resolve_paths(self.folder, col_value(self.ws, row, self.headers, DOCUMENT_HEADERS))
            send_flag = col_value(self.ws, row, self.headers, SEND_FLAG_HEADERS)
            status = col_value(self.ws, row, self.headers, STATUS_HEADERS)

            if not any((channel, object_type_raw, contact, group_name, fallback_target, scheduled_at, message, message_type, image_paths, document_paths, status)):
                continue

            parts = parse_parts(message_type, message, [str(p) for p in image_paths], [str(p) for p in document_paths])
            reasons: list[str] = []
            if not is_wecom_channel(channel):
                reasons.append(f"渠道不是企业微信：{channel or '空'}")
            if object_type not in {"个人", "群聊"}:
                reasons.append(f"对象类型无法识别：{object_type_raw or '空'}")
            if not target:
                reasons.append("发送对象为空")
            if status in SENT_STATUSES and not allow_sent:
                reasons.append("已发送，避免重复发送")
            if send_flag and send_flag.lower() in NO_SEND_FLAGS:
                reasons.append(f"发送标记为 {send_flag}")
            if not any(parts.values()):
                reasons.append(f"无法识别消息类型：{message_type or '空'}")
            if parts["text"] and not message:
                reasons.append("消息类型包含文字但文本内容为空")
            if parts["image"]:
                if not image_paths:
                    reasons.append("消息类型包含图片但图片列为空")
                else:
                    missing_images = [str(path) for path in image_paths if not path.exists()]
                    if missing_images:
                        reasons.append(f"图片不存在：{'; '.join(missing_images)}")
            if parts["file"]:
                if not document_paths:
                    reasons.append("消息类型包含文件/文档但文档列为空")
                else:
                    missing_docs = [str(path) for path in document_paths if not path.exists()]
                    if missing_docs:
                        reasons.append(f"文档不存在：{'; '.join(missing_docs)}")

            rows.append(
                SendRow(
                    row=row,
                    channel=channel,
                    object_type=object_type,
                    target=target,
                    target_aliases=target_aliases,
                    scheduled_at=scheduled_at,
                    message=message,
                    message_type=message_type or self.infer_message_type(parts),
                    image_paths=image_paths,
                    document_paths=document_paths,
                    parts=parts,
                    prior_status=status,
                    should_send=not reasons,
                    reason="；".join(reasons),
                )
            )
        return rows

    @staticmethod
    def infer_message_type(parts: dict[str, bool]) -> str:
        labels = []
        if parts["text"]:
            labels.append("文字")
        if parts["image"]:
            labels.append("图片")
        if parts["file"]:
            labels.append("文档")
        return "、".join(labels) or "未知"

    def mark_row(self, item: SendRow, status: str, error: str, stats: RunStats, mode: str) -> None:
        ensure_columns(self.ws, self.headers, STATUS_COLUMNS)
        ended_at = now_text()
        values = {
            "发送状态": status,
            "错误原因": error,
            "发送时间": ended_at,
            "发送前核对": item.precheck_status,
            "发送前核对时间": item.precheck_time,
            "发送前核对详情": item.precheck_detail,
            "目标会话核对": item.targetcheck_status,
            "目标会话核对时间": item.targetcheck_time,
            "目标会话核对详情": item.targetcheck_detail,
            "目标会话截图": item.target_screenshot,
            "发送后核对": item.postcheck_status,
            "发送后核对时间": item.postcheck_time,
            "发送后核对详情": item.postcheck_detail,
            "发送后截图": item.after_screenshot,
            "发送后OCR": item.after_ocr_text[:500],
            "执行批次": stats.batch_id,
            "程序开始时间": stats.started_at,
            "程序结束时间": ended_at,
            "程序耗时秒": round(item.total_seconds, 3),
            "批次总耗时秒": "",
            "搜索耗时秒": round(item.search_seconds, 3),
            "发送耗时秒": round(item.send_seconds, 3),
            "执行模式": mode,
        }
        for header, value in values.items():
            if header in STATUS_COLUMNS and header in self.headers:
                self.ws.cell(item.row, self.headers[header]).value = value

    def mark_batch_total(self, rows: list[SendRow], stats: RunStats) -> None:
        ensure_columns(self.ws, self.headers, STATUS_COLUMNS)
        values = {
            "批次总耗时秒": round(stats.total_seconds, 3),
            "发送完成时间": stats.dispatch_completed_at,
            "发送完成耗时秒": round(stats.dispatch_seconds, 3) if stats.dispatch_seconds else "",
            "验证完成时间": stats.verification_completed_at,
            "验证完成耗时秒": round(stats.verification_seconds, 3) if stats.verification_seconds else "",
        }
        for item in rows:
            for header, value in values.items():
                if header in STATUS_COLUMNS and header in self.headers:
                    self.ws.cell(item.row, self.headers[header]).value = value

    def save(self) -> Path:
        try:
            self.wb.save(self.saved_path)
            return self.saved_path
        except PermissionError:
            if self.saved_path == self.path:
                self.saved_path = self.recovery_path()
                self.writeback_mode = "recovery_copy"
                self.wb.save(self.saved_path)
                return self.saved_path
            raise


class WinPoint(ctypes.Structure):
    _fields_ = [("x", ctypes.c_long), ("y", ctypes.c_long)]


class WinDropFiles(ctypes.Structure):
    _fields_ = [
        ("pFiles", ctypes.c_uint32),
        ("pt", WinPoint),
        ("fNC", ctypes.c_int32),
        ("fWide", ctypes.c_int32),
    ]


class WinRect(ctypes.Structure):
    _fields_ = [
        ("left", ctypes.c_long),
        ("top", ctypes.c_long),
        ("right", ctypes.c_long),
        ("bottom", ctypes.c_long),
    ]


class WeComGui:
    CF_UNICODETEXT = 13
    CF_HDROP = 15
    GMEM_MOVEABLE = 0x0002
    KEYEVENTF_KEYUP = 0x0002
    MOUSEEVENTF_LEFTDOWN = 0x0002
    MOUSEEVENTF_LEFTUP = 0x0004
    WM_PASTE = 0x0302
    SW_RESTORE = 9
    HWND_TOPMOST = ctypes.c_void_p(-1)
    HWND_NOTOPMOST = ctypes.c_void_p(-2)
    SWP_NOMOVE = 0x0002
    SWP_NOSIZE = 0x0001
    SWP_SHOWWINDOW = 0x0040
    VK_MENU = 0x12
    VK_BACK = 0x08
    VK_CONTROL = 0x11
    VK_ESCAPE = 0x1B
    VK_SHIFT = 0x10
    VK_INSERT = 0x2D
    VK_RETURN = 0x0D
    VK_A = 0x41
    VK_S = 0x53
    VK_V = 0x56

    def __init__(
        self,
        folder: Path,
        batch_id: str,
        run_dir: Path,
        wecom_exe: Path | None = None,
        verify_target: bool = True,
        verify_search_box: bool = True,
        capture_evidence: bool = True,
        low_evidence: bool = False,
        ocr_script: Path = OCR_SCRIPT_PATH,
        search_retries: int = 1,
        search_wait: float = 0.22,
        chat_wait: float = 0.22,
        text_wait: float = 0.22,
        file_wait: float = 0.45,
        file_wait_max: float = 1.25,
        file_wait_per_mb: float = 0.65,
        send_settle_timeout: float = 45.0,
        send_settle_interval: float = 0.6,
        between_rows: float = 0.02,
        paste_method_order: str = DEFAULT_PASTE_METHOD_ORDER,
        safe_fast: bool = False,
        trust_clipboard_paste: bool = False,
        fast_input_check: bool = False,
        lean_evidence_fast: bool = False,
        ultra_fast_dispatch: bool = False,
        dispatch_only_fast: bool = False,
        blind_dispatch_fast: bool = False,
        normalize_window_size: bool = False,
        window_width: int = 1092,
        window_height: int = 818,
    ) -> None:
        if platform.system() != "Windows":
            raise RuntimeError("企业微信 GUI 发送仅支持 Windows")
        self.enable_dpi_awareness()
        self.folder = folder
        self.batch_id = batch_id
        self.run_dir = run_dir
        self.wecom_exe = wecom_exe
        self.verify_target = verify_target
        self.verify_search_box = verify_search_box
        self.capture_evidence = capture_evidence
        self.low_evidence = low_evidence
        self.ocr_script = ocr_script
        self.search_retries = max(1, search_retries)
        self.search_wait = search_wait
        self.chat_wait = chat_wait
        self.text_wait = text_wait
        self.file_wait = file_wait
        self.file_wait_max = file_wait_max
        self.file_wait_per_mb = file_wait_per_mb
        self.send_settle_timeout = send_settle_timeout
        self.send_settle_interval = send_settle_interval
        self.between_rows = between_rows
        self.paste_method_order = paste_method_order
        self.safe_fast = safe_fast
        self.trust_clipboard_paste = trust_clipboard_paste
        self.fast_input_check = fast_input_check
        self.lean_evidence_fast = lean_evidence_fast
        self.ultra_fast_dispatch = ultra_fast_dispatch
        self.dispatch_only_fast = dispatch_only_fast
        self.blind_dispatch_fast = blind_dispatch_fast
        self.normalize_window_size = normalize_window_size
        self.normalized_window_width = max(620, int(window_width or 1092))
        self.normalized_window_height = max(360, int(window_height or 818))
        self.current_target_key = ""
        self.confirmed_target_key = ""
        self.pending_search_result_key = ""
        self.pending_search_result_level = ""
        self.pending_search_result_detail = ""
        self.pending_open_attempt_method = ""
        self.pending_before_open_title_screenshot = ""
        self.open_failure_cache: dict[str, str] = {}
        self.window_hwnd = 0
        self.last_rect: WinRect | None = None
        self.user32 = ctypes.WinDLL("user32", use_last_error=True)
        self.kernel32 = ctypes.WinDLL("kernel32", use_last_error=True)
        self._configure_apis()

    def input_locator_cache_path(self) -> Path:
        return input_locator_cache_path_for_folder(self.folder)

    def input_locator_cache_key(self) -> str:
        app_name = cell_text(getattr(self, "_active_mac_app", "")) or cell_text(getattr(self, "mac_app_name", ""))
        return f"{platform.system()}::{app_name or 'WeCom'}"

    def load_input_locator_cache(self) -> dict[str, Any]:
        path = self.input_locator_cache_path()
        try:
            if not path.exists():
                return {}
            with path.open("r", encoding="utf-8") as src:
                payload = json.load(src)
            return payload if isinstance(payload, dict) else {}
        except Exception:
            return {}

    def remember_input_locator(self, rect: WinRect, point: tuple[int, int], source: str) -> None:
        x, y = point
        if not self.point_on_screen((x, y)):
            return
        width = max(1, rect.right - rect.left)
        height = max(1, rect.bottom - rect.top)
        rel_x = x - rect.left
        rel_y = y - rect.top
        if rel_x <= 0 or rel_y <= 0 or rel_x >= width or rel_y >= height:
            return
        cache = self.load_input_locator_cache()
        cache[self.input_locator_cache_key()] = {
            "updated_at": now_text(),
            "source": source,
            "window_width": width,
            "window_height": height,
            "input_point": [int(rel_x), int(rel_y)],
        }
        path = self.input_locator_cache_path()
        try:
            path.parent.mkdir(parents=True, exist_ok=True)
            with path.open("w", encoding="utf-8") as out:
                json.dump(cache, out, ensure_ascii=False, indent=2)
        except Exception:
            return

    def cached_input_locator(self, rect: WinRect) -> tuple[tuple[int, int], str] | None:
        record = self.load_input_locator_cache().get(self.input_locator_cache_key())
        if not isinstance(record, dict):
            return None
        point = record.get("input_point")
        if not isinstance(point, list) or len(point) != 2:
            return None
        try:
            rel_x = int(point[0])
            rel_y = int(point[1])
        except Exception:
            return None
        width = rect.right - rect.left
        height = rect.bottom - rect.top
        if rel_x <= 0 or rel_y <= 0 or rel_x >= width or rel_y >= height:
            return None
        absolute = (rect.left + rel_x, rect.top + rel_y)
        if not self.point_on_screen(absolute):
            return None
        detail = f"来源={record.get('source', 'unknown')}；更新时间={record.get('updated_at', '')}；相对坐标={rel_x},{rel_y}"
        return absolute, detail

    def clear_input_locator_cache_entry(self) -> None:
        cache = self.load_input_locator_cache()
        key = self.input_locator_cache_key()
        if key not in cache:
            return
        cache.pop(key, None)
        path = self.input_locator_cache_path()
        try:
            path.parent.mkdir(parents=True, exist_ok=True)
            with path.open("w", encoding="utf-8") as out:
                json.dump(cache, out, ensure_ascii=False, indent=2)
        except Exception:
            return

    def normalize_window_if_needed(self, stats: RunStats | None = None) -> WinRect:
        return self.activate()

    @staticmethod
    def enable_dpi_awareness() -> None:
        try:
            user32 = ctypes.WinDLL("user32")
            if hasattr(user32, "SetProcessDpiAwarenessContext"):
                user32.SetProcessDpiAwarenessContext(ctypes.c_void_p(-4))
                return
            user32.SetProcessDPIAware()
        except Exception:
            pass

    def _configure_apis(self) -> None:
        enum_type = ctypes.WINFUNCTYPE(ctypes.c_bool, ctypes.c_void_p, ctypes.c_void_p)
        self._enum_type = enum_type
        self.user32.EnumWindows.argtypes = [enum_type, ctypes.c_void_p]
        self.user32.EnumWindows.restype = ctypes.c_bool
        self.user32.GetWindowTextW.argtypes = [ctypes.c_void_p, ctypes.c_wchar_p, ctypes.c_int]
        self.user32.GetClassNameW.argtypes = [ctypes.c_void_p, ctypes.c_wchar_p, ctypes.c_int]
        self.user32.IsWindow.argtypes = [ctypes.c_void_p]
        self.user32.IsWindow.restype = ctypes.c_bool
        self.user32.IsWindowVisible.argtypes = [ctypes.c_void_p]
        self.user32.GetWindowRect.argtypes = [ctypes.c_void_p, ctypes.POINTER(WinRect)]
        self.user32.GetSystemMetrics.argtypes = [ctypes.c_int]
        self.user32.GetSystemMetrics.restype = ctypes.c_int
        self.user32.MoveWindow.argtypes = [ctypes.c_void_p, ctypes.c_int, ctypes.c_int, ctypes.c_int, ctypes.c_int, ctypes.c_bool]
        self.user32.ShowWindow.argtypes = [ctypes.c_void_p, ctypes.c_int]
        self.user32.SetForegroundWindow.argtypes = [ctypes.c_void_p]
        self.user32.SetWindowPos.argtypes = [
            ctypes.c_void_p,
            ctypes.c_void_p,
            ctypes.c_int,
            ctypes.c_int,
            ctypes.c_int,
            ctypes.c_int,
            ctypes.c_uint,
        ]
        self.user32.BringWindowToTop.argtypes = [ctypes.c_void_p]
        self.user32.GetForegroundWindow.argtypes = []
        self.user32.GetForegroundWindow.restype = ctypes.c_void_p
        self.user32.SetCursorPos.argtypes = [ctypes.c_int, ctypes.c_int]
        self.user32.mouse_event.argtypes = [ctypes.c_uint, ctypes.c_uint, ctypes.c_uint, ctypes.c_uint, ctypes.c_void_p]
        self.user32.keybd_event.argtypes = [ctypes.c_ubyte, ctypes.c_ubyte, ctypes.c_uint, ctypes.c_void_p]
        self.user32.WindowFromPoint.argtypes = [WinPoint]
        self.user32.WindowFromPoint.restype = ctypes.c_void_p
        self.user32.SendMessageW.argtypes = [ctypes.c_void_p, ctypes.c_uint, ctypes.c_void_p, ctypes.c_void_p]
        self.user32.SendMessageW.restype = ctypes.c_void_p
        self.user32.OpenClipboard.argtypes = [ctypes.c_void_p]
        self.user32.EmptyClipboard.argtypes = []
        self.user32.SetClipboardData.argtypes = [ctypes.c_uint, ctypes.c_void_p]
        self.user32.GetClipboardData.argtypes = [ctypes.c_uint]
        self.user32.GetClipboardData.restype = ctypes.c_void_p
        self.user32.CloseClipboard.argtypes = []
        self.kernel32.GlobalAlloc.argtypes = [ctypes.c_uint, ctypes.c_size_t]
        self.kernel32.GlobalAlloc.restype = ctypes.c_void_p
        self.kernel32.GlobalLock.argtypes = [ctypes.c_void_p]
        self.kernel32.GlobalLock.restype = ctypes.c_void_p
        self.kernel32.GlobalUnlock.argtypes = [ctypes.c_void_p]

    def _window_text(self, hwnd: int) -> str:
        buf = ctypes.create_unicode_buffer(256)
        self.user32.GetWindowTextW(hwnd, buf, len(buf))
        return buf.value

    def _class_name(self, hwnd: int) -> str:
        buf = ctypes.create_unicode_buffer(256)
        self.user32.GetClassNameW(hwnd, buf, len(buf))
        return buf.value

    def is_wecom_window(self, hwnd: int) -> bool:
        if not hwnd:
            return False
        title = self._window_text(hwnd)
        class_name = self._class_name(hwnd)
        return class_name == "WeWorkWindow" or title == "企业微信"

    def foreground_is_wecom(self, hwnd: int) -> bool:
        foreground = int(self.user32.GetForegroundWindow() or 0)
        return foreground == hwnd or self.is_wecom_window(foreground)

    def find_window(self) -> int:
        matches: list[tuple[int, int]] = []

        @self._enum_type
        def callback(hwnd: int, _lparam: int) -> bool:
            if self.user32.IsWindowVisible(hwnd):
                if self.is_wecom_window(hwnd):
                    rect = WinRect()
                    if self.user32.GetWindowRect(hwnd, ctypes.byref(rect)):
                        area = max(rect.right - rect.left, 0) * max(rect.bottom - rect.top, 0)
                        if area > 10000:
                            matches.append((area, hwnd))
            return True

        self.user32.EnumWindows(callback, None)
        if not matches:
            return 0
        return sorted(matches, reverse=True)[0][1]

    def force_foreground(self, hwnd: int, rect: WinRect) -> bool:
        flags = self.SWP_NOMOVE | self.SWP_NOSIZE | self.SWP_SHOWWINDOW
        for attempt in range(4):
            self.user32.ShowWindow(hwnd, self.SW_RESTORE)
            self.user32.BringWindowToTop(hwnd)
            self.user32.SetForegroundWindow(hwnd)
            time.sleep(0.08)
            if self.foreground_is_wecom(hwnd):
                return True
            if attempt == 1:
                self.user32.SetWindowPos(hwnd, self.HWND_TOPMOST, 0, 0, 0, 0, flags)
                self.user32.SetWindowPos(hwnd, self.HWND_NOTOPMOST, 0, 0, 0, 0, flags)
                time.sleep(0.08)
                if self.foreground_is_wecom(hwnd):
                    return True

            focus_x = rect.left + min(max((rect.right - rect.left) // 2, 120), 420)
            focus_y = rect.top + 14
            if self.point_on_screen((focus_x, focus_y)):
                self.click(focus_x, focus_y)
                time.sleep(0.08)
            if self.foreground_is_wecom(hwnd):
                return True
        return False

    def activate(self) -> WinRect:
        hwnd = self.window_hwnd if self.window_hwnd and self.user32.IsWindow(self.window_hwnd) else self.find_window()
        if not hwnd:
            self.launch_wecom()
            deadline = time.perf_counter() + 8
            while time.perf_counter() < deadline:
                hwnd = self.find_window()
                if hwnd:
                    break
                time.sleep(0.1)
        if not hwnd:
            raise RuntimeError("未找到企业微信主窗口")

        rect = WinRect()
        if not self.user32.GetWindowRect(hwnd, ctypes.byref(rect)):
            raise RuntimeError("无法读取企业微信窗口位置")
        rect = self.ensure_window_operable(hwnd, rect)

        if not self.force_foreground(hwnd, rect):
            self.window_hwnd = 0
            raise RuntimeError("企业微信主窗口未成功置前，已停止，避免误操作其他窗口")

        if not self.user32.GetWindowRect(hwnd, ctypes.byref(rect)):
            raise RuntimeError("无法读取企业微信窗口位置")
        rect = self.ensure_window_operable(hwnd, rect)
        self.window_hwnd = hwnd
        self.last_rect = rect
        return rect

    def virtual_screen_bounds(self) -> tuple[int, int, int, int]:
        left = self.user32.GetSystemMetrics(76)
        top = self.user32.GetSystemMetrics(77)
        width = self.user32.GetSystemMetrics(78)
        height = self.user32.GetSystemMetrics(79)
        return left, top, left + width, top + height

    def point_on_screen(self, point: tuple[int, int], margin: int = 8) -> bool:
        left, top, right, bottom = self.virtual_screen_bounds()
        x, y = point
        return left + margin <= x <= right - margin and top + margin <= y <= bottom - margin

    def ensure_window_operable(self, hwnd: int, rect: WinRect) -> WinRect:
        search_ok = self.point_on_screen(self.search_point(rect))
        input_ok = self.point_on_screen(self.input_point(rect))
        if search_ok and input_ok:
            return rect

        screen_left, screen_top, screen_right, screen_bottom = self.virtual_screen_bounds()
        width = max(rect.right - rect.left, 760)
        height = max(rect.bottom - rect.top, 560)
        max_left = max(screen_left + 20, screen_right - width - 20)
        max_top = max(screen_top + 20, screen_bottom - height - 20)
        new_left = min(max(rect.left, screen_left + 20), max_left)
        new_top = min(max(rect.top, screen_top + 20), max_top)
        self.user32.MoveWindow(hwnd, int(new_left), int(new_top), int(width), int(height), True)
        time.sleep(0.15)
        moved = WinRect()
        if self.user32.GetWindowRect(hwnd, ctypes.byref(moved)):
            return moved
        return rect

    def launch_wecom(self) -> None:
        candidates = []
        if self.wecom_exe:
            candidates.append(self.wecom_exe)
        candidates.extend(
            [
                Path(r"D:\Program Files (x86)\WXWork\WXWork.exe"),
                Path(r"C:\Program Files (x86)\WXWork\WXWork.exe"),
                Path(r"C:\Program Files\WXWork\WXWork.exe"),
            ]
        )
        exe = next((path for path in candidates if path and path.exists()), None)
        if not exe:
            raise RuntimeError("未找到 WXWork.exe，请用 --wecom-exe 指定路径")
        subprocess.Popen([str(exe)], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)

    def search_point(self, rect: WinRect) -> tuple[int, int]:
        return rect.left + 215, rect.top + 38

    def input_point(self, rect: WinRect) -> tuple[int, int]:
        cached = self.cached_input_locator(rect)
        if cached:
            return cached[0]
        left, _top, right, bottom = self.input_panel_bbox(rect)
        width = right - left
        point = (left + max(80, min(width // 2, width - 220)), bottom - 86)
        self.remember_input_locator(rect, point, "input_panel_bbox")
        return point

    def input_text_bbox(self, rect: WinRect) -> tuple[int, int, int, int]:
        left, top, right, bottom = self.input_panel_bbox(rect)
        return (
            left + 10,
            max(rect.top + 130, top + 12),
            max(left + 80, right - 10),
            max(top + 80, bottom - 24),
        )

    def send_button_point(self, rect: WinRect) -> tuple[int, int]:
        _left, _top, right, bottom = self.input_panel_bbox(rect)
        return right - 44, bottom - 36

    def submit_message(self, rect: WinRect, item: SendRow, label: str) -> None:
        send_x, send_y = self.send_button_point(rect)
        if not self.point_on_screen((send_x, send_y)):
            raise RuntimeError(f"发送按钮坐标不可见：{send_x},{send_y}")
        item.send_action_time = now_text()
        self.click(send_x, send_y)
        add_step_event(item, "点击发送", "已执行", f"{label}；坐标={send_x},{send_y}")

    def input_panel_bbox(self, rect: WinRect) -> tuple[int, int, int, int]:
        fallback = (
            rect.left + max(320, min(425, (rect.right - rect.left) - 580)),
            max(rect.top + 130, rect.bottom - 160),
            rect.right - 24,
            rect.bottom - 16,
        )
        key = (rect.left, rect.top, rect.right, rect.bottom)
        cache = getattr(self, "_input_panel_cache", None)
        if cache and cache[0] == key and time.perf_counter() - cache[1] <= 1.0:
            return cache[2]
        try:
            from PIL import ImageGrab

            image = ImageGrab.grab(bbox=self.window_bbox(rect)).convert("RGB")
            detected = detect_input_panel_bbox_from_image(image)
            if detected:
                left, top, right, bottom = detected
                panel = (rect.left + left, rect.top + top, rect.left + right, rect.top + bottom)
                now = time.perf_counter()
                self._input_panel_cache = (key, now, panel)
                self._input_panel_detected_cache = (key, now, True)
                return panel
        except Exception:
            pass
        now = time.perf_counter()
        self._input_panel_cache = (key, now, fallback)
        self._input_panel_detected_cache = (key, now, False)
        return fallback

    def input_panel_detected(self, rect: WinRect) -> bool:
        key = (rect.left, rect.top, rect.right, rect.bottom)
        self.input_panel_bbox(rect)
        cache = getattr(self, "_input_panel_detected_cache", None)
        return bool(cache and cache[0] == key and time.perf_counter() - cache[1] <= 1.0 and cache[2])

    def search_box_bbox(self, rect: WinRect) -> tuple[int, int, int, int]:
        return rect.left + 70, rect.top + 18, rect.left + 355, rect.top + 64

    def search_results_bbox(self, rect: WinRect) -> tuple[int, int, int, int]:
        return rect.left + 96, rect.top + 64, min(rect.left + 410, rect.right - 18), min(rect.top + 280, rect.bottom - 90)

    def first_search_result_bbox(self, rect: WinRect) -> tuple[int, int, int, int]:
        left, top, right, _ = self.search_results_bbox(rect)
        return left, top, right, min(top + 88, rect.bottom - 90)

    def first_search_result_point(self, rect: WinRect) -> tuple[int, int]:
        left, top, right, bottom = self.first_search_result_bbox(rect)
        return left + max(60, min((right - left) // 2, 180)), top + max(20, min((bottom - top) // 2, 44))

    def click(self, x: int, y: int) -> None:
        self.user32.SetCursorPos(x, y)
        time.sleep(0.02)
        self.user32.mouse_event(self.MOUSEEVENTF_LEFTDOWN, 0, 0, 0, None)
        time.sleep(0.015)
        self.user32.mouse_event(self.MOUSEEVENTF_LEFTUP, 0, 0, 0, None)

    def key_down(self, vk: int) -> None:
        self.user32.keybd_event(vk, 0, 0, None)

    def key_up(self, vk: int) -> None:
        self.user32.keybd_event(vk, 0, self.KEYEVENTF_KEYUP, None)

    def press(self, vk: int) -> None:
        self.key_down(vk)
        time.sleep(0.015)
        self.key_up(vk)

    def hotkey(self, *keys: int) -> None:
        for key in keys:
            self.key_down(key)
            time.sleep(0.01)
        for key in reversed(keys):
            self.key_up(key)
            time.sleep(0.01)

    def paste_clipboard(self) -> None:
        self.key_down(self.VK_CONTROL)
        time.sleep(0.05)
        self.press(self.VK_V)
        time.sleep(0.04)
        self.key_up(self.VK_CONTROL)

    def paste_clipboard_to_point(self, x: int, y: int) -> bool:
        target_hwnd = self.user32.WindowFromPoint(WinPoint(x, y))
        if not target_hwnd:
            return False
        self.user32.SendMessageW(target_hwnd, self.WM_PASTE, ctypes.c_void_p(0), ctypes.c_void_p(0))
        time.sleep(0.12)
        return True

    def paste_clipboard_to_main_window(self) -> bool:
        if not self.window_hwnd:
            return False
        self.user32.SendMessageW(self.window_hwnd, self.WM_PASTE, ctypes.c_void_p(0), ctypes.c_void_p(0))
        time.sleep(0.12)
        return True

    def clear_input_box(self, rect: WinRect) -> None:
        input_x, input_y = self.input_point(rect)
        if not self.point_on_screen((input_x, input_y)):
            return
        self.click(input_x, input_y)
        self.hotkey(self.VK_CONTROL, self.VK_A)
        self.press(self.VK_BACK)
        time.sleep(0.04)

    def _with_clipboard(self, setter: Any) -> None:
        for _attempt in range(8):
            if self.user32.OpenClipboard(None):
                break
            time.sleep(0.03)
        else:
            raise RuntimeError("无法打开剪贴板")
        try:
            self.user32.EmptyClipboard()
            setter()
        finally:
            self.user32.CloseClipboard()

    def set_text_clipboard(self, text: str) -> None:
        def setter() -> None:
            payload = (text + "\0").encode("utf-16le")
            handle = self.kernel32.GlobalAlloc(self.GMEM_MOVEABLE, len(payload))
            if not handle:
                raise RuntimeError("文本剪贴板内存分配失败")
            pointer = self.kernel32.GlobalLock(handle)
            ctypes.memmove(pointer, payload, len(payload))
            self.kernel32.GlobalUnlock(handle)
            if not self.user32.SetClipboardData(self.CF_UNICODETEXT, handle):
                raise RuntimeError("写入文本剪贴板失败")

        self._with_clipboard(setter)

    def set_text_clipboard_system(self, text: str) -> None:
        command = (
            "[Console]::InputEncoding=[System.Text.UTF8Encoding]::new();"
            "$text=[Console]::In.ReadToEnd();"
            "Set-Clipboard -Value $text"
        )
        completed = subprocess.run(
            [
                "powershell.exe",
                "-NoProfile",
                "-ExecutionPolicy",
                "Bypass",
                "-Command",
                command,
            ],
            input=text,
            capture_output=True,
            text=True,
            encoding="utf-8",
            timeout=8,
            creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
        )
        if completed.returncode != 0:
            raise RuntimeError((completed.stderr or completed.stdout or "系统剪贴板写入失败").strip())
        time.sleep(0.08)

    def get_text_clipboard(self) -> str:
        for _attempt in range(8):
            if self.user32.OpenClipboard(None):
                break
            time.sleep(0.03)
        else:
            return ""
        try:
            handle = self.user32.GetClipboardData(self.CF_UNICODETEXT)
            if not handle:
                return ""
            pointer = self.kernel32.GlobalLock(handle)
            if not pointer:
                return ""
            try:
                return ctypes.wstring_at(pointer)
            finally:
                self.kernel32.GlobalUnlock(handle)
        finally:
            self.user32.CloseClipboard()

    def set_verified_text_clipboard(self, text: str, reliable: bool = False) -> None:
        self.set_text_clipboard(text)
        if reliable:
            self.set_text_clipboard_system(text)
        clipboard_text = self.get_text_clipboard()
        text_key = compact_text(text)
        clipboard_key = compact_text(clipboard_text)
        direct_match = bool(text_key and (text_key == clipboard_key or text_key in clipboard_key))
        if not direct_match and not ocr_matches_message(text, clipboard_text):
            if not reliable:
                self.set_text_clipboard_system(text)
                clipboard_text = self.get_text_clipboard()
                clipboard_key = compact_text(clipboard_text)
                direct_match = bool(text_key and (text_key == clipboard_key or text_key in clipboard_key))
        if not direct_match and not ocr_matches_message(text, clipboard_text):
            expected = ",".join(message_ocr_fragments(text)[:4])
            actual = " ".join(cell_text(clipboard_text).split())[:120]
            raise RuntimeError(f"剪贴板写入核对失败：预期片段={expected}；实际={actual}")

    def paste_methods(self, input_x: int, input_y: int) -> list[tuple[str, Any]]:
        registry: dict[str, tuple[str, Any]] = {
            "ctrl-v": ("Ctrl+V", lambda: (self.paste_clipboard() or True)),
            "wm-point": ("WM_PASTE输入框", lambda: self.paste_clipboard_to_point(input_x, input_y)),
            "wm-window": ("WM_PASTE主窗口", self.paste_clipboard_to_main_window),
            "shift-insert": ("Shift+Insert", lambda: (self.hotkey(self.VK_SHIFT, self.VK_INSERT) or True)),
        }
        order = [cell_text(part).lower() for part in self.paste_method_order.split(",") if cell_text(part)]
        methods: list[tuple[str, Any]] = []
        for key in order:
            if key in registry:
                methods.append(registry[key])
        if not methods:
            methods.extend([registry["ctrl-v"], registry["ctrl-v"], registry["ctrl-v"]])
        return methods

    def set_file_clipboard(self, paths: list[Path]) -> None:
        try:
            self.set_file_clipboard_system(paths)
            return
        except Exception as system_error:
            fallback_error = system_error

        def setter() -> None:
            files = "\0".join(str(path) for path in paths) + "\0\0"
            files_payload = files.encode("utf-16le")
            drop = WinDropFiles()
            drop.pFiles = ctypes.sizeof(WinDropFiles)
            drop.pt = WinPoint(0, 0)
            drop.fNC = 0
            drop.fWide = 1
            drop_payload = bytes(drop)
            total_size = len(drop_payload) + len(files_payload)
            handle = self.kernel32.GlobalAlloc(self.GMEM_MOVEABLE, total_size)
            if not handle:
                raise RuntimeError("文件剪贴板内存分配失败")
            pointer = self.kernel32.GlobalLock(handle)
            ctypes.memmove(pointer, drop_payload, len(drop_payload))
            ctypes.memmove(pointer + len(drop_payload), files_payload, len(files_payload))
            self.kernel32.GlobalUnlock(handle)
            if not self.user32.SetClipboardData(self.CF_HDROP, handle):
                raise RuntimeError("写入文件剪贴板失败")

        try:
            self._with_clipboard(setter)
        except Exception as ctypes_error:
            raise RuntimeError(f"系统文件剪贴板失败：{fallback_error}；CF_HDROP失败：{ctypes_error}") from ctypes_error

    def set_file_clipboard_system(self, paths: list[Path]) -> None:
        payload = json.dumps([str(path) for path in paths], ensure_ascii=False)
        command = (
            "Add-Type -AssemblyName System.Windows.Forms;"
            "$raw=[Console]::In.ReadToEnd();"
            "$items=$raw | ConvertFrom-Json;"
            "$files=New-Object System.Collections.Specialized.StringCollection;"
            "foreach($item in $items){[void]$files.Add([string]$item)};"
            "[System.Windows.Forms.Clipboard]::SetFileDropList($files)"
        )
        completed = subprocess.run(
            [
                "powershell.exe",
                "-STA",
                "-NoProfile",
                "-ExecutionPolicy",
                "Bypass",
                "-Command",
                command,
            ],
            input=payload,
            capture_output=True,
            text=True,
            encoding="utf-8",
            timeout=8,
            creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
        )
        if completed.returncode != 0:
            raise RuntimeError((completed.stderr or completed.stdout or "系统文件剪贴板写入失败").strip())
        time.sleep(0.08)

    def evidence_path(self, item: SendRow, stage: str) -> Path:
        filename = f"row{item.row}_{safe_filename(item.target)}_{stage}_{timestamp_slug()}.png"
        return self.run_dir / filename

    def record_evidence(
        self,
        item: SendRow,
        stage: str,
        path: Path,
        bbox: tuple[int, int, int, int] | None = None,
    ) -> Path:
        record: dict[str, Any] = {
            "time": now_text(),
            "stage": stage,
            "path": str(path),
        }
        if bbox is not None:
            record["bbox"] = list(bbox)
        if getattr(self, "low_evidence", False):
            record["temporary"] = True
        item.evidence_files.append(record)
        return path

    def cleanup_row_evidence(self, item: SendRow) -> int:
        if not getattr(self, "low_evidence", False):
            return 0
        deleted = 0
        seen_paths: set[Path] = set()
        for record in item.evidence_files:
            if not record.get("temporary"):
                continue
            path = Path(cell_text(record.get("path")))
            if not path:
                continue
            seen_paths.add(path)
            try:
                if path.exists() and path.is_file() and path.parent == self.run_dir:
                    path.unlink()
                    deleted += 1
                    record["deleted"] = True
            except OSError as exc:
                record["delete_error"] = str(exc)
        row_prefix = f"row{item.row}_{safe_filename(item.target)}_"
        for path in self.run_dir.glob(f"{row_prefix}*.png"):
            if path in seen_paths:
                continue
            try:
                if path.is_file():
                    path.unlink()
                    deleted += 1
            except OSError:
                pass
        item.low_evidence_deleted_files += deleted
        item.target_screenshot = ""
        item.before_screenshot = ""
        item.after_screenshot = ""
        for event in item.step_events:
            screenshot = Path(cell_text(event.get("screenshot")))
            if screenshot.parent == self.run_dir:
                event["screenshot"] = ""
        return deleted

    def capture_bbox(self, bbox: tuple[int, int, int, int], path: Path) -> Path:
        try:
            from PIL import ImageGrab
        except Exception as exc:
            raise RuntimeError(f"截图组件不可用：{exc}") from exc

        path.parent.mkdir(parents=True, exist_ok=True)
        image = ImageGrab.grab(bbox=bbox)
        image.save(path)
        return path

    def raise_if_wecom_blocked_from_screenshot(
        self,
        item: SendRow,
        screenshot: Path,
        step: str,
    ) -> None:
        ocr_text = self.ocr_image(screenshot)
        marker = wecom_global_block_marker(ocr_text)
        if not marker:
            return
        item.after_ocr_text = ocr_text
        detail = f"企业微信当前不可发送：{marker}；需完成手机安全验证或重新登录；截图={screenshot}"
        add_step_event(item, step, "阻断", detail, str(screenshot))
        raise WeComBlockedError(detail)

    def raise_if_wecom_blocked_without_input_panel(self, item: SendRow, rect: WinRect, step: str) -> None:
        if not self.capture_evidence:
            return
        if self.input_panel_detected(rect):
            return
        screenshot = self.capture_window(item, rect, "wecom_state")
        self.raise_if_wecom_blocked_from_screenshot(item, screenshot, step)

    def title_bbox(self, rect: WinRect) -> tuple[int, int, int, int]:
        width = rect.right - rect.left
        left = rect.left + min(max(width // 4 - 70, 245), 320)
        top = rect.top + 28
        right = min(rect.right - 70, left + max(420, width // 3))
        bottom = rect.top + 62
        if right <= left + 120:
            right = rect.right - 20
        return left, top, right, bottom

    def window_bbox(self, rect: WinRect) -> tuple[int, int, int, int]:
        return rect.left, rect.top, rect.right, rect.bottom

    def send_status_bbox(self, rect: WinRect) -> tuple[int, int, int, int]:
        left = rect.left + min(230, max(180, (rect.right - rect.left) // 5))
        top = max(rect.top + 120, rect.bottom - 380)
        return left, top, rect.right - 10, rect.bottom - 28

    def capture_title(self, item: SendRow, rect: WinRect) -> Path:
        path = self.evidence_path(item, "target_title")
        bbox = self.title_bbox(rect)
        return self.record_evidence(item, "target_title", self.capture_bbox(bbox, path), bbox)

    def capture_search_box(self, item: SendRow, rect: WinRect) -> Path:
        path = self.evidence_path(item, "search_box")
        bbox = self.search_box_bbox(rect)
        return self.record_evidence(item, "search_box", self.capture_bbox(bbox, path), bbox)

    def capture_search_results(self, item: SendRow, rect: WinRect) -> Path:
        path = self.evidence_path(item, "search_results")
        bbox = self.search_results_bbox(rect)
        return self.record_evidence(item, "search_results", self.capture_bbox(bbox, path), bbox)

    def capture_first_search_result(self, item: SendRow, rect: WinRect) -> Path:
        path = self.evidence_path(item, "search_result_first")
        bbox = self.first_search_result_bbox(rect)
        return self.record_evidence(item, "search_result_first", self.capture_bbox(bbox, path), bbox)

    def capture_input_text(self, item: SendRow, rect: WinRect) -> Path:
        path = self.evidence_path(item, "input_text")
        bbox = self.input_text_bbox(rect)
        try:
            from PIL import Image
        except Exception:
            return self.record_evidence(item, "input_text", self.capture_bbox(bbox, path), bbox)

        window_path = self.evidence_path(item, "input_text_window")
        window_bbox = self.window_bbox(rect)
        self.capture_bbox(window_bbox, window_path)
        with Image.open(window_path) as image:
            scale_x = image.width / max(1, rect.right - rect.left)
            scale_y = image.height / max(1, rect.bottom - rect.top)
            left = int(max(0, (bbox[0] - rect.left) * scale_x))
            top = int(max(0, (bbox[1] - rect.top) * scale_y))
            right = int(min(image.width, (bbox[2] - rect.left) * scale_x))
            bottom = int(min(image.height, (bbox[3] - rect.top) * scale_y))
            image.crop((left, top, right, bottom)).save(path)
        return self.record_evidence(item, "input_text", path, bbox)

    def capture_window(self, item: SendRow, rect: WinRect, stage: str) -> Path:
        path = self.evidence_path(item, stage)
        bbox = self.window_bbox(rect)
        captured = self.record_evidence(item, stage, self.capture_bbox(bbox, path), bbox)
        if stage == "before_send":
            item.before_screenshot = str(captured)
        if stage == "after_send":
            item.after_screenshot = str(captured)
        return captured

    def capture_send_status(self, item: SendRow, rect: WinRect) -> Path:
        path = self.evidence_path(item, "send_status")
        bbox = self.send_status_bbox(rect)
        return self.record_evidence(item, "send_status", self.capture_bbox(bbox, path), bbox)

    def capture_latest_message_status(self, item: SendRow, screenshot_path: Path) -> Path:
        path = self.evidence_path(item, "latest_message")
        return self.record_evidence(item, "latest_message", crop_latest_outgoing_message(screenshot_path, path))

    def prepare_ocr_image(self, image_path: Path) -> Path:
        try:
            from PIL import Image, ImageEnhance, ImageOps
        except Exception:
            return image_path
        try:
            image = Image.open(image_path)
            small_text = image.width < 360 or image.height < 100
            scale = 8 if small_text else 3
            resized = image.resize((image.width * scale, image.height * scale), Image.Resampling.LANCZOS)
            gray = ImageOps.grayscale(resized)
            enhanced = ImageEnhance.Contrast(gray).enhance(2.5 if small_text else 1.8)
            ocr_path = image_path.with_name(f"{image_path.stem}_ocr.png")
            enhanced.save(ocr_path)
            return ocr_path
        except Exception:
            return image_path

    def run_ocr_file(self, image_path: Path) -> str:
        if not self.ocr_script.exists():
            raise RuntimeError(f"OCR脚本不存在：{self.ocr_script}")
        completed = subprocess.run(
            [
                "powershell.exe",
                "-NoProfile",
                "-ExecutionPolicy",
                "Bypass",
                "-File",
                str(self.ocr_script),
                "-Path",
                str(image_path),
            ],
            capture_output=True,
            text=True,
            timeout=10,
            creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
        )
        if completed.returncode != 0:
            raise RuntimeError((completed.stderr or completed.stdout or "OCR执行失败").strip())
        return completed.stdout.strip()

    def ocr_image(self, image_path: Path) -> str:
        return self.run_ocr_file(self.prepare_ocr_image(image_path))

    def accessibility_text_in_bbox(self, bbox: tuple[int, int, int, int]) -> str:
        _ = bbox
        return ""

    def save_ocr_variant(
        self,
        image_path: Path,
        suffix: str,
        *,
        crop_box: tuple[int, int, int, int] | None = None,
        scale: int = 8,
        mode: str = "color",
    ) -> Path | None:
        try:
            from PIL import Image, ImageEnhance, ImageOps
        except Exception:
            return None
        try:
            image = Image.open(image_path).convert("RGB")
            if crop_box is not None:
                image = image.crop(crop_box)
            resized = image.resize((image.width * scale, image.height * scale), Image.Resampling.LANCZOS)
            if mode == "gray":
                output = ImageEnhance.Contrast(ImageOps.grayscale(resized)).enhance(3.0)
            elif mode == "bw":
                gray = ImageOps.grayscale(resized)
                output = gray.point(lambda value: 0 if value < 190 else 255)
            else:
                output = resized
            path = image_path.with_name(f"{image_path.stem}_{suffix}.png")
            output.save(path)
            return path
        except Exception:
            return None

    def unique_paths(self, paths: Iterable[Path | None]) -> list[Path]:
        unique: list[Path] = []
        seen: set[Path] = set()
        for path in paths:
            if path is None or path in seen:
                continue
            unique.append(path)
            seen.add(path)
        return unique

    def prepare_target_title_ocr_images(self, image_path: Path) -> list[Path]:
        paths: list[Path | None] = [image_path, self.prepare_ocr_image(image_path)]
        try:
            from PIL import Image
            image = Image.open(image_path)
            width, height = image.size
            left_width = min(width, 220)
            paths.extend(
                [
                    self.save_ocr_variant(image_path, "title_color8", scale=8, mode="color"),
                    self.save_ocr_variant(image_path, "title_gray8", scale=8, mode="gray"),
                    self.save_ocr_variant(image_path, "title_bw16", scale=16, mode="bw"),
                    self.save_ocr_variant(
                        image_path,
                        "title_left_color8",
                        crop_box=(0, 0, left_width, height),
                        scale=8,
                        mode="color",
                    ),
                    self.save_ocr_variant(
                        image_path,
                        "title_left_gray8",
                        crop_box=(0, 0, left_width, height),
                        scale=8,
                        mode="gray",
                    ),
                ]
            )
        except Exception:
            pass
        return self.unique_paths(paths)

    def ocr_target_title(self, image_path: Path, candidates: Iterable[str]) -> tuple[str, str]:
        texts: list[str] = []
        for ocr_input in self.prepare_target_title_ocr_images(image_path):
            try:
                text = self.run_ocr_file(ocr_input)
            except OcrUnavailableError:
                text = ""
            if text:
                texts.append(text)
            combined = "\n".join(texts)
            matched = ocr_contains_any_title_candidate(candidates, combined)
            if matched:
                return combined, matched
        if texts:
            return "\n".join(texts), ""
        if tesseract_command() is None:
            raise OcrUnavailableError("OCR不可用：macOS Vision OCR无结果，且未安装 tesseract，当前行无法通过 OCR 核对")
        raise OcrNoTextError("OCR未识别到可核对文字：macOS Vision/tesseract 均返回空结果，当前行未发送")

    def run_ocr_file_quick(self, image_path: Path, timeout_seconds: float = 3.0) -> str:
        _ = timeout_seconds
        try:
            return self.run_ocr_file(image_path)
        except Exception:
            return ""

    def ocr_target_title_quick(
        self,
        image_path: Path,
        candidates: Iterable[str],
        timeout_seconds: float = 3.0,
    ) -> tuple[str, str]:
        text = self.run_ocr_file_quick(image_path, timeout_seconds=timeout_seconds)
        matched = ocr_contains_any_title_candidate(candidates, text) if text else ""
        if not matched and text:
            matched = ocr_contains_any_relaxed_opened_title_candidate(candidates, text)
        return text, matched

    def prepare_search_result_ocr_images(self, image_path: Path) -> list[Path]:
        paths: list[Path | None] = [image_path, self.prepare_ocr_image(image_path)]
        try:
            from PIL import Image
        except Exception:
            return self.unique_paths(paths)
        try:
            image = Image.open(image_path).convert("RGB")
            width, height = image.size
            paths.extend(
                [
                    self.save_ocr_variant(image_path, "search_color6", scale=6, mode="color"),
                    self.save_ocr_variant(
                        image_path,
                        "search_wide_color6",
                        crop_box=(min(20, width - 1), 0, width, height),
                        scale=6,
                        mode="color",
                    ),
                ]
            )
            if width > 60 and height > 20:
                # Exclude the avatar column; Windows OCR often misses small blue names
                # when the row includes the icon and the section label.
                crop = image.crop((min(35, width - 1), 0, width, height))
                scaled = crop.resize((crop.width * 8, crop.height * 8), Image.Resampling.LANCZOS)
                path = image_path.with_name(f"{image_path.stem}_name_ocr.png")
                scaled.save(path)
                paths.append(path)
                paths.append(
                    self.save_ocr_variant(
                        image_path,
                        "search_name_color6",
                        crop_box=(min(35, width - 1), 0, width, height),
                        scale=6,
                        mode="color",
                    )
                )
        except Exception:
            pass
        return self.unique_paths(paths)

    def ocr_search_result(self, image_path: Path, candidates: Iterable[str]) -> tuple[str, str]:
        texts: list[str] = []
        for ocr_input in self.prepare_search_result_ocr_images(image_path):
            try:
                text = self.run_ocr_file(ocr_input)
            except OcrUnavailableError:
                text = ""
            if text:
                texts.append(text)
            combined = "\n".join(texts)
            matched = ocr_contains_any_target(candidates, combined)
            if not matched:
                matched = ocr_contains_any_search_result_target(candidates, combined)
            if matched:
                return combined, matched
        if texts:
            return "\n".join(texts), ""
        if tesseract_command() is None:
            raise OcrUnavailableError("OCR不可用：macOS Vision OCR无结果，且未安装 tesseract，当前行无法通过 OCR 核对")
        raise OcrNoTextError("OCR未识别到可核对文字：macOS Vision/tesseract 均返回空结果，当前行未发送")

    def remember_search_result_confirmation(self, item: SendRow, level: str, detail: str) -> None:
        self.pending_search_result_key = item.cache_key
        self.pending_search_result_level = level
        self.pending_search_result_detail = detail

    def clear_search_result_confirmation(self) -> None:
        self.pending_search_result_key = ""
        self.pending_search_result_level = ""
        self.pending_search_result_detail = ""
        self.pending_before_open_title_screenshot = ""

    def remember_title_before_open(self, item: SendRow, rect: WinRect) -> None:
        if not self.capture_evidence:
            self.pending_before_open_title_screenshot = ""
            return
        path = self.evidence_path(item, "before_open_title")
        bbox = self.title_bbox(rect)
        captured = self.record_evidence(item, "before_open_title", self.capture_bbox(bbox, path), bbox)
        self.pending_before_open_title_screenshot = str(captured)

    def title_changed_after_open(self, title_screenshot: Path) -> tuple[bool, float]:
        before = getattr(self, "pending_before_open_title_screenshot", "")
        if not before:
            return False, 0.0
        ratio = image_difference_ratio(Path(before), title_screenshot)
        return ratio >= 0.015, ratio

    def selected_conversation_text(self, item: SendRow, rect: WinRect) -> tuple[str, str]:
        _ = item, rect
        return "", ""

    def confirm_target(self, item: SendRow, rect: WinRect) -> None:
        item.targetcheck_time = now_text()
        if not self.verify_target:
            item.targetcheck_status = "已跳过"
            item.targetcheck_detail = "已按参数跳过目标会话 OCR 核对"
            add_step_event(item, "目标会话核对", "已跳过", item.targetcheck_detail)
            return

        candidates = target_ocr_candidates(item.target, item.target_aliases, include_variants=True)
        quick_title_confirm = (
            self.pending_search_result_key == item.cache_key
            and self.pending_open_attempt_method == "enter"
        )
        if getattr(self, "ultra_fast_dispatch", False) and quick_title_confirm:
            screenshot = self.capture_title(item, rect)
            item.target_screenshot = str(screenshot)
            title_changed, title_change_ratio = self.title_changed_after_open(screenshot)
            if not title_changed:
                item.targetcheck_status = "未通过"
                item.targetcheck_detail = (
                    f"Ultra 快速确认失败：Enter 后标题区域未变化，疑似仍停留在上一会话，已停止发送："
                    f"预期={item.target}；标题变化={title_change_ratio:.4f}；"
                    f"打开证据={self.pending_search_result_detail[:130]}"
                )
                add_step_event(item, "目标会话核对", "未通过", item.targetcheck_detail, item.target_screenshot)
                raise RuntimeError(item.targetcheck_detail)
            item.targetcheck_status = "已通过"
            item.targetcheck_detail = (
                f"Ultra 快速确认通过：Enter 后标题区域已变化，跳过标题 OCR；"
                f"预期={item.target}；标题变化={title_change_ratio:.4f}；"
                f"打开证据={self.pending_search_result_detail[:130]}"
            )
            self.confirmed_target_key = item.cache_key
            add_step_event(item, "目标会话核对", "Ultra标题变化通过", item.targetcheck_detail, item.target_screenshot)
            return
        name_ax_text = self.accessibility_text_in_bbox(self.title_name_bbox(rect))
        if name_ax_text:
            matched = ocr_contains_any_title_candidate(candidates, name_ax_text)
            match_source = "AX-NAME"
            if not matched:
                matched = ocr_contains_any_relaxed_opened_title_candidate(candidates, name_ax_text)
                if matched:
                    match_source = "AX-NAME-RELAXED"
            if matched:
                compact_ax = " ".join(name_ax_text.split())
                item.targetcheck_status = "已通过"
                item.targetcheck_detail = f"目标会话核对命中：预期={item.target}；命中={matched}；来源={match_source}；文本={compact_ax[:180]}"
                self.confirmed_target_key = item.cache_key
                add_step_event(item, "目标会话核对", "已通过", item.targetcheck_detail)
                return

        name_screenshot = self.capture_title_name(item, rect)
        if quick_title_confirm:
            name_ocr_text, name_matched = self.ocr_target_title_quick(name_screenshot, candidates, timeout_seconds=2.5)
            if not name_matched:
                compact_quick_name_ocr = " ".join(name_ocr_text.split())
                add_step_event(
                    item,
                    "目标会话核对",
                    "快速标题姓名OCR未确认",
                    f"预期={item.target}；OCR={compact_quick_name_ocr[:160]}",
                    str(name_screenshot),
                )
        else:
            try:
                name_ocr_text, name_matched = self.ocr_target_title(name_screenshot, candidates)
            except (OcrUnavailableError, OcrNoTextError):
                name_ocr_text, name_matched = "", ""
        if name_matched:
            compact_name_ocr = " ".join(name_ocr_text.split())
            item.targetcheck_status = "已通过"
            item.targetcheck_detail = f"目标会话核对命中：预期={item.target}；命中={name_matched}；来源=OCR-NAME；文本={compact_name_ocr[:180]}"
            self.confirmed_target_key = item.cache_key
            add_step_event(item, "目标会话核对", "已通过", item.targetcheck_detail, str(name_screenshot))
            return
        if self.pending_search_result_key == item.cache_key and self.pending_search_result_level == "matched":
            near_name_matched = ocr_short_cjk_title_near_target(candidates, name_ocr_text)
            if near_name_matched:
                shared_count = ocr_short_cjk_shared_chars_count(near_name_matched, name_ocr_text)
                compact_name_ocr = " ".join(name_ocr_text.split())
                item.targetcheck_status = "已通过"
                item.targetcheck_detail = (
                    f"搜索结果已精确确认且标题姓名 OCR 近似命中，允许短中文名兜底：预期={item.target}；"
                    f"命中={near_name_matched}；共同字数={shared_count}；"
                    f"搜索结果={self.pending_search_result_detail[:130]}；标题姓名OCR={compact_name_ocr[:130]}"
                )
                self.confirmed_target_key = item.cache_key
                add_step_event(item, "目标会话核对", "短中文名标题姓名兜底通过", item.targetcheck_detail, str(name_screenshot))
                return

        ax_text = self.accessibility_text_in_bbox(self.title_bbox(rect))
        if ax_text:
            matched = ocr_contains_any_title_candidate(candidates, ax_text)
            match_source = "AX"
            if not matched:
                matched = ocr_contains_any_relaxed_opened_title_candidate(candidates, ax_text)
                if matched:
                    match_source = "AX-RELAXED"
            if matched:
                compact_ax = " ".join(ax_text.split())
                item.targetcheck_status = "已通过"
                item.targetcheck_detail = f"目标会话核对命中：预期={item.target}；命中={matched}；来源={match_source}；文本={compact_ax[:180]}"
                self.confirmed_target_key = item.cache_key
                add_step_event(item, "目标会话核对", "已通过", item.targetcheck_detail)
                return

        screenshot = self.capture_title(item, rect)
        item.target_screenshot = str(screenshot)
        if quick_title_confirm:
            ocr_text, matched = self.ocr_target_title_quick(screenshot, candidates, timeout_seconds=2.5)
            match_source = "OCR-QUICK"
            if not matched:
                compact_quick_ocr = " ".join(ocr_text.split())
                add_step_event(
                    item,
                    "目标会话核对",
                    "快速标题OCR未确认",
                    f"预期={item.target}；OCR={compact_quick_ocr[:160]}",
                    str(screenshot),
                )
        else:
            try:
                ocr_text, matched = self.ocr_target_title(screenshot, candidates)
            except OcrUnavailableError as exc:
                add_step_event(
                    item,
                    "当前会话预核对",
                    "OCR不可用继续搜索",
                    f"预期={item.target}；{exc}",
                    str(screenshot),
                )
                raise RuntimeError(f"目标会话无法安全确认：预期={item.target}；{exc}") from exc
            match_source = "OCR"
        if not matched:
            matched = ocr_contains_any_title_candidate(candidates, ocr_text)
        if not matched:
            matched = ocr_contains_any_relaxed_opened_title_candidate(candidates, ocr_text)
            if matched:
                match_source = "OCR-RELAXED"
        compact_ocr = " ".join(ocr_text.split())
        if matched:
            item.targetcheck_status = "已通过"
            item.targetcheck_detail = f"目标会话核对命中：预期={item.target}；命中={matched}；来源={match_source}；文本={compact_ocr[:180]}"
            self.confirmed_target_key = item.cache_key
            add_step_event(item, "目标会话核对", "已通过", item.targetcheck_detail, item.target_screenshot)
            return

        title_changed, title_change_ratio = self.title_changed_after_open(screenshot)
        context_marker = opened_title_has_context_marker(ocr_text)
        selected_text, selected_screenshot = "", ""
        if (
            self.pending_search_result_key == item.cache_key
            and self.pending_search_result_level in {"matched", "enter"}
        ):
            if quick_title_confirm:
                if not title_changed:
                    item.targetcheck_status = "未通过"
                    item.targetcheck_detail = (
                        f"搜索结果虽已命中，但 Enter/打开后标题区域未变化，疑似仍停留在上一会话，已停止发送："
                        f"预期={item.target}；标题变化={title_change_ratio:.4f}；"
                        f"打开证据={self.pending_search_result_detail[:130]}；标题OCR={compact_ocr[:100]}"
                    )
                    add_step_event(item, "目标会话核对", "未通过", item.targetcheck_detail, item.target_screenshot)
                    raise RuntimeError(item.targetcheck_detail)
                item.targetcheck_status = "已通过"
                search_level_text = "搜索结果已精确确认" if self.pending_search_result_level == "matched" else "已执行 Enter 打开"
                item.targetcheck_detail = (
                    f"{search_level_text}且标题区域已变化，标题 OCR 未命中，允许受限兜底：预期={item.target}；"
                    f"标题变化={title_change_ratio:.4f}；打开证据={self.pending_search_result_detail[:130]}；"
                    f"标题OCR={compact_ocr[:100]}"
                )
                self.confirmed_target_key = item.cache_key
                add_step_event(item, "目标会话核对", "Enter主路径兜底通过", item.targetcheck_detail, item.target_screenshot)
                return
            selected_text, selected_screenshot = self.selected_conversation_text(item, rect)
            selected_matched = ocr_contains_any_target(candidates, selected_text) if selected_text else ""
            if not selected_matched and selected_text:
                selected_matched = ocr_contains_any_search_result_target(candidates, selected_text)
            selected_near_matched = ""
            if not selected_matched and selected_text:
                selected_near_matched = ocr_near_matches_any_target(candidates, selected_text)
            if selected_matched:
                compact_selected = " ".join(selected_text.split())
                item.targetcheck_status = "已通过"
                item.targetcheck_detail = (
                    f"搜索结果已精确确认且左侧选中会话命中，允许兜底：预期={item.target}；"
                    f"命中={selected_matched}；搜索结果={self.pending_search_result_detail[:130]}；"
                    f"左侧选中={compact_selected[:130]}；标题OCR={compact_ocr[:90]}"
                )
                self.confirmed_target_key = item.cache_key
                add_step_event(item, "目标会话核对", "左侧选中会话兜底通过", item.targetcheck_detail, selected_screenshot or item.target_screenshot)
                return
            if selected_near_matched and self.input_panel_detected(rect):
                compact_selected = " ".join(selected_text.split())
                shared_count = cjk_shared_chars_count(selected_near_matched, selected_text)
                item.targetcheck_status = "已通过"
                item.targetcheck_detail = (
                    f"搜索结果已精确确认、左侧选中会话近似命中且输入区存在，允许受限兜底：预期={item.target}；"
                    f"命中={selected_near_matched}；共同字数={shared_count}；"
                    f"搜索结果={self.pending_search_result_detail[:130]}；左侧选中={compact_selected[:130]}；标题OCR={compact_ocr[:90]}"
                )
                self.confirmed_target_key = item.cache_key
                add_step_event(item, "目标会话核对", "左侧选中会话近似兜底通过", item.targetcheck_detail, selected_screenshot or item.target_screenshot)
                return
            near_title_matched = ocr_short_cjk_title_near_target(candidates, ocr_text)
            if not near_title_matched:
                near_title_matched = ocr_near_matches_any_target(candidates, ocr_text)
            if near_title_matched:
                shared_count = ocr_short_cjk_shared_chars_count(near_title_matched, ocr_text)
                if not shared_count:
                    shared_count = cjk_shared_chars_count(near_title_matched, ocr_text)
                item.targetcheck_status = "已通过"
                item.targetcheck_detail = (
                    f"搜索结果已精确确认且标题 OCR 近似命中，允许短中文名兜底：预期={item.target}；"
                    f"命中={near_title_matched}；共同字数={shared_count}；"
                    f"搜索结果={self.pending_search_result_detail[:130]}；标题OCR={compact_ocr[:130]}"
                )
                self.confirmed_target_key = item.cache_key
                add_step_event(item, "目标会话核对", "短中文名标题兜底通过", item.targetcheck_detail, item.target_screenshot)
                return
        if (
            self.pending_open_attempt_method == "enter"
            and self.pending_search_result_key == item.cache_key
            and self.pending_search_result_level == "enter"
            and title_changed
            and context_marker
        ):
            item.targetcheck_status = "已通过"
            item.targetcheck_detail = (
                f"Enter打开后标题区域已切换且标题含职位/状态上下文，允许受限兜底：预期={item.target}；"
                f"标题变化={title_change_ratio:.3f}；上下文={context_marker}；标题OCR={compact_ocr[:150]}"
            )
            self.confirmed_target_key = item.cache_key
            add_step_event(item, "目标会话核对", "Enter标题上下文兜底通过", item.targetcheck_detail, item.target_screenshot)
            return
        if self.pending_search_result_key == item.cache_key and self.pending_search_result_level == "matched":
            input_panel_ok = self.input_panel_detected(rect)
            selected_near_matched = ocr_near_matches_any_target(candidates, selected_text) if selected_text else ""
            title_near_matched = ocr_near_matches_any_target(candidates, ocr_text)
            if input_panel_ok and (selected_near_matched or title_near_matched):
                near_matched = selected_near_matched or title_near_matched
                shared_source = "左侧选中会话" if selected_near_matched else "标题OCR"
                shared_text = selected_text if selected_near_matched else ocr_text
                shared_count = cjk_shared_chars_count(near_matched, shared_text)
                compact_selected = " ".join(selected_text.split())
                item.targetcheck_status = "已通过"
                item.targetcheck_detail = (
                    f"搜索结果已精确确认，标题区域变化不足但{shared_source}近似命中且输入区存在，允许受限兜底：预期={item.target}；"
                    f"命中={near_matched}；共同字数={shared_count}；标题变化={title_change_ratio:.3f}；"
                    f"搜索结果={self.pending_search_result_detail[:130]}；左侧选中={compact_selected[:120]}；标题OCR={compact_ocr[:100]}"
                )
                self.confirmed_target_key = item.cache_key
                add_step_event(item, "目标会话核对", "搜索结果低变化受限兜底通过", item.targetcheck_detail, selected_screenshot or item.target_screenshot)
                return
            add_step_event(
                item,
                "目标会话核对",
                "兜底拒绝",
                (
                    f"搜索结果已精确确认，但点击后标题区域未发生变化，且未取得左侧选中/标题近似命中或输入区证据，"
                    f"拒绝发送避免停留在上一聊天；标题变化={title_change_ratio:.3f}"
                ),
                item.target_screenshot,
            )

        if (
            self.pending_search_result_key == item.cache_key
            and self.pending_search_result_level == "near"
            and near_search_result_fallback_allowed(item.target, ocr_text)
        ):
            if title_changed:
                item.targetcheck_status = "已通过"
                item.targetcheck_detail = (
                    f"搜索结果疑似精确命中、标题区域已切换且标题 OCR 有同名特征，允许受限兜底：预期={item.target}；"
                    f"标题变化={title_change_ratio:.3f}；搜索结果={self.pending_search_result_detail[:150]}；标题OCR={compact_ocr[:110]}"
                )
                self.confirmed_target_key = item.cache_key
                add_step_event(item, "目标会话核对", "搜索结果受限兜底通过", item.targetcheck_detail, item.target_screenshot)
                return
            add_step_event(
                item,
                "目标会话核对",
                "受限兜底拒绝",
                f"搜索结果疑似命中，但点击后标题区域未发生变化，拒绝发送避免停留在上一聊天；标题变化={title_change_ratio:.3f}",
                item.target_screenshot,
            )

        item.targetcheck_status = "未通过"
        item.targetcheck_detail = (
            f"目标会话 OCR 未命中：预期={item.target}；候选={','.join(candidates)}；OCR={compact_ocr[:180]}；截图={screenshot}"
        )
        add_step_event(item, "目标会话核对", "未通过", item.targetcheck_detail, item.target_screenshot)
        raise RuntimeError(item.targetcheck_detail)

    def use_current_target_if_matched(self, item: SendRow, rect: WinRect) -> bool:
        if not self.verify_target:
            return False
        candidates = target_ocr_candidates(item.target, item.target_aliases, include_variants=False)
        ax_text = self.accessibility_text_in_bbox(self.title_bbox(rect))
        if ax_text:
            matched = ocr_contains_any_title_candidate(candidates, ax_text)
            if matched:
                compact_ax = " ".join(ax_text.split())
                item.targetcheck_time = now_text()
                item.targetcheck_status = "已通过"
                item.target_screenshot = ""
                item.targetcheck_detail = f"当前会话核对命中：预期={item.target}；命中={matched}；来源=AX；文本={compact_ax[:180]}"
                self.confirmed_target_key = item.cache_key
                self.current_target_key = item.cache_key
                add_step_event(item, "当前会话预核对", "已通过", item.targetcheck_detail)
                return True

        screenshot = self.capture_title(item, rect)
        ocr_text, matched = self.ocr_target_title(screenshot, candidates)
        compact_ocr = " ".join(ocr_text.split())
        if not matched:
            add_step_event(
                item,
                "当前会话预核对",
                "未命中继续搜索",
                f"预期={item.target}；OCR={compact_ocr[:160]}",
                str(screenshot),
            )
            return False
        item.targetcheck_time = now_text()
        item.targetcheck_status = "已通过"
        item.target_screenshot = str(screenshot)
        item.targetcheck_detail = f"当前会话核对命中：预期={item.target}；命中={matched}；来源=OCR；文本={compact_ocr[:180]}"
        self.confirmed_target_key = item.cache_key
        self.current_target_key = item.cache_key
        add_step_event(item, "当前会话预核对", "已通过", item.targetcheck_detail, item.target_screenshot)
        return True

    def search_queries(self, item: SendRow) -> list[str]:
        queries: list[str] = []
        for value in target_ocr_candidates(item.target, item.target_aliases, include_variants=False):
            text = cell_text(value)
            if text and text not in queries:
                queries.append(text)
        return queries or [item.target]

    def confirm_search_box(self, item: SendRow, rect: WinRect, query: str) -> None:
        if not self.verify_search_box:
            return
        screenshot = self.capture_search_box(item, rect)
        ax_text = self.accessibility_text_in_bbox(self.search_box_bbox(rect))
        ocr_text = ax_text or self.ocr_image(screenshot)
        candidates = target_ocr_candidates(item.target, [*item.target_aliases, query], include_variants=False)
        if ocr_contains_any_target(candidates, ocr_text):
            source = "AX" if ax_text else "OCR"
            add_step_event(item, "搜索框核对", "已通过", f"输入={query}；来源={source}", str(screenshot))
            return
        compact_ocr = " ".join(ocr_text.split())
        if platform.system() == "Darwin":
            add_step_event(
                item,
                "搜索框核对",
                "预警继续",
                (
                    f"输入={query}；OCR/AX={compact_ocr[:120]}；"
                    "macOS 企业微信搜索框区域容易读到会话列表，继续由搜索结果和会话标题做安全核对"
                ),
                str(screenshot),
            )
            return
        if not compact_ocr:
            add_step_event(item, "搜索框核对", "OCR为空重试", f"输入={query}；搜索框未确认收到目标名", str(screenshot))
            raise RuntimeError(f"搜索框 OCR 为空，未确认输入目标名：输入={query}；截图={screenshot}")
        query_key = compact_text(query)
        ocr_key = compact_text(ocr_text)
        if query_key and ocr_key and any(ch in ocr_key for ch in query_key):
            add_step_event(
                item,
                "搜索框核对",
                "部分识别继续",
                f"输入={query}；OCR={compact_ocr[:120]}；点击后继续目标会话核对",
                str(screenshot),
            )
            return
        if marker_in_ocr(ocr_text, ("Search", "搜索")) or len(ocr_key) >= 2:
            add_step_event(
                item,
                "搜索框核对",
                "未输入目标",
                f"输入={query}；OCR={compact_ocr[:120]}；搜索框未识别到目标名，重试搜索",
                str(screenshot),
            )
            raise RuntimeError(f"搜索框未输入目标名：输入={query}；OCR={compact_ocr[:120]}；截图={screenshot}")
        add_step_event(
            item,
            "搜索框核对",
            "OCR不一致继续",
            (
                f"输入={query}；OCR={compact_ocr[:120]}；"
                "搜索框截图仅作预警，继续打开结果并由目标会话标题精确核对"
            ),
            str(screenshot),
        )
        return

    def set_search_box_text(self, rect: WinRect, text: str) -> bool:
        _ = (rect, text)
        return False

    def confirm_first_search_result(self, item: SendRow, rect: WinRect, query: str) -> None:
        self.clear_search_result_confirmation()
        if not self.verify_target:
            return
        screenshot = self.capture_first_search_result(item, rect)
        candidates = target_ocr_candidates(item.target, [*item.target_aliases, query], include_variants=False)
        ax_text = self.accessibility_text_in_bbox(self.first_search_result_bbox(rect))
        ax_matched = ocr_contains_any_target(candidates, ax_text) if ax_text else ""
        if not ax_matched and ax_text:
            ax_matched = ocr_contains_any_search_result_target(candidates, ax_text)
        if ax_matched:
            ocr_text = ax_text
            matched = ax_matched
        else:
            ocr_text, matched = self.ocr_search_result(screenshot, candidates)
        compact_ocr = " ".join(ocr_text.split())
        if matched:
            source = "AX" if ax_matched else "OCR"
            add_step_event(
                item,
                "搜索结果核对",
                "已通过",
                f"第一个搜索结果匹配：预期={item.target}；命中={matched}；来源={source}；文本={compact_ocr[:160]}",
                str(screenshot),
            )
            self.remember_search_result_confirmation(item, "matched", f"命中={matched}；OCR={compact_ocr[:160]}")
            return
        near_miss = ocr_near_miss_target(candidates, ocr_text)
        if near_miss:
            add_step_event(
                item,
                "搜索结果核对",
                "疑似命中继续强核对",
                f"预期={item.target}；疑似命中={near_miss}；OCR={compact_ocr[:160]}；将打开候选后做目标标题强核对",
                str(screenshot),
            )
            self.remember_search_result_confirmation(item, "near", f"疑似={near_miss}；OCR={compact_ocr[:160]}")
            return
        compact_key = compact_text_strict(ocr_text)
        raw_key = strip_ocr_spaces_strict(ocr_text)
        if any(compact_text_strict(marker) in compact_key or strip_ocr_spaces_strict(marker) in raw_key for marker in SEARCH_GLOBAL_MARKERS):
            time.sleep(max(self.search_wait, 0.45))
            retry_screenshot = self.capture_first_search_result(item, rect)
            retry_ax_text = self.accessibility_text_in_bbox(self.first_search_result_bbox(rect))
            if retry_ax_text:
                retry_ocr_text = retry_ax_text
                retry_matched = ocr_contains_any_target(candidates, retry_ax_text)
                if not retry_matched:
                    retry_matched = ocr_contains_any_search_result_target(candidates, retry_ax_text)
            else:
                try:
                    retry_ocr_text, retry_matched = self.ocr_search_result(retry_screenshot, candidates)
                except OcrUnavailableError as exc:
                    add_step_event(
                        item,
                        "搜索结果核对",
                        "OCR不可用跳过",
                        f"预期={item.target}；搜索结果无法通过 AX/OCR 安全确认；{exc}",
                        str(retry_screenshot),
                    )
                    raise RuntimeError(f"搜索结果无法安全确认，已跳过发送：预期={item.target}；{exc}") from exc
            retry_compact_ocr = " ".join(retry_ocr_text.split())
            if retry_matched:
                add_step_event(
                    item,
                    "搜索结果核对",
                    "已通过",
                    f"本地结果刷新后匹配：预期={item.target}；命中={retry_matched}；OCR={retry_compact_ocr[:160]}",
                    str(retry_screenshot),
                )
                self.remember_search_result_confirmation(item, "matched", f"命中={retry_matched}；OCR={retry_compact_ocr[:160]}")
                return
            retry_near_miss = ocr_near_miss_target(candidates, retry_ocr_text)
            if retry_near_miss:
                add_step_event(
                    item,
                    "搜索结果核对",
                    "疑似命中继续强核对",
                    f"刷新后预期={item.target}；疑似命中={retry_near_miss}；OCR={retry_compact_ocr[:160]}；将打开候选后做目标标题强核对",
                    str(retry_screenshot),
                )
                self.remember_search_result_confirmation(item, "near", f"疑似={retry_near_miss}；OCR={retry_compact_ocr[:160]}")
                return
            ocr_text = retry_ocr_text
            compact_ocr = retry_compact_ocr
            compact_key = compact_text_strict(ocr_text)
            raw_key = strip_ocr_spaces_strict(ocr_text)
        if any(compact_text_strict(marker) in compact_key or strip_ocr_spaces_strict(marker) in raw_key for marker in SEARCH_GLOBAL_MARKERS):
            add_step_event(
                item,
                "搜索结果核对",
                "未找到本地精确结果",
                f"不打开聊天窗口：预期={item.target}；OCR={compact_ocr[:160]}",
                str(screenshot),
            )
            raise RuntimeError(
                f"搜索结果未找到本地精确目标，已停止打开聊天窗口：预期={item.target}；OCR={compact_ocr[:160]}；截图={screenshot}"
            )
        if not compact_ocr and not screenshot_has_visible_search_candidate(screenshot):
            time.sleep(max(self.search_wait, 0.35))
            retry_screenshot = self.capture_first_search_result(item, rect)
            try:
                retry_ocr_text, retry_matched = self.ocr_search_result(retry_screenshot, candidates)
            except OcrUnavailableError as exc:
                add_step_event(
                    item,
                    "搜索结果核对",
                    "OCR不可用跳过",
                    f"预期={item.target}；空白等待后仍无法 OCR 核对；{exc}",
                    str(retry_screenshot),
                )
                raise RuntimeError(f"搜索结果无法安全确认，已跳过发送：预期={item.target}；{exc}") from exc
            retry_compact_ocr = " ".join(retry_ocr_text.split())
            if retry_matched:
                add_step_event(
                    item,
                    "搜索结果核对",
                    "已通过",
                    f"空白等待后匹配：预期={item.target}；命中={retry_matched}；OCR={retry_compact_ocr[:160]}",
                    str(retry_screenshot),
                )
                self.remember_search_result_confirmation(item, "matched", f"命中={retry_matched}；OCR={retry_compact_ocr[:160]}")
                return
            retry_near_miss = ocr_near_miss_target(candidates, retry_ocr_text)
            if retry_near_miss:
                add_step_event(
                    item,
                    "搜索结果核对",
                    "疑似命中继续强核对",
                    f"空白等待后预期={item.target}；疑似命中={retry_near_miss}；OCR={retry_compact_ocr[:160]}；将打开候选后做目标标题强核对",
                    str(retry_screenshot),
                )
                self.remember_search_result_confirmation(item, "near", f"疑似={retry_near_miss}；OCR={retry_compact_ocr[:160]}")
                return
            if not retry_compact_ocr and not screenshot_has_visible_search_candidate(retry_screenshot):
                add_step_event(
                    item,
                    "搜索结果核对",
                    "搜索结果为空",
                    f"不打开聊天窗口：预期={item.target}；首个本地结果区域为空白",
                    str(retry_screenshot),
                )
                raise RuntimeError(f"搜索结果为空，已停止打开聊天窗口：预期={item.target}；截图={retry_screenshot}")
            screenshot = retry_screenshot
            ocr_text = retry_ocr_text
            compact_ocr = retry_compact_ocr
        if compact_ocr:
            add_step_event(
                item,
                "搜索结果核对",
                "OCR未确认继续",
                f"预期={item.target}；OCR={compact_ocr[:160]}；将打开候选后做目标标题强核对",
                str(screenshot),
            )
            self.remember_search_result_confirmation(item, "unconfirmed", f"OCR={compact_ocr[:160]}")
            return
        add_step_event(
            item,
            "搜索结果核对",
            "OCR为空继续",
            f"预期={item.target}；将打开候选后做目标标题强核对",
            str(screenshot),
        )
        self.remember_search_result_confirmation(item, "unconfirmed", "OCR为空")

    def open_search_result_with_enter(self, item: SendRow, rect: WinRect, query: str) -> WinRect:
        self.confirm_first_search_result(item, rect, query)
        self.remember_title_before_open(item, rect)
        result_x, result_y = self.first_search_result_point(rect)
        if not self.point_on_screen((result_x, result_y)):
            raise RuntimeError(f"搜索结果坐标不可见：{result_x},{result_y}")
        add_step_event(item, "打开搜索结果", "执行中", f"输入={query}；方式=点击候选；坐标={result_x},{result_y}")
        self.click(result_x, result_y)
        time.sleep(max(self.chat_wait, 0.45))
        rect = self.activate()
        try:
            self.confirm_target(item, rect)
            return rect
        except Exception as click_error:
            add_step_event(
                item,
                "打开搜索结果",
                "点击未确认改用Enter",
                f"输入={query}；点击错误={click_error}",
            )
            self.press(self.VK_RETURN)
            time.sleep(max(self.chat_wait, 0.45))
            rect = self.activate()
            self.confirm_target(item, rect)
            return rect

    def inspect_input_text(self, item: SendRow, rect: WinRect) -> tuple[str, Path, str]:
        screenshot = self.capture_input_text(item, rect)
        ax_text = self.accessibility_text_in_bbox(self.input_text_bbox(rect))
        ocr_text = ax_text or self.ocr_image(screenshot)
        presence = input_message_presence(item.message, ocr_text)
        if presence == "matched":
            source = "AX" if ax_text else "OCR"
            add_step_event(item, "输入框文本核对", "已通过", f"已识别本次文案片段；来源={source}", str(screenshot))
            return presence, screenshot, ocr_text
        compact_ocr = " ".join(ocr_text.split())
        fragments = ",".join(
            message_ocr_fragments(item.message)[:4]
            + message_visible_fragments(item.message)[:4]
            + message_short_fragments(item.message)[:4]
        )
        status = "疑似本次内容" if presence == "partial" else "未通过"
        add_step_event(item, "输入框文本核对", status, f"片段={fragments}；OCR={compact_ocr[:120]}", str(screenshot))
        return presence, screenshot, ocr_text

    def inspect_input_text_fast(self, item: SendRow, rect: WinRect) -> tuple[str, Path, str]:
        screenshot = self.capture_input_text(item, rect)
        self.raise_if_multiselect_toolbar_visible(item, rect, "输入框快速核对")
        ax_text = self.accessibility_text_in_bbox(self.input_text_bbox(rect))
        if ax_text:
            presence = input_message_presence(item.message, ax_text)
            if presence in {"matched", "partial"}:
                add_step_event(item, "输入框文本核对", "快速通过", f"来源=AX；状态={presence}", str(screenshot))
                return presence, screenshot, ax_text
        if input_image_has_visible_text(screenshot):
            if input_image_has_left_aligned_content(screenshot):
                add_step_event(
                    item,
                    "输入框文本核对",
                    "视觉快速通过",
                    "输入区呈左对齐文本形态，且未检测到多选工具栏；跳过输入框 OCR",
                    str(screenshot),
                )
                return "matched", screenshot, "FAST_INPUT_LEFT_ALIGNED_TEXT"
            ocr_text = self.ocr_image(screenshot)
            if self.is_multiselect_toolbar_text(ocr_text):
                compact_toolbar = " ".join(ocr_text.split())
                add_step_event(
                    item,
                    "输入框文本核对",
                    "检测到多选工具栏",
                    f"输入区不是可发送编辑框；OCR={compact_toolbar[:140]}",
                    str(screenshot),
                )
                return "empty", screenshot, ocr_text
            presence = input_message_presence(item.message, ocr_text)
            compact_ocr = " ".join(ocr_text.split())
            if presence in {"matched", "partial"}:
                add_step_event(item, "输入框文本核对", "快速OCR通过", f"状态={presence}；OCR={compact_ocr[:120]}", str(screenshot))
                return presence, screenshot, ocr_text
            add_step_event(
                item,
                "输入框文本核对",
                "视觉有字但非本次文案",
                f"输入区截图有可见文本，但未识别到本次文案；OCR={compact_ocr[:120]}",
                str(screenshot),
            )
            return "empty", screenshot, ocr_text
        ocr_text = self.ocr_image(screenshot)
        presence = input_message_presence(item.message, ocr_text)
        compact_ocr = " ".join(ocr_text.split())
        if presence in {"matched", "partial"}:
            add_step_event(item, "输入框文本核对", "快速OCR通过", f"状态={presence}；OCR={compact_ocr[:120]}", str(screenshot))
            return presence, screenshot, ocr_text
        fragments = ",".join(
            message_ocr_fragments(item.message)[:4]
            + message_visible_fragments(item.message)[:4]
            + message_short_fragments(item.message)[:4]
        )
        status = "疑似本次内容" if presence == "partial" else "未通过"
        add_step_event(item, "输入框文本核对", status, f"片段={fragments}；OCR={compact_ocr[:120]}", str(screenshot))
        return presence, screenshot, ocr_text

    def multiselect_toolbar_bbox(self, rect: WinRect) -> tuple[int, int, int, int]:
        return rect.left + 260, rect.bottom - 190, rect.right - 10, rect.bottom - 5

    def multiselect_toolbar_text(self, rect: WinRect) -> str:
        return self.accessibility_text_in_bbox(self.multiselect_toolbar_bbox(rect))

    def is_multiselect_toolbar_text(self, text: str) -> bool:
        normalized = " ".join(cell_text(text).split()).lower()
        if not normalized:
            return False
        strong_markers = (
            "one-by-one forward",
            "combine forward",
            "逐条转发",
            "合并转发",
        )
        if any(marker in normalized for marker in strong_markers):
            return True
        weak_markers = (
            "favorite",
            "delete",
            "add to",
            "forward",
            "收藏",
            "删除",
            "添加到",
            "转发",
        )
        return sum(1 for marker in weak_markers if marker in normalized) >= 2

    def multiselect_toolbar_visible(self, rect: WinRect) -> tuple[bool, str]:
        text = self.multiselect_toolbar_text(rect)
        return self.is_multiselect_toolbar_text(text), text

    def exit_multiselect_if_visible(self, item: SendRow, rect: WinRect, context: str) -> WinRect:
        visible, text = self.multiselect_toolbar_visible(rect)
        if not visible:
            return rect
        compact = " ".join(text.split())
        add_step_event(
            item,
            "多选状态清理",
            "检测到多选工具栏",
            f"{context}：底部出现转发/收藏/删除工具栏，先按 Esc 退出；AX={compact[:140]}",
        )
        self.press(self.VK_ESCAPE)
        time.sleep(0.3)
        rect = self.activate()
        still_visible, still_text = self.multiselect_toolbar_visible(rect)
        if still_visible:
            compact_still = " ".join(still_text.split())
            raise RuntimeError(f"{context}：企业微信仍处于消息多选状态，已停止避免空发；AX={compact_still[:140]}")
        add_step_event(item, "多选状态清理", "已退出", f"{context}：已退出消息多选状态")
        return rect

    def raise_if_multiselect_toolbar_visible(self, item: SendRow, rect: WinRect, context: str) -> None:
        visible, text = self.multiselect_toolbar_visible(rect)
        if not visible:
            return
        compact = " ".join(text.split())
        raise RuntimeError(f"{context}：检测到消息多选工具栏，输入框未处于可发送状态；AX={compact[:140]}")

    def confirm_blind_input_nonempty(self, item: SendRow, rect: WinRect) -> None:
        screenshot = self.capture_input_text(item, rect)
        if input_image_has_visible_text(screenshot):
            if input_image_has_left_aligned_content(screenshot):
                add_step_event(
                    item,
                    "输入框最小护栏",
                    "视觉快速通过",
                    "blind-dispatch-fast：输入区呈左对齐文本形态，跳过 AX/OCR",
                    str(screenshot),
                )
                return
            ocr_text = self.ocr_image(screenshot)
            compact_ocr = " ".join(ocr_text.split())
            if self.is_multiselect_toolbar_text(ocr_text):
                add_step_event(
                    item,
                    "输入框最小护栏",
                    "检测到多选工具栏",
                    f"blind-dispatch-fast：输入区是消息多选工具栏，已停止避免空发；OCR={compact_ocr[:140]}",
                    str(screenshot),
                )
                raise RuntimeError(f"blind-dispatch-fast 检测到消息多选工具栏，未点击发送；截图={screenshot}")
            presence = input_message_presence(item.message, ocr_text)
            if presence in {"matched", "partial"}:
                add_step_event(
                    item,
                    "输入框最小护栏",
                    "轻量OCR通过",
                    f"blind-dispatch-fast：输入区识别到本次文案；状态={presence}；OCR={compact_ocr[:120]}",
                    str(screenshot),
                )
                return
            add_step_event(
                item,
                "输入框最小护栏",
                "视觉有字但未通过",
                f"blind-dispatch-fast：输入区有可见文本但不是本次文案，已停止避免误发；OCR={compact_ocr[:120]}",
                str(screenshot),
            )
            raise RuntimeError(f"blind-dispatch-fast 输入区未识别到本次文案，未点击发送；截图={screenshot}")
        ax_text = self.accessibility_text_in_bbox(self.input_text_bbox(rect))
        if ax_text:
            if self.is_multiselect_toolbar_text(ax_text):
                compact_ax = " ".join(ax_text.split())
                add_step_event(
                    item,
                    "输入框最小护栏",
                    "检测到多选工具栏",
                    f"blind-dispatch-fast：AX检测到消息多选工具栏，已停止避免空发；AX={compact_ax[:140]}",
                    str(screenshot),
                )
                raise RuntimeError(f"blind-dispatch-fast 检测到消息多选工具栏，未点击发送；截图={screenshot}")
            presence = input_message_presence(item.message, ax_text)
            if presence in {"matched", "partial"}:
                add_step_event(item, "输入框最小护栏", "AX兜底通过", f"来源=AX；状态={presence}", str(screenshot))
                return
        compact_ax = " ".join(ax_text.split())
        add_step_event(
            item,
            "输入框最小护栏",
            "未通过",
            f"blind-dispatch-fast：粘贴后输入区疑似为空，已停止避免空发；AX={compact_ax[:120]}",
            str(screenshot),
        )
        raise RuntimeError(f"blind-dispatch-fast 输入框为空，未点击发送；截图={screenshot}")

    def confirm_input_text(self, item: SendRow, rect: WinRect) -> None:
        presence, screenshot, ocr_text = self.inspect_input_text(item, rect)
        if presence == "matched":
            return
        compact_ocr = " ".join(ocr_text.split())
        fragments = ",".join(
            message_ocr_fragments(item.message)[:4]
            + message_visible_fragments(item.message)[:4]
            + message_short_fragments(item.message)[:4]
        )
        raise RuntimeError(f"文本粘贴核对失败：未在输入框识别到文案片段={fragments}；OCR={compact_ocr[:120]}；截图={screenshot}")

    def confirm_input_cleared_after_send(self, item: SendRow, rect: WinRect, *, context: str) -> bool:
        input_screenshot = self.capture_input_text(item, rect)
        ax_text = self.accessibility_text_in_bbox(self.input_text_bbox(rect))
        input_ocr = ax_text or self.ocr_image(input_screenshot)
        presence = input_message_presence(item.message, input_ocr)
        if presence in {"matched", "partial"}:
            compact_input = " ".join(input_ocr.split())
            add_step_event(
                item,
                "发送后输入框残留核对",
                "未通过",
                f"{context}后输入框仍有本次文案；OCR={compact_input[:140]}",
                str(input_screenshot),
            )
            return False
        add_step_event(item, "发送后输入框残留核对", "已通过", f"{context}后未识别到本次文案残留", str(input_screenshot))
        return True

    def confirm_input_cleared_after_send_visual(self, item: SendRow, rect: WinRect, *, context: str) -> bool:
        input_screenshot = self.capture_input_text(item, rect)
        if input_image_has_visible_text(input_screenshot):
            add_step_event(
                item,
                "发送后输入框残留核对",
                "视觉待复核",
                f"{context}后输入框仍有可见文本，需 OCR 强核对",
                str(input_screenshot),
            )
            return self.confirm_input_cleared_after_send(item, rect, context=context)
        add_step_event(
            item,
            "发送后输入框残留核对",
            "视觉通过",
            f"{context}后输入框未见明显文本，跳过输入框 OCR",
            str(input_screenshot),
        )
        return True

    def open_target(self, item: SendRow) -> float:
        start = time.perf_counter()
        cached_failure = self.open_failure_cache.get(item.cache_key)
        if cached_failure:
            add_step_event(item, "打开目标会话", "已跳过重复搜索", cached_failure)
            raise RuntimeError(f"同批次已确认该目标无法本地精确打开：{cached_failure}")
        rect = self.activate()
        self.raise_if_wecom_blocked_without_input_panel(item, rect, "打开目标前企业微信状态核对")
        if self.current_target_key == item.cache_key:
            self.confirm_target(item, rect)
            return time.perf_counter() - start
        if not getattr(self, "ultra_fast_dispatch", False) and self.use_current_target_if_matched(item, rect):
            return time.perf_counter() - start

        failures: list[str] = []
        for query in self.search_queries(item):
            for attempt in range(1, self.search_retries + 1):
                rect = self.activate()
                self.press(self.VK_ESCAPE)
                time.sleep(0.08)
                rect = self.activate()
                search_x, search_y = self.search_point(rect)
                if not self.point_on_screen((search_x, search_y)):
                    failures.append(f"搜索框坐标不可见：{search_x},{search_y}")
                    continue
                self.click(search_x, search_y)
                time.sleep(0.08)
                if not self.set_search_box_text(rect, query):
                    for _ in range(2):
                        self.hotkey(self.VK_CONTROL, self.VK_A)
                        time.sleep(0.03)
                        self.press(self.VK_BACK)
                        if hasattr(self, "VK_DELETE"):
                            self.press(getattr(self, "VK_DELETE"))
                        time.sleep(0.06)
                        self.click(search_x, search_y)
                    self.click(search_x, search_y)
                    time.sleep(0.08)
                    self.set_verified_text_clipboard(query, reliable=True)
                    for _ in range(2):
                        self.hotkey(self.VK_CONTROL, self.VK_V)
                        time.sleep(0.08)
                time.sleep(max(self.search_wait, 0.25 + 0.15 * (attempt - 1)))
                try:
                    self.confirm_search_box(item, rect, query)
                    rect = self.open_search_result_with_enter(item, rect, query)
                    self.current_target_key = item.cache_key
                    return time.perf_counter() - start
                except WeComBlockedError:
                    raise
                except Exception as exc:
                    failures.append(f"查询={query} 第{attempt}次：{exc}")
                    self.current_target_key = ""
                    self.confirmed_target_key = ""
                    self.clear_search_result_confirmation()
                    self.press(self.VK_ESCAPE)
                    time.sleep(0.12)
        error = "打开目标会话失败；" + " | ".join(failures[-4:])
        if stable_open_failure_reason(error):
            self.open_failure_cache[item.cache_key] = error
        raise RuntimeError(error)

    def paste_text_for_send(self, rect: WinRect, item: SendRow) -> None:
        rect = self.exit_multiselect_if_visible(item, rect, "粘贴前")
        input_x, input_y = self.input_point(rect)
        self.click(input_x, input_y)
        self.clear_input_box(rect)
        if getattr(self, "blind_dispatch_fast", False):
            self.set_verified_text_clipboard(item.message, reliable=True)
            first_empty_error: Exception | None = None
            for attempt in range(1, 3):
                if attempt == 2:
                    result_x, result_y = self.first_search_result_point(rect)
                    if not self.point_on_screen((result_x, result_y), margin=4):
                        break
                    add_step_event(
                        item,
                        "打开搜索结果",
                        "输入框为空后点击候选兜底",
                        f"blind-dispatch-fast：Enter后未检测到输入框内容，点击第一条候选再重试；坐标={result_x},{result_y}",
                    )
                    self.click(result_x, result_y)
                    time.sleep(max(self.chat_wait, 0.55))
                    rect = self.activate()
                    rect = self.exit_multiselect_if_visible(item, rect, "候选兜底后")
                    input_x, input_y = self.input_point(rect)
                    self.clear_input_box(rect)
                    self.set_verified_text_clipboard(item.message, reliable=True)
                rect = self.exit_multiselect_if_visible(item, rect, f"第{attempt}次粘贴前")
                self.click(input_x, input_y)
                self.hotkey(self.VK_CONTROL, self.VK_V)
                time.sleep(max(self.text_wait, 0.18))
                try:
                    self.confirm_blind_input_nonempty(item, rect)
                    add_step_event(
                        item,
                        "文本粘贴",
                        "盲发最小核对通过",
                        (
                            "blind-dispatch-fast：剪贴板已写入，输入框非空护栏通过，"
                            f"跳过完整输入框 OCR确认；尝试={attempt}"
                        ),
                    )
                    return
                except RuntimeError as exc:
                    if attempt == 1:
                        first_empty_error = exc
                        add_step_event(
                            item,
                            "文本粘贴",
                            "准备候选兜底",
                            f"首次粘贴后输入框为空，准备点击搜索结果第一条后重试；原因={exc}",
                        )
                        continue
                    raise RuntimeError(f"{exc}；首次失败={first_empty_error}") from exc
            raise RuntimeError(f"blind-dispatch-fast 输入框为空，候选兜底未执行；首次失败={first_empty_error}")
        failures: list[str] = []
        last_presence = "empty"
        last_screenshot: Path | None = None
        clipboard_verified = False
        for attempt, (method_name, paste_method) in enumerate(self.paste_methods(input_x, input_y), start=1):
            self.set_verified_text_clipboard(item.message, reliable=True)
            clipboard_verified = True
            if attempt > 1:
                if last_presence == "partial":
                    if self.trust_clipboard_paste and clipboard_verified:
                        add_step_event(
                            item,
                            "文本粘贴",
                            "剪贴板可信放行",
                            "输入框疑似已有本次内容但核对不足；已按剪贴板核对结果继续发送",
                        )
                        break
                    raise RuntimeError("文本粘贴后输入框已有疑似本次内容，但核对不足；已停止发送并保留输入框内容，避免误清空")
                if last_screenshot is not None:
                    add_step_event(
                        item,
                        "文本粘贴",
                        "保留现场",
                        f"第{attempt - 1}次粘贴后未确认文本，但已保留输入框现场，不执行清空重试；截图={last_screenshot}",
                        str(last_screenshot),
                    )
                    raise RuntimeError(
                        "文本粘贴未确认，已停止并保留输入框内容，避免误删可能已粘贴的文案；"
                        + " | ".join(failures[-2:])
                    )
                self.clear_input_box(rect)
                time.sleep(0.08)
            self.click(input_x, input_y)
            if not paste_method():
                failures.append(f"第{attempt}次：{method_name} 未找到可粘贴窗口")
                continue
            time.sleep(0.12)
            if self.fast_input_check and clipboard_verified:
                presence, screenshot, ocr_text = self.inspect_input_text_fast(item, rect)
            else:
                presence, screenshot, ocr_text = self.inspect_input_text(item, rect)
            if presence == "empty" and clipboard_verified:
                add_step_event(
                    item,
                    "文本粘贴",
                    "等待复查",
                    f"第{attempt}次({method_name})后输入框暂未识别到内容，等待渲染后复查",
                    str(screenshot),
                )
                time.sleep(0.7)
                self.click(input_x, input_y)
                time.sleep(0.08)
                if self.fast_input_check and clipboard_verified:
                    presence, screenshot, ocr_text = self.inspect_input_text_fast(item, rect)
                else:
                    presence, screenshot, ocr_text = self.inspect_input_text(item, rect)
            if presence == "empty" and clipboard_verified:
                cached_locator = self.cached_input_locator(rect)
                if cached_locator:
                    (cached_x, cached_y), locator_detail = cached_locator
                    add_step_event(
                        item,
                        "文本粘贴",
                        "定位缓存放行",
                        (
                            f"第{attempt}次({method_name})后输入框 OCR/AX 未识别到文本；"
                            f"剪贴板已核对，且输入框定位缓存可用，继续发送；坐标={cached_x},{cached_y}；{locator_detail}"
                        ),
                        str(screenshot),
                    )
                    break
            last_presence = presence
            last_screenshot = screenshot
            if presence == "matched":
                add_step_event(item, "文本粘贴", "已通过", f"方式={method_name}；尝试={attempt}")
                break
            compact_ocr = " ".join(ocr_text.split())
            if presence == "partial":
                if (self.trust_clipboard_paste or self.fast_input_check) and clipboard_verified:
                    add_step_event(
                        item,
                        "文本粘贴",
                        "剪贴板可信放行",
                        f"方式={method_name}；尝试={attempt}；输入框疑似本次内容；OCR={compact_ocr[:120]}",
                        str(screenshot),
                    )
                    break
                failure = (
                    f"第{attempt}次({method_name})：输入框疑似已有本次内容但未达到发送核对标准，"
                    f"已停止并保留现场；OCR={compact_ocr[:120]}；截图={screenshot}"
                )
                failures.append(failure)
                add_step_event(item, "文本粘贴", "保留现场", failure, str(screenshot))
                raise RuntimeError("文本粘贴核对不足，已停止发送并保留输入框内容；" + failure)
            failure = f"第{attempt}次({method_name})：输入框未识别到本次文案；状态={presence}；OCR={compact_ocr[:120]}；截图={screenshot}"
            failures.append(failure)
            add_step_event(item, "文本粘贴", "未通过", failure, str(screenshot))
        else:
            if self.trust_clipboard_paste:
                self.clear_input_box(rect)
                add_step_event(item, "文本粘贴", "失败已清空", "未确认输入框收到本次文案，已清空输入框避免草稿残留")
            raise RuntimeError("文本粘贴失败，已停止发送；" + " | ".join(failures[-2:]))

    def send_text(self, rect: WinRect, item: SendRow) -> None:
        self.paste_text_for_send(rect, item)
        if getattr(self, "blind_dispatch_fast", False):
            rect = self.exit_multiselect_if_visible(item, rect, "按 Enter 发送前")
            input_x, input_y = self.input_point(rect)
            self.click(input_x, input_y)
            item.send_action_time = now_text()
            self.press(self.VK_RETURN)
            add_step_event(item, "按 Enter 发送", "盲发已执行", f"文字；坐标={input_x},{input_y}")
            time.sleep(max(self.text_wait, 0.25))
            return
        self.submit_message(rect, item, "文字")
        time.sleep(self.text_wait)
        if not self.confirm_input_cleared_after_send(item, rect, context="点击发送"):
            item.send_action_time = now_text()
            self.press(self.VK_RETURN)
            add_step_event(item, "键盘补发", "已执行", "点击发送后输入框仍有文案，尝试 Enter 发送")
            time.sleep(max(self.text_wait, 0.35))
            if not self.confirm_input_cleared_after_send(item, rect, context="Enter补发"):
                raise RuntimeError("点击发送和 Enter 补发后文案仍在输入框，未确认发送成功")

    def send_text_with_files(self, rect: WinRect, item: SendRow, paths: list[Path]) -> None:
        self.paste_text_for_send(rect, item)
        input_x, input_y = self.input_point(rect)
        self.click(input_x, input_y)
        self.set_file_clipboard(paths)
        self.hotkey(self.VK_CONTROL, self.VK_V)
        file_wait = self.file_wait_for(paths)
        add_step_event(item, "附件粘贴", "已执行", f"数量={len(paths)}；等待={file_wait:.2f}s")
        time.sleep(max(0.25, min(file_wait, 0.8)))
        self.submit_message(rect, item, "文字+附件")
        time.sleep(max(self.text_wait, file_wait))
        if item.parts["text"] and not self.confirm_input_cleared_after_send(item, rect, context="点击发送"):
            raise RuntimeError("点击发送后文案仍在输入框，组合消息未确认发送成功")

    def send_files(self, rect: WinRect, item: SendRow, paths: list[Path]) -> None:
        if not paths:
            return
        input_x, input_y = self.input_point(rect)
        self.click(input_x, input_y)
        self.clear_input_box(rect)
        self.set_file_clipboard(paths)
        self.hotkey(self.VK_CONTROL, self.VK_V)
        time.sleep(0.12)
        item.send_action_time = now_text()
        self.press(self.VK_RETURN)
        wait_seconds = self.file_wait_for(paths)
        add_step_event(item, "发送附件", "已执行", f"数量={len(paths)}；大小等待={wait_seconds:.2f}s")
        time.sleep(wait_seconds)

    def wait_for_send_settle(self, item: SendRow, rect: WinRect, before_screenshot: Path | None = None) -> None:
        if not self.capture_evidence:
            return
        if getattr(self, "safe_fast", False):
            try:
                self.wait_for_send_settle_fast(item, rect, before_screenshot)
                return
            except RuntimeError as exc:
                fast_error = str(exc)
                if "输入框仍有本次文案" in fast_error or "红色失败标记" in fast_error:
                    raise
                add_step_event(
                    item,
                    "快速发送后核对",
                    "转完整核对",
                    f"快速核对未确认发送，改用完整核对：{brief_log_text(fast_error)}",
                    item.after_screenshot,
                )
        deadline = time.perf_counter() + self.send_settle_timeout
        last_pending = ""
        before_bubble = latest_outgoing_bubble_from_path(before_screenshot) if before_screenshot else None
        while True:
            item.after_screenshot = str(self.capture_window(item, rect, "after_send"))
            status_area = self.capture_send_status(item, rect)
            status_area_ocr = self.ocr_image(status_area)
            status_area_error = send_error_marker(status_area_ocr)
            if status_area_error:
                item.after_ocr_text = status_area_ocr
                add_step_event(item, "发送后状态区域核对", "未通过", status_area_error, str(status_area))
                raise RuntimeError(f"发送后界面出现错误提示：{status_area_error}；截图={status_area}")

            red_failure = screenshot_red_failure_indicator(Path(item.after_screenshot))
            if red_failure:
                item.after_ocr_text = red_failure
                add_step_event(item, "发送后红色失败标记核对", "未通过", red_failure, item.after_screenshot)
                raise RuntimeError(f"{red_failure}；截图={item.after_screenshot}")

            after_bubble = latest_outgoing_bubble_from_path(Path(item.after_screenshot))
            status_image = self.capture_latest_message_status(item, Path(item.after_screenshot))
            status_ocr = self.ocr_image(status_image)
            item.after_ocr_text = status_ocr
            error = send_error_marker(status_ocr)
            pending = send_pending_marker(status_ocr)
            if error:
                add_step_event(item, "最新消息核对", "未通过", error, str(status_image))
                raise RuntimeError(f"发送后界面出现错误标记：{error}；截图={status_image}")
            if not pending:
                if item.parts["text"] and ocr_matches_message(item.message, status_ocr):
                    if not self.confirm_input_cleared_after_send(item, rect, context="最新气泡匹配"):
                        raise RuntimeError("点击发送后文案仍在输入框，发送按钮可能未命中")
                    detail = "文字+附件内容/状态核对通过" if item.file_paths else "最新消息内容核对通过"
                    add_step_event(item, "最新消息核对", "已通过", detail, str(status_image))
                    return
                if item.parts["text"] and after_bubble is None:
                    if not self.confirm_input_cleared_after_send(item, rect, context="未检测到气泡"):
                        raise RuntimeError("点击发送后文案仍在输入框，发送按钮可能未命中")
                    add_step_event(item, "最新消息核对", "未通过", "未检测到已发送消息气泡", str(status_image))
                    raise RuntimeError(f"发送后未检测到已发送消息气泡；截图={status_image}")
                if item.parts["text"] and item.file_paths:
                    compact_ocr = " ".join(status_ocr.split())
                    add_step_event(item, "最新消息核对", "未通过", f"组合消息内容未匹配；OCR={compact_ocr[:140]}", str(status_image))
                    raise RuntimeError(f"发送后组合消息内容未匹配本次文本；OCR={compact_ocr[:140]}；截图={status_image}")
                if item.parts["text"] and not ocr_matches_message(item.message, status_ocr):
                    compact_ocr = " ".join(status_ocr.split())
                    add_step_event(item, "最新消息核对", "未通过", f"内容未匹配；OCR={compact_ocr[:140]}", str(status_image))
                    raise RuntimeError(f"发送后最新消息内容未匹配本次文本；OCR={compact_ocr[:140]}；截图={status_image}")
                if same_bubble_area(before_bubble, after_bubble) and not ocr_contains_time_marker(
                    status_ocr,
                    item.send_action_time,
                ):
                    add_step_event(item, "最新消息核对", "未通过", "未检测到新增消息气泡", str(status_image))
                    raise RuntimeError(f"发送后未检测到新增消息气泡；截图={status_image}")
                add_step_event(item, "最新消息核对", "已通过", "最新消息时间/内容/状态核对通过", str(status_image))
                return
            last_pending = pending
            if time.perf_counter() >= deadline:
                item.after_ocr_text += f"\nPENDING_TIMEOUT={last_pending}"
                add_step_event(item, "最新消息核对", "等待超时", last_pending, item.after_screenshot)
                raise RuntimeError(f"发送后界面仍处于待完成状态：{last_pending}；截图={item.after_screenshot}")
            add_step_event(item, "最新消息核对", "等待中", pending, str(status_image))
            time.sleep(self.send_settle_interval)

    def wait_for_send_settle_fast(
        self,
        item: SendRow,
        rect: WinRect,
        before_screenshot: Path | None = None,
        *,
        strict_new_bubble: bool = False,
    ) -> None:
        time.sleep(max(self.send_settle_interval, 0.25))
        item.after_screenshot = str(self.capture_window(item, rect, "after_send"))
        red_failure = screenshot_red_failure_indicator(Path(item.after_screenshot))
        if red_failure:
            item.after_ocr_text = red_failure
            add_step_event(item, "发送后红色失败标记核对", "未通过", red_failure, item.after_screenshot)
            raise RuntimeError(f"{red_failure}；截图={item.after_screenshot}")

        if getattr(self, "lean_evidence_fast", False):
            add_step_event(
                item,
                "发送后输入框残留核对",
                "快速跳过",
                "少证据快速模式：已截图并核对红色失败标记，跳过输入框残留 OCR",
                item.after_screenshot,
            )
        elif item.parts["text"] and not self.confirm_input_cleared_after_send(item, rect, context="快速核对"):
            raise RuntimeError("快速核对发现输入框仍有本次文案，未确认发送成功")

        before_bubble = latest_outgoing_bubble_from_path(before_screenshot) if before_screenshot else None
        after_bubble = latest_outgoing_bubble_from_path(Path(item.after_screenshot))
        if item.parts["text"] and after_bubble is None:
            add_step_event(item, "快速发送后核对", "未通过", "未检测到已发送消息气泡", item.after_screenshot)
            raise RuntimeError(f"快速核对未检测到已发送消息气泡；截图={item.after_screenshot}")
        if item.file_paths and before_bubble is not None and same_bubble_area(before_bubble, after_bubble):
            add_step_event(item, "快速发送后核对", "未通过", "组合消息未检测到新增消息气泡", item.after_screenshot)
            raise RuntimeError(f"快速核对未检测到新增消息气泡；截图={item.after_screenshot}")

        if before_bubble is not None and same_bubble_area(before_bubble, after_bubble):
            status_image = self.capture_latest_message_status(item, Path(item.after_screenshot))
            status_ocr = self.ocr_image(status_image)
            if latest_bubble_confirmed_new(before_bubble, after_bubble, status_ocr, item.send_action_time):
                detail = "输入框清空、未检测到失败标记，且最新消息时间/气泡变化已确认"
                add_step_event(item, "快速发送后核对", "已通过", detail, item.after_screenshot)
            else:
                if strict_new_bubble:
                    detail = "输入框清空，但未检测到新增消息气泡或本次发送时间；批量快速派发不放行"
                    add_step_event(item, "快速发送后核对", "预警待复核", detail, item.after_screenshot)
                detail = "输入框清空、未检测到失败标记；气泡位置未明显变化，按文本快速核验放行"
                add_step_event(item, "快速发送后核对", "预警放行", detail, item.after_screenshot)
        else:
            detail = "输入框清空、未检测到失败标记，且检测到右侧消息气泡"
            add_step_event(item, "快速发送后核对", "已通过", detail, item.after_screenshot)
        item.after_ocr_text = "FAST_SEND_OK"

    def file_wait_for(self, paths: list[Path]) -> float:
        total_bytes = 0
        for path in paths:
            try:
                total_bytes += path.stat().st_size
            except OSError:
                total_bytes += 0
        total_mb = total_bytes / (1024 * 1024)
        return min(self.file_wait_max, max(self.file_wait, self.file_wait + total_mb * self.file_wait_per_mb))

    def send_message(
        self,
        item: SendRow,
        wait_settle: bool = True,
        *,
        dispatch_confirm: bool = False,
    ) -> tuple[float, list[str]]:
        start = time.perf_counter()
        rect = self.last_rect if self.last_rect is not None else self.activate()
        lean_evidence_fast = getattr(self, "lean_evidence_fast", False)
        need_before_screenshot = ((wait_settle and not lean_evidence_fast) or dispatch_confirm) and not getattr(self, "blind_dispatch_fast", False)
        before_screenshot = self.capture_window(item, rect, "before_send") if self.capture_evidence and need_before_screenshot else None
        if before_screenshot:
            add_step_event(item, "发送前整窗截图", "已保存", "发送动作前完整窗口截图", str(before_screenshot))
        elif lean_evidence_fast:
            add_step_event(item, "发送前整窗截图", "快速跳过", "少证据快速模式：跳过发送前完整窗口截图")
        steps: list[str] = []
        file_paths = item.file_paths
        if item.parts["text"] and file_paths:
            self.send_text_with_files(rect, item, file_paths)
            steps.append("文字+附件")
            item.sent_steps = steps.copy()
        else:
            if item.parts["text"]:
                self.send_text(rect, item)
                steps.append("文字")
                item.sent_steps = steps.copy()
            if item.parts["image"]:
                self.send_files(rect, item, item.image_paths)
                steps.append(f"图片({len(item.image_paths)})")
                item.sent_steps = steps.copy()
            if item.parts["file"]:
                self.send_files(rect, item, item.document_paths)
                steps.append(f"文档({len(item.document_paths)})")
                item.sent_steps = steps.copy()
        time.sleep(self.between_rows)
        if wait_settle:
            self.wait_for_send_settle(item, rect, before_screenshot)
        elif dispatch_confirm:
            if getattr(self, "dispatch_only_fast", False):
                time.sleep(max(self.between_rows, 0.35))
                item.after_ocr_text = "DISPATCH_ONLY_FAST"
                add_step_event(
                    item,
                    "发送后核对",
                    "派发快速跳过",
                    "dispatch-only-fast：点击发送后短等待，不做发送后截图/气泡核对，发送后内容核对留待抽查或后续复核",
                )
            else:
                self.wait_for_send_settle_fast(item, rect, before_screenshot, strict_new_bubble=True)
                add_step_event(item, "发送后核对", "待统一复核", "派发即时核对已通过，发送后内容核对将在批量派发结束后统一执行", item.after_screenshot)
        else:
            add_step_event(item, "发送后核对", "待统一复核", "已点击发送，发送后核对将在批量派发结束后统一执行", item.after_screenshot)
        return time.perf_counter() - start, steps


class MacWeComGui(WeComGui):
    VK_BACK = "backspace"
    VK_DELETE = "delete"
    VK_CONTROL = "command"
    VK_ESCAPE = "esc"
    VK_SHIFT = "shift"
    VK_INSERT = "insert"
    VK_RETURN = "enter"
    VK_A = "a"
    VK_S = "s"
    VK_V = "v"

    def __init__(
        self,
        folder: Path,
        batch_id: str,
        run_dir: Path,
        wecom_exe: Path | None = None,
        verify_target: bool = True,
        verify_search_box: bool = True,
        capture_evidence: bool = True,
        low_evidence: bool = False,
        ocr_script: Path = OCR_SCRIPT_PATH,
        search_retries: int = 1,
        search_wait: float = 0.22,
        chat_wait: float = 0.22,
        text_wait: float = 0.22,
        file_wait: float = 0.45,
        file_wait_max: float = 1.25,
        file_wait_per_mb: float = 0.65,
        send_settle_timeout: float = 45.0,
        send_settle_interval: float = 0.6,
        between_rows: float = 0.02,
        paste_method_order: str = DEFAULT_PASTE_METHOD_ORDER,
        safe_fast: bool = False,
        trust_clipboard_paste: bool = False,
        fast_input_check: bool = False,
        lean_evidence_fast: bool = False,
        ultra_fast_dispatch: bool = False,
        dispatch_only_fast: bool = False,
        blind_dispatch_fast: bool = False,
        normalize_window_size: bool = False,
        window_width: int = 1092,
        window_height: int = 818,
        mac_app_name: str = "企业微信",
    ) -> None:
        if platform.system() != "Darwin":
            raise RuntimeError("MacWeComGui 仅支持 macOS")
        self.folder = folder
        self.batch_id = batch_id
        self.run_dir = run_dir
        self.wecom_exe = wecom_exe
        self.verify_target = verify_target
        self.verify_search_box = verify_search_box
        self.capture_evidence = capture_evidence
        self.low_evidence = low_evidence
        self.ocr_script = ocr_script
        self.search_retries = max(1, search_retries)
        self.search_wait = search_wait
        self.chat_wait = chat_wait
        self.text_wait = text_wait
        self.file_wait = file_wait
        self.file_wait_max = file_wait_max
        self.file_wait_per_mb = file_wait_per_mb
        self.send_settle_timeout = send_settle_timeout
        self.send_settle_interval = send_settle_interval
        self.between_rows = between_rows
        self.paste_method_order = paste_method_order
        self.safe_fast = safe_fast
        self.trust_clipboard_paste = trust_clipboard_paste
        self.fast_input_check = fast_input_check
        self.lean_evidence_fast = lean_evidence_fast
        self.ultra_fast_dispatch = ultra_fast_dispatch
        self.dispatch_only_fast = dispatch_only_fast
        self.blind_dispatch_fast = blind_dispatch_fast
        self.normalize_window_size = normalize_window_size
        self.normalized_window_width = max(620, int(window_width or 1092))
        self.normalized_window_height = max(360, int(window_height or 818))
        self._window_normalized_once = False
        self.current_target_key = ""
        self.confirmed_target_key = ""
        self.pending_search_result_key = ""
        self.pending_search_result_level = ""
        self.pending_search_result_detail = ""
        self.pending_before_open_title_screenshot = ""
        self.open_failure_cache: dict[str, str] = {}
        self.window_hwnd = 0
        self.last_rect: WinRect | None = None
        self.mac_app_name = cell_text(mac_app_name) or "企业微信"
        self.mac_app_candidates = [self.mac_app_name, "企业微信", "WeCom", "WXWork", "com.tencent.WeWorkMac"]
        self._active_mac_app = self.mac_app_name
        self._search_candidate_point: tuple[int, int] | None = None
        self._search_candidate_text = ""
        self._search_candidate_elements: list[Any] = []
        self._last_search_candidate_ax_text = ""
        self._last_search_candidate_ocr_text = ""
        try:
            import Quartz  # noqa: F401
            import Vision  # noqa: F401
            self.macos_vision_available = True
        except Exception:
            self.macos_vision_available = False

    def _run_osascript(self, script: str) -> str:
        completed = subprocess.run(
            ["osascript", "-e", script],
            capture_output=True,
            text=True,
            timeout=8,
        )
        if completed.returncode != 0:
            raise RuntimeError((completed.stderr or completed.stdout or "osascript 执行失败").strip())
        return cell_text(completed.stdout)

    def _window_bounds(self, app_name: str) -> tuple[int, int, int, int] | None:
        script = f'''
tell application "System Events"
    if not (exists process "{app_name}") then return ""
    tell process "{app_name}"
        set frontmost to true
        if (count of windows) = 0 then return ""
        set bestArea to 0
        set bestText to ""
        repeat with w in windows
            try
                set p to position of w
                set s to size of w
                set windowWidth to item 1 of s
                set windowHeight to item 2 of s
                set area to windowWidth * windowHeight
                if windowWidth >= 620 and windowHeight >= 360 and area > bestArea then
                    set bestArea to area
                    set bestText to (item 1 of p as text) & "," & (item 2 of p as text) & "," & (windowWidth as text) & "," & (windowHeight as text)
                end if
            end try
        end repeat
        if bestText is not "" then return bestText
        set p to position of front window
        set s to size of front window
        return (item 1 of p as text) & "," & (item 2 of p as text) & "," & (item 1 of s as text) & "," & (item 2 of s as text)
    end tell
end tell
'''
        try:
            output = self._run_osascript(script)
        except Exception:
            return None
        if not output:
            return None
        parts = [cell_text(part) for part in output.replace("\n", "").split(",")]
        if len(parts) != 4:
            return None
        try:
            left, top, width, height = [int(float(part)) for part in parts]
        except ValueError:
            return None
        if width < 620 or height < 360:
            return None
        return left, top, left + width, top + height

    def _set_front_window_bounds(self, app_name: str, left: int, top: int, width: int, height: int) -> bool:
        script = f'''
tell application "System Events"
    if not (exists process "{app_name}") then return "missing"
    tell process "{app_name}"
        set frontmost to true
        if (count of windows) = 0 then return "no-window"
        set position of front window to {{{int(left)}, {int(top)}}}
        delay 0.1
        set size of front window to {{{int(width)}, {int(height)}}}
        delay 0.1
        set position of front window to {{{int(left)}, {int(top)}}}
        delay 0.2
        set p to position of front window
        set s to size of front window
        return (item 1 of p as text) & "," & (item 2 of p as text) & "," & (item 1 of s as text) & "," & (item 2 of s as text)
    end tell
end tell
'''
        try:
            output = self._run_osascript(script)
        except Exception:
            return False
        return bool(output and output not in {"missing", "no-window"})

    def normalize_window_if_needed(self, stats: RunStats | None = None) -> WinRect:
        if not self.normalize_window_size:
            return self.activate()
        rect = self.activate()
        width = rect.right - rect.left
        height = rect.bottom - rect.top
        target_width = self.normalized_window_width
        target_height = self.normalized_window_height
        if self._window_normalized_once and abs(width - target_width) <= 8 and abs(height - target_height) <= 8:
            return rect
        screen_left, screen_top, screen_right, screen_bottom = self.virtual_screen_bounds()
        available_width = max(620, screen_right - screen_left)
        available_height = max(360, screen_bottom - screen_top)
        target_width = min(target_width, available_width - 20 if available_width > 640 else available_width)
        target_height = min(target_height, available_height - 40 if available_height > 400 else available_height)
        left = max(screen_left + 20, min(screen_left + 31, screen_right - target_width - 10))
        top = max(screen_top + 40, min(screen_top + 40, screen_bottom - target_height - 10))
        ok = self._set_front_window_bounds(self._active_mac_app, left, top, target_width, target_height)
        time.sleep(0.35)
        rect = self.activate()
        actual_width = rect.right - rect.left
        actual_height = rect.bottom - rect.top
        if abs(actual_width - target_width) > 8 or abs(actual_height - target_height) > 8 or abs(rect.top - top) > 8:
            ok = self._set_front_window_bounds(self._active_mac_app, left, top, target_width, target_height) or ok
            time.sleep(0.35)
            rect = self.activate()
        self._window_normalized_once = ok
        self.last_rect = rect
        self.current_target_key = ""
        self.confirmed_target_key = ""
        self.clear_input_locator_cache_entry()
        detail = f"目标={target_width}x{target_height}；实际={rect.right - rect.left}x{rect.bottom - rect.top}；位置={rect.left},{rect.top}"
        if stats is not None:
            log_line(stats, f"[{now_text()}] 企业微信窗口统一 | {detail}")
        return rect

    def _open_app(self, app_name: str) -> bool:
        command = ["open", "-b", app_name] if "." in app_name else ["open", "-a", app_name]
        completed = subprocess.run(
            command,
            capture_output=True,
            text=True,
            timeout=8,
        )
        return completed.returncode == 0

    def _frontmost_process_name(self) -> str:
        script = 'tell application "System Events" to get name of first application process whose frontmost is true'
        try:
            return self._run_osascript(script)
        except Exception:
            return ""

    def _frontmost_is_wecom(self) -> bool:
        return self._frontmost_process_name() in {"企业微信", "WeCom", "WXWork"}

    def _require_wecom_frontmost(self, context: str = "操作") -> None:
        frontmost = self._frontmost_process_name()
        if frontmost in {"企业微信", "WeCom", "WXWork"}:
            return
        for app_name in ["企业微信", "WeCom", "WXWork", self._active_mac_app, *self.mac_app_candidates]:
            if self._focus_app(app_name):
                return
        for _ in range(3):
            time.sleep(0.25)
            frontmost = self._frontmost_process_name()
            if frontmost in {"企业微信", "WeCom", "WXWork"}:
                return
        frontmost = self._frontmost_process_name()
        raise RuntimeError(f"{context}前台窗口不是企业微信，已停止避免误操作；当前前台={frontmost or '未知'}")

    def _visible_window_count(self, app_name: str) -> int:
        script = f'''
tell application "System Events"
    if not (exists process "{app_name}") then return "-1"
    tell process "{app_name}"
        return count of windows
    end tell
end tell
'''
        try:
            output = self._run_osascript(script)
            return int(output) if output else 0
        except Exception:
            return 0

    def _focus_app(self, app_name: str) -> bool:
        script = f'''
try
    tell application "{app_name}" to activate
end try
delay 0.15
tell application "System Events"
    if not (exists process "{app_name}") then return "missing"
    tell process "{app_name}"
        set frontmost to true
    end tell
end tell
'''
        try:
            self._run_osascript(script)
            time.sleep(0.35)
        except Exception:
            pass
        if self._frontmost_is_wecom():
            return True
        if self._open_app(app_name):
            time.sleep(0.8)
        return self._frontmost_is_wecom()

    def activate(self) -> WinRect:
        quartz_window = self._quartz_main_window()
        if quartz_window is not None:
            if not self._focus_app(self._active_mac_app):
                raise RuntimeError(
                    f"企业微信窗口存在但未能置前，已停止避免误操作；当前前台={self._frontmost_process_name() or '未知'}"
                )
            self._require_wecom_frontmost("激活企业微信")
            _window_id, bounds = quartz_window
            rect = WinRect()
            rect.left, rect.top, rect.right, rect.bottom = bounds
            self.last_rect = rect
            return rect

        for app_name in self.mac_app_candidates:
            self._focus_app(app_name)
            window_count = self._visible_window_count(app_name)
            if window_count == 0:
                continue
            bounds = self._window_bounds(app_name)
            if bounds:
                self._active_mac_app = app_name
                break
        else:
            launched = False
            for app_name in self.mac_app_candidates:
                if self._open_app(app_name):
                    self._active_mac_app = app_name
                    launched = True
                    break
            if not launched:
                raise RuntimeError("未找到企业微信应用，请通过 --mac-app-name 指定应用名（例如 企业微信 / WeCom）")
            deadline = time.perf_counter() + 10
            bounds = None
            while time.perf_counter() < deadline:
                self._focus_app(self._active_mac_app)
                if self._visible_window_count(self._active_mac_app) == 0:
                    time.sleep(0.2)
                    continue
                bounds = self._window_bounds(self._active_mac_app)
                if bounds:
                    break
                time.sleep(0.2)
            if not bounds:
                raise GuiUnavailableError(
                    "企业微信没有可操作主窗口：系统能看到企业微信进程，但当前桌面没有可见窗口。"
                    "请手动打开企业微信主窗口到当前桌面后再运行。"
                )
        rect = WinRect()
        rect.left, rect.top, rect.right, rect.bottom = bounds
        self._require_wecom_frontmost("激活企业微信")
        self.last_rect = rect
        return rect

    def virtual_screen_bounds(self) -> tuple[int, int, int, int]:
        try:
            from PIL import ImageGrab
        except Exception as exc:
            raise RuntimeError(f"无法读取屏幕尺寸：{exc}") from exc
        width, height = ImageGrab.grab().size
        return 0, 0, int(width), int(height)

    def input_point(self, rect: WinRect) -> tuple[int, int]:
        cached = self.cached_input_locator(rect)
        if cached and "来源=mac_default" not in cached[1]:
            return cached[0]
        width = rect.right - rect.left
        # macOS WeCom keeps the editor close to the bottom of the chat window.
        # Older fixed offsets such as bottom-185 can land in the message area on
        # shorter windows, which focuses a message bubble instead of the editor.
        point = (rect.left + max(360, min(width // 2 + 160, width - 260)), rect.bottom - 58)
        self.remember_input_locator(rect, point, "mac_bottom_editor")
        return point

    def input_text_bbox(self, rect: WinRect) -> tuple[int, int, int, int]:
        return rect.left + 300, rect.bottom - 145, rect.right - 30, rect.bottom - 24

    def point_on_screen(self, point: tuple[int, int], margin: int = 8) -> bool:
        left, top, right, bottom = self.virtual_screen_bounds()
        x, y = point
        return left + margin <= x <= right - margin and top + margin <= y <= bottom - margin

    def click(self, x: int, y: int) -> None:
        self.activate()
        try:
            import pyautogui

            pyautogui.click(x=int(x), y=int(y), duration=0.02)
            time.sleep(0.02)
            return
        except Exception as pyautogui_exc:
            pyautogui_error = pyautogui_exc
        try:
            import Quartz

            point = Quartz.CGPointMake(int(x), int(y))
            down = Quartz.CGEventCreateMouseEvent(
                None,
                Quartz.kCGEventLeftMouseDown,
                point,
                Quartz.kCGMouseButtonLeft,
            )
            up = Quartz.CGEventCreateMouseEvent(
                None,
                Quartz.kCGEventLeftMouseUp,
                point,
                Quartz.kCGMouseButtonLeft,
            )
            Quartz.CGEventPost(Quartz.kCGHIDEventTap, down)
            time.sleep(0.015)
            Quartz.CGEventPost(Quartz.kCGHIDEventTap, up)
        except Exception as exc:
            raise RuntimeError(f"macOS 鼠标点击失败：坐标=({x},{y})；pyautogui={pyautogui_error}；Quartz={exc}") from exc
        time.sleep(0.02)

    def key_down(self, vk: str) -> None:
        _ = vk

    def key_up(self, vk: str) -> None:
        _ = vk

    def press(self, vk: str) -> None:
        key_codes = {
            self.VK_BACK: 51,
            self.VK_DELETE: 117,
            self.VK_ESCAPE: 53,
            self.VK_RETURN: 36,
        }
        if vk in key_codes:
            self._run_osascript(f'tell application "System Events" to key code {key_codes[vk]}')
            return
        if len(vk) == 1:
            self._run_osascript(f'tell application "System Events" to keystroke "{vk}"')
            return
        raise RuntimeError(f"macOS 暂不支持按键：{vk}")

    def hotkey(self, *keys: str) -> None:
        if keys == (self.VK_CONTROL, self.VK_A):
            self._run_osascript('tell application "System Events" to keystroke "a" using command down')
            return
        if keys == (self.VK_CONTROL, self.VK_V):
            self._run_osascript('tell application "System Events" to keystroke "v" using command down')
            return
        if keys == (self.VK_SHIFT, self.VK_INSERT):
            self._run_osascript('tell application "System Events" to key code 114 using shift down')
            return
        raise RuntimeError(f"macOS 暂不支持快捷键：{'+'.join(keys)}")

    def paste_clipboard(self) -> None:
        self.hotkey(self.VK_CONTROL, self.VK_V)

    def paste_clipboard_to_point(self, x: int, y: int) -> bool:
        _ = (x, y)
        return False

    def paste_clipboard_to_main_window(self) -> bool:
        return False

    def clear_input_box(self, rect: WinRect) -> None:
        input_x, input_y = self.input_point(rect)
        if not self.point_on_screen((input_x, input_y)):
            return
        self.click(input_x, input_y)
        time.sleep(0.08)
        self.hotkey(self.VK_CONTROL, self.VK_A)
        self.press(self.VK_BACK)
        time.sleep(0.08)

    def set_text_clipboard(self, text: str) -> None:
        completed = subprocess.run(
            ["pbcopy"],
            input=text,
            capture_output=True,
            text=True,
            timeout=6,
        )
        if completed.returncode != 0:
            raise RuntimeError((completed.stderr or completed.stdout or "pbcopy 写入失败").strip())
        time.sleep(0.05)

    def set_text_clipboard_system(self, text: str) -> None:
        self.set_text_clipboard(text)

    def get_text_clipboard(self) -> str:
        completed = subprocess.run(
            ["pbpaste"],
            capture_output=True,
            text=True,
            timeout=6,
        )
        if completed.returncode != 0:
            return ""
        return completed.stdout

    def paste_methods(self, input_x: int, input_y: int) -> list[tuple[str, Any]]:
        _ = (input_x, input_y)
        return [("Cmd+V", lambda: (self.paste_clipboard() or True))] * 3

    def set_file_clipboard(self, paths: list[Path]) -> None:
        if not paths:
            return
        try:
            from AppKit import NSPasteboard, NSURL
        except Exception as exc:
            raise RuntimeError(f"macOS 附件剪贴板不可用：缺少 AppKit；详情：{exc}") from exc
        urls = []
        missing = [str(path) for path in paths if not path.exists()]
        if missing:
            raise RuntimeError(f"附件文件不存在：{', '.join(missing)}")
        for path in paths:
            urls.append(NSURL.fileURLWithPath_(str(path.resolve())))
        pasteboard = NSPasteboard.generalPasteboard()
        pasteboard.clearContents()
        if not pasteboard.writeObjects_(urls):
            raise RuntimeError("macOS 附件剪贴板写入失败")
        time.sleep(0.08)

    def submit_message(self, rect: WinRect, item: SendRow, label: str) -> None:
        input_x, input_y = self.input_point(rect)
        if not self.point_on_screen((input_x, input_y)):
            raise RuntimeError(f"输入框坐标不可见，无法聚焦后 Enter 发送：{input_x},{input_y}")
        self.click(input_x, input_y)
        time.sleep(0.12)
        item.send_action_time = now_text()
        self.press(self.VK_RETURN)
        add_step_event(item, "按 Enter 发送", "已执行", f"{label}；发送前已重新聚焦输入框；坐标={input_x},{input_y}")

    def launch_wecom(self) -> None:
        if not self._open_app(self._active_mac_app):
            raise RuntimeError(f"无法启动企业微信应用：{self._active_mac_app}")

    def _active_process_pid(self) -> int | None:
        script = f'''
tell application "System Events"
    if not (exists process "{self._active_mac_app}") then return ""
    return unix id of process "{self._active_mac_app}"
end tell
'''
        try:
            output = self._run_osascript(script)
            return int(output) if output else None
        except Exception:
            return None

    def _quartz_main_window(self) -> tuple[int, tuple[int, int, int, int]] | None:
        try:
            import Quartz
        except Exception:
            return None
        owner_names = {self._active_mac_app, "WeCom", "企业微信", "WXWork"}
        try:
            windows = Quartz.CGWindowListCopyWindowInfo(Quartz.kCGWindowListOptionAll, Quartz.kCGNullWindowID)
        except Exception:
            return None
        if not windows:
            return None
        screen_left, screen_top, screen_right, screen_bottom = self.virtual_screen_bounds()
        best: tuple[int, tuple[int, int, int, int], int] | None = None
        for window in windows:
            owner = cell_text(window.get("kCGWindowOwnerName", ""))
            if owner not in owner_names:
                continue
            try:
                layer = int(window.get("kCGWindowLayer", 99))
            except Exception:
                layer = 99
            if layer != 0:
                continue
            bounds = window.get("kCGWindowBounds") or {}
            try:
                left = int(float(bounds.get("X", 0)))
                top = int(float(bounds.get("Y", 0)))
                width = int(float(bounds.get("Width", 0)))
                height = int(float(bounds.get("Height", 0)))
                window_id = int(window.get("kCGWindowNumber"))
            except Exception:
                continue
            if width < 620 or height < 360:
                continue
            area = width * height
            visible_width = max(0, min(left + width, screen_right) - max(left, screen_left))
            visible_height = max(0, min(top + height, screen_bottom) - max(top, screen_top))
            visible_area = visible_width * visible_height
            if visible_area < area * 0.6:
                continue
            name = cell_text(window.get("kCGWindowName", ""))
            score = visible_area + (1_000_000 if name in {"WeCom", "企业微信"} else 0)
            if best is None or score > best[2]:
                best = (window_id, (left, top, left + width, top + height), score)
        if best is None:
            return None
        return best[0], best[1]

    def _ax_rect_for_element(self, element: Any) -> tuple[float, float, float, float] | None:
        try:
            from ApplicationServices import AXUIElementCopyAttributeValue, AXValueGetValue, kAXValueCGRectType
        except Exception:
            return None
        try:
            err, value = AXUIElementCopyAttributeValue(element, "AXFrame", None)
            if err != 0 or value is None:
                return None
            ok, rect = AXValueGetValue(value, kAXValueCGRectType, None)
            if not ok:
                return None
            left = float(rect.origin.x)
            top = float(rect.origin.y)
            return left, top, left + float(rect.size.width), top + float(rect.size.height)
        except Exception:
            return None

    def set_search_box_text(self, rect: WinRect, text: str) -> bool:
        search_x, search_y = self.search_point(rect)
        try:
            self.click(search_x, search_y)
            time.sleep(0.12)
            self.hotkey(self.VK_CONTROL, self.VK_A)
            time.sleep(0.04)
            self.press(self.VK_BACK)
            time.sleep(0.08)
            self.set_verified_text_clipboard(text, reliable=True)
            self.hotkey(self.VK_CONTROL, self.VK_V)
            time.sleep(max(self.search_wait, 0.35))
            return True
        except Exception:
            return False

    def _mac_search_candidate_bounds(self, rect: WinRect) -> tuple[int, int, int, int]:
        return (
            rect.left + 75,
            rect.top + 86,
            min(rect.left + 470, rect.right - 12),
            min(rect.top + 470, rect.bottom - 90),
        )

    def search_results_bbox(self, rect: WinRect) -> tuple[int, int, int, int]:
        return (
            rect.left + 70,
            rect.top + 62,
            min(rect.left + 470, rect.right - 12),
            min(rect.top + 470, rect.bottom - 90),
        )

    def _find_search_candidate_ax(self, item: SendRow, rect: WinRect, query: str) -> tuple[str, tuple[int, int], str, list[Any]] | None:
        try:
            from ApplicationServices import AXUIElementCopyAttributeValue, AXUIElementCreateApplication
        except Exception:
            return None
        try:
            self.activate()
            pid = self._active_process_pid()
            if not pid:
                return None
            app = AXUIElementCreateApplication(pid)
            roots: list[Any] = []
            err, focused = AXUIElementCopyAttributeValue(app, "AXFocusedWindow", None)
            if err == 0 and focused is not None:
                roots.append(focused)
            err, windows = AXUIElementCopyAttributeValue(app, "AXWindows", None)
            if err == 0 and windows is not None:
                try:
                    roots.extend(list(windows))
                except Exception:
                    pass
            roots.append(app)
        except Exception:
            return None

        left, top, right, bottom = self._mac_search_candidate_bounds(rect)
        candidates = target_ocr_candidates(item.target, [*item.target_aliases, query], include_variants=False)
        text_attrs = ("AXValue", "AXTitle", "AXDescription", "AXHelp")
        child_attrs = (
            "AXChildren",
            "AXChildrenInNavigationOrder",
            "AXVisibleChildren",
            "AXContents",
            "AXRows",
            "AXSelectedChildren",
        )
        seen_elements: set[str] = set()
        matches: list[tuple[float, float, str, tuple[int, int], str, list[Any]]] = []
        ax_text_nodes: list[tuple[float, float, float, float, str, str, Any, list[Any]]] = []
        diagnostic_texts: list[str] = []
        deadline = time.perf_counter() + 1.5

        def remember_diagnostic_text(text: str) -> None:
            text = " ".join(cell_text(text).split())
            if text and text not in diagnostic_texts:
                diagnostic_texts.append(text[:180])

        def in_candidate_area(ax_rect: tuple[float, float, float, float] | None) -> bool:
            if ax_rect is None:
                return False
            x1, y1, x2, y2 = ax_rect
            if x2 < left or x1 > right or y2 < top or y1 > bottom:
                return False
            if y2 - y1 > 180 or x2 - x1 > 720:
                return False
            return True

        def children_for(element: Any, attr: str) -> list[Any]:
            try:
                err, value = AXUIElementCopyAttributeValue(element, attr, None)
                if err != 0 or not value:
                    return []
                return list(value)
            except Exception:
                return []

        def subtree_text(element: Any, max_nodes: int = 80) -> str:
            collected: list[str] = []
            seen: set[str] = set()

            def add(value: str) -> None:
                value = cell_text(value)
                if value and value not in collected:
                    collected.append(value)

            def walk_text(node: Any) -> None:
                if len(seen) >= max_nodes:
                    return
                key = str(node)
                if key in seen:
                    return
                seen.add(key)
                for attr in text_attrs:
                    add(self._ax_value_for_element(node, attr))
                for attr in child_attrs:
                    for child in children_for(node, attr):
                        walk_text(child)

            walk_text(element)
            return "\n".join(collected)

        def remember_text_node(
            ax_rect: tuple[float, float, float, float],
            text: str,
            element: Any,
            parents: list[Any],
        ) -> None:
            normalized = "\n".join(dict.fromkeys(part for part in (cell_text(line) for line in text.splitlines()) if part))
            if not normalized or normalized == "WeCom" or re.fullmatch(r"[0-9.]+", normalized):
                return
            role = self._ax_value_for_element(element, "AXRole")
            ax_text_nodes.append((*ax_rect, normalized, role, element, parents[-5:]))
            remember_diagnostic_text(normalized)

        def strict_match_text(text: str) -> str:
            return ocr_contains_any_target(candidates, text) if text else ""

        def row_match_text(text: str) -> str:
            matched = ocr_contains_any_target(candidates, text) if text else ""
            if not matched and text:
                matched = ocr_contains_any_search_result_target(candidates, text)
            return matched

        def match_structured_rows() -> tuple[str, tuple[int, int], str, list[Any]] | None:
            if not ax_text_nodes:
                return None
            rows: list[dict[str, Any]] = []
            for x1, y1, x2, y2, text, role, element, parents in sorted(ax_text_nodes, key=lambda node: (node[1], node[0])):
                if y2 - y1 > 125:
                    continue
                center_y = (y1 + y2) / 2
                row: dict[str, Any] | None = None
                for existing in rows:
                    existing_center = (existing["top"] + existing["bottom"]) / 2
                    vertical_overlap = min(y2, existing["bottom"]) - max(y1, existing["top"])
                    if abs(center_y - existing_center) <= 34 or vertical_overlap >= min(y2 - y1, existing["bottom"] - existing["top"]) * 0.45:
                        row = existing
                        break
                if row is None:
                    row = {"top": y1, "bottom": y2, "left": x1, "right": x2, "nodes": []}
                    rows.append(row)
                row["top"] = min(row["top"], y1)
                row["bottom"] = max(row["bottom"], y2)
                row["left"] = min(row["left"], x1)
                row["right"] = max(row["right"], x2)
                row["nodes"].append((x1, y1, x2, y2, text, role, element, parents))

            for index, row in enumerate(sorted(rows, key=lambda value: (value["top"], value["left"]))[:6], start=1):
                nodes = sorted(row["nodes"], key=lambda node: (node[0], node[1]))
                all_parts = [node[4] for node in nodes]
                name_parts = [
                    node[4]
                    for node in nodes
                    if node[0] <= left + 310 and not re.fullmatch(r"[0-9:/： .-]+", node[4])
                ]
                name_text = "\n".join(dict.fromkeys(name_parts))
                row_text = "\n".join(dict.fromkeys(all_parts))
                matched = row_match_text(name_text)
                match_scope = "name"
                if not matched:
                    matched = row_match_text(row_text)
                    match_scope = "row"
                if not matched:
                    continue
                non_chat_marker = search_non_chat_result_marker(row_text)
                if match_scope == "row" and non_chat_marker:
                    remember_diagnostic_text(f"跳过疑似非会话结果({non_chat_marker})：{row_text[:120]}")
                    continue
                click_x = int(max(left + 35, min(row["left"] + min(max(row["right"] - row["left"], 130), 260), right - 25)))
                click_y = int(max(top + 10, min((row["top"] + row["bottom"]) / 2, bottom - 10)))
                action_elements: list[Any] = []
                for _x1, _y1, _x2, _y2, _text, _role, element, parents in nodes:
                    action_elements.extend([element, *reversed(parents)])
                structured_text = (
                    f"AX-ROW row={index};scope={match_scope};"
                    f"name={name_text[:120]};text={row_text[:220]}"
                )
                return matched, (click_x, click_y), structured_text, action_elements
            return None

        def walk(element: Any, parents: list[Any] | None = None) -> None:
            parents = parents or []
            if time.perf_counter() > deadline or len(seen_elements) > 1500:
                return
            key = str(element)
            if key in seen_elements:
                return
            seen_elements.add(key)
            ax_rect = self._ax_rect_for_element(element)
            if in_candidate_area(ax_rect):
                collected: list[str] = []
                for attr in text_attrs:
                    text = self._ax_value_for_element(element, attr)
                    if text:
                        collected.append(text)
                direct_text = "\n".join(dict.fromkeys(collected))
                text = direct_text or subtree_text(element)
                if direct_text:
                    remember_text_node(ax_rect, direct_text, element, parents)
                    child_text = subtree_text(element, max_nodes=40)
                    if child_text:
                        text = "\n".join(dict.fromkeys([direct_text, *child_text.splitlines()]))
                remember_diagnostic_text(text)
                matched = strict_match_text(text)
                if matched and ax_rect is not None:
                    if search_non_chat_result_marker(text):
                        remember_diagnostic_text(f"跳过疑似非会话结果：{text[:120]}")
                    else:
                        x1, y1, x2, y2 = ax_rect
                        click_anchor = x1 + min(max((x2 - x1) / 2, 80), 220)
                        click_x = int(max(left + 35, min(click_anchor, right - 25)))
                        click_y = int(max(top + 10, min((y1 + y2) / 2, bottom - 10)))
                        action_elements = [element, *reversed(parents[-5:])]
                        matches.append((y1, x1, matched, (click_x, click_y), text, action_elements))
            for attr in child_attrs:
                for child in children_for(element, attr):
                    walk(child, [*parents, element])

        for root in roots:
            walk(root)
        structured_match = match_structured_rows()
        if structured_match:
            matched, point, text, elements = structured_match
            self._last_search_candidate_ax_text = " | ".join(diagnostic_texts[:8])
            return matched, point, text, elements
        if not matches:
            self._last_search_candidate_ax_text = " | ".join(diagnostic_texts[:8])
            return None
        matches.sort(key=lambda item_match: (item_match[0], item_match[1]))
        _y, _x, matched, point, text, elements = matches[0]
        self._last_search_candidate_ax_text = " | ".join(diagnostic_texts[:8])
        return matched, point, text, elements

    def confirm_first_search_result(self, item: SendRow, rect: WinRect, query: str) -> None:
        self._search_candidate_point = None
        self._search_candidate_text = ""
        self._search_candidate_elements = []
        self._search_candidate_near = False
        self._last_search_candidate_ax_text = ""
        self._last_search_candidate_ocr_text = ""
        if not self.verify_target:
            return
        screenshot = self.capture_search_results(item, rect)
        candidate = self._find_search_candidate_ax(item, rect, query)
        if candidate:
            matched, point, text, elements = candidate
            self._search_candidate_point = point
            self._search_candidate_text = text
            self._search_candidate_elements = elements
            compact_text = " ".join(text.split())
            add_step_event(
                item,
                "搜索结果核对",
                "已通过",
                f"macOS AX 候选匹配：预期={item.target}；命中={matched}；坐标={point[0]},{point[1]}；文本={compact_text[:160]}",
                str(screenshot),
            )
            self.remember_search_result_confirmation(item, "matched", f"命中={matched}；AX={compact_text[:160]}")
            return
        visual_candidate = self._find_search_candidate_by_row_ocr(item, rect, query, screenshot)
        if visual_candidate:
            matched, point, text, near = visual_candidate
            self._search_candidate_point = point
            self._search_candidate_text = text
            self._search_candidate_near = near
            compact_text = " ".join(text.split())
            level = "near" if near else "matched"
            status_text = "搜索结果截图分行疑似匹配" if near else "搜索结果截图分行匹配"
            add_step_event(
                item,
                "搜索结果核对",
                "待标题强核对" if near else "已通过",
                (
                    f"{status_text}：预期={item.target}；命中={matched}；坐标={point[0]},{point[1]}；"
                    f"OCR={compact_text[:160]}；"
                    f"{'将打开候选后做目标标题强核对' if near else '候选已确认'}"
                ),
                str(screenshot),
            )
            self.remember_search_result_confirmation(item, level, f"命中={matched}；OCR={compact_text[:160]}")
            return
        add_step_event(
            item,
            "搜索结果核对",
            "未命中跳过",
            (
                f"预期={item.target}；macOS AX/OCR 未在搜索结果列表找到匹配候选，跳过发送；"
                f"AX摘要={brief_log_text(self._last_search_candidate_ax_text, 260) or '无'}；"
                f"OCR摘要={brief_log_text(self._last_search_candidate_ocr_text, 320) or '无'}"
            ),
            str(screenshot),
        )
        raise RuntimeError(f"搜索结果未命中目标名，已跳过发送：预期={item.target}；截图={screenshot}")

    def _find_search_candidate_by_row_ocr(
        self,
        item: SendRow,
        rect: WinRect,
        query: str,
        screenshot: Path,
    ) -> tuple[str, tuple[int, int], str, bool] | None:
        try:
            from PIL import Image
        except Exception:
            return None
        candidates = target_ocr_candidates(item.target, [*item.target_aliases, query], include_variants=False)
        left, top, right, bottom = self.search_results_bbox(rect)
        try:
            image = Image.open(screenshot).convert("RGB")
        except Exception:
            return None
        gap = 2
        fast_mode = getattr(self, "lean_evidence_fast", False)
        row_heights = (108, 96) if fast_mode else (72, 84, 96, 108)
        max_rows = 3
        seen_crops: set[tuple[int, int]] = set()
        ocr_diagnostics: list[str] = []

        def remember_ocr_diagnostic(label: str, text: str) -> None:
            text = " ".join(cell_text(text).split())
            if text:
                ocr_diagnostics.append(f"{label}={text[:160]}")
                self._last_search_candidate_ocr_text = " | ".join(ocr_diagnostics[-12:])

        def match_text(text: str) -> str:
            matched = ocr_contains_any_target(candidates, text) if text else ""
            if not matched and text:
                matched = ocr_contains_any_search_result_target(candidates, text)
            return matched

        def near_match_text(text: str) -> str:
            if not text:
                return ""
            return ocr_search_result_short_tail_target(candidates, text)

        def candidate_point_for_row(y1: int, y2: int) -> tuple[int, int]:
            click_x = int(left + min(max((right - left) * 0.38, 130), 260))
            click_y = int(top + (y1 + y2) / 2)
            return click_x, click_y

        for row_height in row_heights:
            for index in range(max_rows):
                y1 = index * row_height
                if y1 >= image.height:
                    break
                y2 = min(image.height, y1 + row_height - gap)
                if y2 - y1 < 38:
                    break
                crop_key = (y1, y2)
                if crop_key in seen_crops:
                    continue
                seen_crops.add(crop_key)
                row_image = image.crop((0, y1, image.width, y2))
                row_label = f"row{index + 1}/h{row_height}"
                ocr_attempts: list[tuple[str, str]] = []

                def run_row_crop(label: str, crop_box: tuple[int, int, int, int]) -> str:
                    crop_left, crop_top, crop_right, crop_bottom = crop_box
                    if crop_right - crop_left < 24 or crop_bottom - crop_top < 24:
                        return ""
                    crop_path = screenshot.with_name(f"{screenshot.stem}_{row_label.replace('/', '_')}_{label}_ocr.png")
                    try:
                        row_image.crop(crop_box).save(crop_path)
                        crop_text = self.run_ocr_file(crop_path)
                    except Exception:
                        crop_text = ""
                    remember_ocr_diagnostic(f"{row_label}/{label}", crop_text)
                    if crop_text:
                        ocr_attempts.append((label, crop_text))
                    return crop_text

                name_crops: list[tuple[str, tuple[int, int, int, int]]] = []
                if image.width > 160:
                    name_crops.append(("name_tight", (min(70, image.width - 1), 0, min(image.width, 340), row_image.height)))
                if image.width > 120 and not fast_mode:
                    name_crops.append(("name_medium", (min(42, image.width - 1), 0, min(image.width, 390), row_image.height)))
                if image.width > 90 and not fast_mode:
                    name_crops.append(("name_wide", (min(42, image.width - 1), 0, min(image.width, 520), row_image.height)))

                for label, crop_box in name_crops:
                    name_text = run_row_crop(label, crop_box)
                    matched = match_text(name_text)
                    if matched:
                        return matched, candidate_point_for_row(y1, y2), f"row={index + 1};height={row_height};crop={label}\n{name_text}", False

                combined_name_text = "\n".join(dict.fromkeys(text for _label, text in ocr_attempts))
                matched = near_match_text(combined_name_text)
                if matched:
                    return matched, candidate_point_for_row(y1, y2), f"row={index + 1};height={row_height};crop=name_near\n{combined_name_text}", True

                if not fast_mode:
                    row_text = run_row_crop("row", (0, 0, row_image.width, row_image.height))
                    matched = match_text(row_text)
                    if matched:
                        non_chat_marker = search_non_chat_result_marker(row_text)
                        if non_chat_marker:
                            remember_ocr_diagnostic(f"{row_label}/skip_non_chat", f"{non_chat_marker} {row_text[:140]}")
                            continue
                        return matched, candidate_point_for_row(y1, y2), f"row={index + 1};height={row_height};crop=row\n{row_text}", False
        if fast_mode:
            row_heights = (72, 84)
            for row_height in row_heights:
                for index in range(max_rows):
                    y1 = index * row_height
                    if y1 >= image.height:
                        break
                    y2 = min(image.height, y1 + row_height - gap)
                    if y2 - y1 < 38:
                        break
                    crop_key = (y1, y2)
                    if crop_key in seen_crops:
                        continue
                    seen_crops.add(crop_key)
                    row_image = image.crop((0, y1, image.width, y2))
                    row_label = f"row{index + 1}/h{row_height}"
                    crop_box = (min(70, image.width - 1), 0, min(image.width, 340), row_image.height)
                    crop_path = screenshot.with_name(f"{screenshot.stem}_{row_label.replace('/', '_')}_name_tight_ocr.png")
                    try:
                        row_image.crop(crop_box).save(crop_path)
                        name_text = self.run_ocr_file(crop_path)
                    except Exception:
                        name_text = ""
                    remember_ocr_diagnostic(f"{row_label}/name_tight", name_text)
                    matched = match_text(name_text)
                    if matched:
                        return matched, candidate_point_for_row(y1, y2), f"row={index + 1};height={row_height};crop=name_tight\n{name_text}", False
        return None

    def _press_search_candidate_ax(self) -> bool:
        if not self._search_candidate_elements:
            return False
        try:
            from ApplicationServices import AXUIElementPerformAction
        except Exception:
            return False
        for element in self._search_candidate_elements:
            for action in ("AXPress", "AXConfirm"):
                try:
                    if AXUIElementPerformAction(element, action) == 0:
                        time.sleep(max(self.chat_wait, 0.45))
                        return True
                except Exception:
                    pass
        return False

    def _click_confirmed_search_candidate_point(self) -> bool:
        if self._search_candidate_point is None:
            return False
        x, y = self._search_candidate_point
        if not self.point_on_screen((x, y), margin=4):
            return False
        try:
            self.click(x, y)
            time.sleep(max(self.chat_wait, 0.45))
            return True
        except Exception:
            return False

    def _click_confirmed_search_candidate_variant(self, rect: WinRect, variant: str) -> bool:
        if self._search_candidate_point is None:
            return False
        left, _top, right, _bottom = self.search_results_bbox(rect)
        _x, y = self._search_candidate_point
        offsets = {
            "left": left + 92,
            "middle": left + 180,
            "right": min(right - 45, left + 310),
        }
        x = int(offsets.get(variant, _x))
        if not self.point_on_screen((x, y), margin=4):
            return False
        try:
            self.click(x, y)
            time.sleep(max(self.chat_wait, 0.45))
            return True
        except Exception:
            return False

    def _confirm_target_after_open_with_recovery(
        self,
        item: SendRow,
        rect: WinRect,
        query: str,
        opened_by: str,
    ) -> WinRect:
        try:
            self.confirm_target(item, rect)
            return rect
        except Exception as first_error:
            add_step_event(
                item,
                "打开搜索结果",
                "标题未确认准备补救",
                f"输入={query}；方式={opened_by}；首次标题核对失败={first_error}",
                item.target_screenshot,
            )
            last_error: Exception = first_error

        recovery_methods = (
            "二次坐标点击候选",
            "候选左侧点击",
            "候选中部点击",
            "候选右侧点击",
            "Enter打开候选",
        )
        for recovery_method in recovery_methods:
            self.remember_title_before_open(item, rect)
            try:
                if recovery_method == "二次坐标点击候选":
                    if not self._click_confirmed_search_candidate_point():
                        continue
                elif recovery_method == "候选左侧点击":
                    if not self._click_confirmed_search_candidate_variant(rect, "left"):
                        continue
                elif recovery_method == "候选中部点击":
                    if not self._click_confirmed_search_candidate_variant(rect, "middle"):
                        continue
                elif recovery_method == "候选右侧点击":
                    if not self._click_confirmed_search_candidate_variant(rect, "right"):
                        continue
                else:
                    self.press(self.VK_RETURN)
                    time.sleep(max(self.chat_wait, 0.55))
                add_step_event(
                    item,
                    "打开搜索结果",
                    "已执行补救",
                    f"输入={query}；方式={opened_by}->{recovery_method}；文本={' '.join(self._search_candidate_text.split())[:120]}",
                )
                rect = self.activate()
                self.confirm_target(item, rect)
                add_step_event(
                    item,
                    "打开搜索结果",
                    "补救后已通过",
                    f"输入={query}；方式={opened_by}->{recovery_method}；目标标题已确认",
                    item.target_screenshot,
                )
                return rect
            except Exception as exc:
                last_error = exc
                add_step_event(
                    item,
                    "打开搜索结果",
                    "补救未确认",
                    f"输入={query}；方式={opened_by}->{recovery_method}；错误={exc}",
                    item.target_screenshot,
                )
        raise last_error

    def open_search_result_with_enter(self, item: SendRow, rect: WinRect, query: str) -> WinRect:
        if getattr(self, "blind_dispatch_fast", False):
            self.pending_search_result_key = item.cache_key
            self.pending_search_result_level = "enter"
            self.pending_search_result_detail = f"输入={query}；方式=盲发Enter；跳过搜索结果/标题确认"
            self.pending_open_attempt_method = "enter"
            add_step_event(
                item,
                "打开搜索结果",
                "盲发Enter",
                f"输入={query}；跳过搜索结果OCR和标题确认，短等待后进入粘贴发送",
            )
            self.press(self.VK_RETURN)
            time.sleep(max(self.chat_wait, 0.65))
            rect = self.activate()
            item.targetcheck_status = "盲发跳过"
            item.targetcheck_detail = "blind-dispatch-fast：未做目标标题确认"
            self.current_target_key = item.cache_key
            self.confirmed_target_key = item.cache_key
            return rect
        if getattr(self, "lean_evidence_fast", False):
            self.remember_title_before_open(item, rect)
            self.pending_search_result_key = item.cache_key
            self.pending_search_result_level = "enter"
            self.pending_search_result_detail = f"输入={query}；方式=快速Enter主路径；未做搜索结果OCR"
            self.pending_open_attempt_method = "enter"
            add_step_event(
                item,
                "打开搜索结果",
                "快速Enter主路径",
                f"输入={query}；跳过搜索结果OCR，Enter后做目标标题核对",
            )
            self.press(self.VK_RETURN)
            time.sleep(max(self.chat_wait, 0.45))
            rect = self.activate()
            try:
                self.confirm_target(item, rect)
                add_step_event(item, "打开搜索结果", "已通过", f"输入={query}；方式=快速Enter主路径；目标标题已确认")
                return rect
            except Exception as enter_error:
                add_step_event(
                    item,
                    "打开搜索结果",
                    "快速Enter未确认转候选核对",
                    f"输入={query}；Enter后目标标题未确认={enter_error}",
                    item.target_screenshot,
                )
                self.clear_search_result_confirmation()
                self._search_candidate_point = None
                self._search_candidate_text = ""
                self._search_candidate_elements = []
                self._search_candidate_near = False
                rect = self.activate()
                self.press(self.VK_ESCAPE)
                time.sleep(0.08)
                search_x, search_y = self.search_point(rect)
                self.click(search_x, search_y)
                time.sleep(0.06)
                self.set_verified_text_clipboard(query, reliable=True)
                self.hotkey(self.VK_CONTROL, self.VK_A)
                time.sleep(0.03)
                self.hotkey(self.VK_CONTROL, self.VK_V)
                time.sleep(max(self.search_wait, 0.25))

        try:
            self.confirm_first_search_result(item, rect, query)
        except Exception as candidate_error:
            self.remember_title_before_open(item, rect)
            self.pending_search_result_key = item.cache_key
            self.pending_search_result_level = "enter"
            self.pending_search_result_detail = f"输入={query}；方式=候选未确认回退Enter；候选错误={candidate_error}"
            self.pending_open_attempt_method = "enter"
            add_step_event(
                item,
                "打开搜索结果",
                "候选未确认回退Enter",
                f"输入={query}；候选错误={candidate_error}",
            )
            self.press(self.VK_RETURN)
            time.sleep(max(self.chat_wait, 0.45))
            rect = self.activate()
            self.confirm_target(item, rect)
            add_step_event(item, "打开搜索结果", "已通过", f"输入={query}；方式=Enter回退；目标标题已确认")
            return rect

        self.remember_title_before_open(item, rect)
        self.pending_open_attempt_method = "enter"
        opened_by = "Enter打开候选"
        if self._search_candidate_near:
            opened_by = f"{opened_by}(疑似候选)"
        add_step_event(
            item,
            "打开搜索结果",
            "已执行",
            f"输入={query}；方式={opened_by}；文本={' '.join(self._search_candidate_text.split())[:120]}",
        )
        self.press(self.VK_RETURN)
        time.sleep(max(self.chat_wait, 0.45))
        rect = self.activate()
        return self._confirm_target_after_open_with_recovery(item, rect, query, opened_by)

    def first_search_result_point(self, rect: WinRect) -> tuple[int, int]:
        _ = rect
        if self._search_candidate_point is not None:
            return self._search_candidate_point
        return super().first_search_result_point(rect)

    def _ax_value_for_element(self, element: Any, attr: str) -> str:
        try:
            from ApplicationServices import AXUIElementCopyAttributeValue
        except Exception:
            return ""
        try:
            err, value = AXUIElementCopyAttributeValue(element, attr, None)
            if err != 0 or value is None:
                return ""
            return cell_text(value)
        except Exception:
            return ""

    def _accessibility_text_in_bbox_osascript(self, bbox: tuple[int, int, int, int]) -> str:
        left, top, right, bottom = [int(value) for value in bbox]
        script = f'''
on appendLine(currentText, nextText)
    if nextText is "" then return currentText
    if currentText is "" then return nextText
    return currentText & linefeed & nextText
end appendLine

on collectText(elementRef, leftBound, topBound, rightBound, bottomBound, depth)
    if depth > 9 then return ""
    set overlapsBounds to true
    try
        set p to position of elementRef
        set s to size of elementRef
        set x1 to item 1 of p
        set y1 to item 2 of p
        set x2 to x1 + (item 1 of s)
        set y2 to y1 + (item 2 of s)
        set overlapsBounds to not (x2 < leftBound or x1 > rightBound or y2 < topBound or y1 > bottomBound)
    end try
    if overlapsBounds is false then return ""

    set outputText to ""
    try
        set outputText to my appendLine(outputText, (name of elementRef as text))
    end try
    try
        set outputText to my appendLine(outputText, (value of elementRef as text))
    end try
    try
        set outputText to my appendLine(outputText, (description of elementRef as text))
    end try
    try
        repeat with childRef in UI elements of elementRef
            set outputText to my appendLine(outputText, my collectText(childRef, leftBound, topBound, rightBound, bottomBound, depth + 1))
        end repeat
    end try
    return outputText
end collectText

tell application "System Events"
    if not (exists process "{self._active_mac_app}") then return ""
    tell process "{self._active_mac_app}"
        set frontmost to true
        if (count of windows) = 0 then return ""
        return my collectText(front window, {left}, {top}, {right}, {bottom}, 0)
    end tell
end tell
'''
        try:
            output = self._run_osascript(script)
        except Exception:
            return ""
        texts: list[str] = []
        seen: set[str] = set()
        for raw_line in output.splitlines():
            text = cell_text(raw_line)
            if not text or text == "WeCom" or re.fullmatch(r"[0-9.]+", text):
                continue
            if text not in seen:
                seen.add(text)
                texts.append(text)
        return "\n".join(texts)

    def accessibility_text_in_bbox(self, bbox: tuple[int, int, int, int]) -> str:
        try:
            from ApplicationServices import AXUIElementCopyAttributeValue, AXUIElementCreateApplication
        except Exception:
            return self._accessibility_text_in_bbox_osascript(bbox)

        try:
            self.activate()
            pid = self._active_process_pid()
            if not pid:
                return ""
            app = AXUIElementCreateApplication(pid)
            err, root = AXUIElementCopyAttributeValue(app, "AXFocusedWindow", None)
            roots = [root] if err == 0 and root is not None else []
            err, windows = AXUIElementCopyAttributeValue(app, "AXWindows", None)
            if err == 0 and windows is not None:
                try:
                    roots.extend(list(windows))
                except Exception:
                    pass
            roots.append(app)
        except Exception:
            return ""

        text_attrs = ("AXValue", "AXTitle", "AXDescription", "AXPlaceholderValue")
        child_attrs = ("AXChildren", "AXChildrenInNavigationOrder", "AXVisibleChildren", "AXContents")
        bbox_f = tuple(float(value) for value in bbox)
        texts: list[str] = []
        seen_texts: set[str] = set()
        seen_elements: set[str] = set()
        deadline = time.perf_counter() + 1.5

        def overlaps(rect: tuple[float, float, float, float] | None) -> bool:
            if rect is None:
                return False
            return not (rect[2] < bbox_f[0] or rect[0] > bbox_f[2] or rect[3] < bbox_f[1] or rect[1] > bbox_f[3])

        def add_text(value: str) -> None:
            text = cell_text(value)
            if not text or text == "WeCom":
                return
            if re.fullmatch(r"[0-9.]+", text):
                return
            if text not in seen_texts:
                seen_texts.add(text)
                texts.append(text)

        def children_for(element: Any, attr: str) -> list[Any]:
            try:
                err, value = AXUIElementCopyAttributeValue(element, attr, None)
                if err != 0 or not value:
                    return []
                return list(value)
            except Exception:
                return []

        def walk(element: Any) -> None:
            if time.perf_counter() > deadline:
                return
            if len(seen_elements) > 1200:
                return
            key = str(element)
            if key in seen_elements:
                return
            seen_elements.add(key)
            if overlaps(self._ax_rect_for_element(element)):
                for attr in text_attrs:
                    add_text(self._ax_value_for_element(element, attr))
            for attr in child_attrs:
                if texts and time.perf_counter() > deadline - 0.3:
                    return
                for child in children_for(element, attr):
                    walk(child)

        for root in roots:
            walk(root)
        return "\n".join(texts)

    def capture_bbox(self, bbox: tuple[int, int, int, int], path: Path) -> Path:
        left, top, right, bottom = bbox
        if right <= left or bottom <= top:
            raise RuntimeError(f"截图区域无效：{bbox}")
        last_error = ""
        for attempt in range(3):
            try:
                rect = self.activate()
                self._require_wecom_frontmost("截图")
                window_info = self._quartz_main_window()
                if window_info is not None:
                    try:
                        import Quartz
                        from PIL import Image
                    except Exception as exc:
                        last_error = str(exc)
                        window_info = None
                    else:
                        window_id, window_bounds = window_info
                        image = Quartz.CGWindowListCreateImage(
                            Quartz.CGRectNull,
                            Quartz.kCGWindowListOptionIncludingWindow,
                            window_id,
                            Quartz.kCGWindowImageBoundsIgnoreFraming,
                        )
                        if image is not None:
                            path.parent.mkdir(parents=True, exist_ok=True)
                            tmp_path = path.with_name(f"{path.stem}_window_full{path.suffix}")
                            url_bytes = str(tmp_path).encode()
                            url = Quartz.CFURLCreateFromFileSystemRepresentation(None, url_bytes, len(url_bytes), False)
                            dest = Quartz.CGImageDestinationCreateWithURL(url, "public.png", 1, None)
                            Quartz.CGImageDestinationAddImage(dest, image, None)
                            if Quartz.CGImageDestinationFinalize(dest):
                                full = Image.open(tmp_path)
                                win_left, win_top, win_right, win_bottom = window_bounds
                                scale_x = full.width / max(1, win_right - win_left)
                                scale_y = full.height / max(1, win_bottom - win_top)
                                crop = (
                                    max(0, int(round((left - win_left) * scale_x))),
                                    max(0, int(round((top - win_top) * scale_y))),
                                    min(full.width, int(round((right - win_left) * scale_x))),
                                    min(full.height, int(round((bottom - win_top) * scale_y))),
                                )
                                if crop[2] > crop[0] and crop[3] > crop[1]:
                                    full.crop(crop).save(path)
                                    try:
                                        tmp_path.unlink()
                                    except OSError:
                                        pass
                                    return path
                        last_error = "Quartz 未返回可用窗口图像"
                else:
                    try:
                        from PIL import ImageGrab

                        path.parent.mkdir(parents=True, exist_ok=True)
                        ImageGrab.grab(bbox=bbox).convert("RGB").save(path)
                        return path
                    except Exception as exc:
                        last_error = f"未找到企业微信主窗口；前台截图兜底失败={exc}"
            except Exception as exc:
                last_error = str(exc)
            if attempt < 2:
                time.sleep(0.5)
        detail = f"；最后错误={last_error}" if last_error else ""
        raise RuntimeError(f"macOS 企业微信窗口截图失败，已重试3次，停止避免截取其他前台窗口{detail}")

    def title_bbox(self, rect: WinRect) -> tuple[int, int, int, int]:
        width = rect.right - rect.left
        left = rect.left + min(max(int(width * 0.18), 300), 390)
        top = rect.top + 24
        right = min(rect.right - 220, left + 520)
        bottom = rect.top + 92
        if right <= left + 120:
            right = rect.right - 20
        return left, top, right, bottom

    def title_name_bbox(self, rect: WinRect) -> tuple[int, int, int, int]:
        left, top, right, _bottom = self.title_bbox(rect)
        return left, top, min(right, left + 380), top + 40

    def selected_conversation_bbox(self, rect: WinRect) -> tuple[int, int, int, int]:
        return (
            rect.left + 70,
            rect.top + 58,
            min(rect.left + 430, rect.right - 20),
            max(rect.top + 120, rect.bottom - 45),
        )

    def selected_conversation_text(self, item: SendRow, rect: WinRect) -> tuple[str, str]:
        try:
            from PIL import Image
        except Exception:
            return "", ""
        bbox = self.selected_conversation_bbox(rect)
        full_path = self.evidence_path(item, "selected_chat_list")
        try:
            captured = self.record_evidence(item, "selected_chat_list", self.capture_bbox(bbox, full_path), bbox)
            image = Image.open(captured).convert("RGB")
        except Exception:
            return "", ""

        width, height = image.size
        if width < 80 or height < 40:
            return "", str(captured)
        row_scores: list[int] = []
        for y in range(height):
            count = 0
            for x in range(width):
                r, g, b = image.getpixel((x, y))
                if b >= 170 and 75 <= g <= 180 and r <= 120:
                    count += 1
            row_scores.append(count)

        threshold = max(8, int(width * 0.18))
        segments: list[tuple[int, int, int]] = []
        start: int | None = None
        score_sum = 0
        for y, score in enumerate(row_scores):
            if score >= threshold:
                if start is None:
                    start = y
                    score_sum = 0
                score_sum += score
            elif start is not None:
                if y - start >= 24:
                    segments.append((score_sum, start, y))
                start = None
                score_sum = 0
        if start is not None and height - start >= 24:
            segments.append((score_sum, start, height))
        if not segments:
            return "", str(captured)

        _score, y1, y2 = max(segments, key=lambda segment: (segment[0], segment[2] - segment[1]))
        row_top = max(0, y1 - 8)
        row_bottom = min(height, y2 + 8)
        row_image = image.crop((0, row_top, width, row_bottom))
        row_path = captured.with_name(f"{captured.stem}_selected_row.png")
        try:
            row_image.save(row_path)
            abs_bbox = (bbox[0], bbox[1] + row_top, bbox[2], bbox[1] + row_bottom)
            self.record_evidence(item, "selected_chat", row_path, abs_bbox)
        except Exception:
            pass

        texts: list[str] = []
        crop_lefts = (0, min(52, max(width - 1, 0)), min(78, max(width - 1, 0)))
        for index, crop_left in enumerate(crop_lefts):
            if crop_left >= width - 20:
                continue
            crop = row_image.crop((crop_left, 0, width, row_image.height))
            scale = 6 if crop.height < 120 else 3
            variant = crop.resize((crop.width * scale, crop.height * scale), Image.Resampling.LANCZOS)
            variant_path = captured.with_name(f"{captured.stem}_selected_row_ocr{index}.png")
            try:
                variant.save(variant_path)
                text = self.run_ocr_file(variant_path)
            except Exception:
                text = ""
            text = cell_text(text)
            if text:
                texts.append(text)

        ax_text = self.accessibility_text_in_bbox((bbox[0], bbox[1] + row_top, bbox[2], bbox[1] + row_bottom))
        if ax_text:
            texts.append(ax_text)
        seen: set[str] = set()
        cleaned: list[str] = []
        for text in texts:
            for line in text.splitlines():
                value = cell_text(line)
                if value and value not in seen:
                    seen.add(value)
                    cleaned.append(value)
        return "\n".join(cleaned), str(row_path)

    def capture_title_name(self, item: SendRow, rect: WinRect) -> Path:
        path = self.evidence_path(item, "target_title_name")
        bbox = self.title_name_bbox(rect)
        return self.record_evidence(item, "target_title_name", self.capture_bbox(bbox, path), bbox)

    def run_ocr_file(self, image_path: Path) -> str:
        if self.macos_vision_available and MACOS_VISION_OCR_SCRIPT_PATH.exists():
            try:
                completed = subprocess.run(
                    [sys.executable, str(MACOS_VISION_OCR_SCRIPT_PATH), str(image_path)],
                    capture_output=True,
                    text=True,
                    timeout=12,
                )
                if completed.returncode == 0 and completed.stdout.strip():
                    return completed.stdout.strip()
            except subprocess.TimeoutExpired:
                pass
        if self.ocr_script.exists() and self.ocr_script.suffix.lower() == ".py":
            completed = subprocess.run(
                [sys.executable, str(self.ocr_script), str(image_path)],
                capture_output=True,
                text=True,
                timeout=12,
            )
            if completed.returncode == 0:
                return completed.stdout.strip()
        errors: list[str] = []
        tesseract = tesseract_command()
        if tesseract is None:
            raise OcrUnavailableError("OCR不可用：macOS Vision OCR无结果，且未安装 tesseract，当前行无法通过 OCR 核对")
        for lang in ("chi_sim+eng", "chi_sim", "eng"):
            completed = subprocess.run(
                [tesseract, str(image_path), "stdout", "-l", lang, "--psm", "6"],
                capture_output=True,
                text=True,
                timeout=12,
            )
            if completed.returncode == 0:
                return completed.stdout.strip()
            errors.append((completed.stderr or completed.stdout or f"tesseract OCR失败({lang})").strip())
        raise RuntimeError(
            "OCR执行失败。请先安装 tesseract：`brew install tesseract tesseract-lang`；"
            f"详细错误：{' | '.join(errors[-2:])}"
        )

    def run_ocr_file_quick(self, image_path: Path, timeout_seconds: float = 3.0) -> str:
        timeout = max(0.8, float(timeout_seconds))
        if self.macos_vision_available and MACOS_VISION_OCR_SCRIPT_PATH.exists():
            try:
                completed = subprocess.run(
                    [sys.executable, str(MACOS_VISION_OCR_SCRIPT_PATH), str(image_path)],
                    capture_output=True,
                    text=True,
                    timeout=timeout,
                )
                if completed.returncode == 0 and completed.stdout.strip():
                    return completed.stdout.strip()
            except subprocess.TimeoutExpired:
                return ""
            except Exception:
                pass
        if self.ocr_script.exists() and self.ocr_script.suffix.lower() == ".py":
            try:
                completed = subprocess.run(
                    [sys.executable, str(self.ocr_script), str(image_path)],
                    capture_output=True,
                    text=True,
                    timeout=timeout,
                )
                if completed.returncode == 0:
                    return completed.stdout.strip()
            except subprocess.TimeoutExpired:
                return ""
            except Exception:
                pass
        tesseract = tesseract_command()
        if tesseract is None:
            return ""
        try:
            completed = subprocess.run(
                [tesseract, str(image_path), "stdout", "-l", "chi_sim+eng", "--psm", "6"],
                capture_output=True,
                text=True,
                timeout=timeout,
            )
            if completed.returncode == 0:
                return completed.stdout.strip()
        except Exception:
            return ""
        return ""


def path_is_inside(path: Path, folder: Path) -> bool:
    try:
        path.resolve().relative_to(folder.resolve())
        return True
    except ValueError:
        return False


def is_generated_send_workbook(path: Path) -> bool:
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return False
    lower_name = path.name.lower()
    return any(lower_name.startswith(prefix.lower()) for prefix in GENERATED_SEND_WORKBOOK_PREFIXES)


def unique_path(folder: Path, filename: str) -> Path:
    candidate = folder / filename
    if not candidate.exists():
        return candidate
    stem = candidate.stem
    suffix = candidate.suffix
    for index in range(2, 1000):
        renamed = folder / f"{stem}_{index:02d}{suffix}"
        if not renamed.exists():
            return renamed
    return folder / f"{stem}_{timestamp_slug()}{suffix}"


def archive_generated_workbook_to_run_dir(stats: RunStats, workbook_path: Path) -> tuple[Path, str]:
    if stats.run_dir is None:
        return workbook_path, ""
    source = workbook_path.resolve()
    if path_is_inside(source, stats.run_dir) or not is_generated_send_workbook(source) or not source.exists():
        return workbook_path, ""
    destination = unique_path(stats.run_dir, source.name)
    try:
        shutil.move(str(source), str(destination))
        return destination, "已移动"
    except OSError as move_error:
        try:
            shutil.copy2(str(source), str(destination))
            return destination, f"已复制，源文件保留：{move_error}"
        except OSError as copy_error:
            return workbook_path, f"归档失败：{copy_error}"


def ordered_rows(rows: list[SendRow], group_targets: bool = False) -> list[SendRow]:
    sendable = [row for row in rows if row.should_send]
    if not group_targets:
        return sendable
    grouped: dict[str, list[SendRow]] = {}
    order: list[str] = []
    for row in sendable:
        if row.cache_key not in grouped:
            grouped[row.cache_key] = []
            order.append(row.cache_key)
        grouped[row.cache_key].append(row)
    return [row for key in order for row in grouped[key]]


def brief_log_text(value: str, limit: int = 90) -> str:
    text = " ".join(cell_text(value).split())
    text = re.sub(r"[；;]\s*截图=.*$", "", text)
    text = re.sub(r"区域=\([^)]+\)-\([^)]+\)", "最新消息红色失败标记", text)
    if len(text) <= limit:
        return text
    return text[: limit - 1] + "…"


def build_gui(store: WorkbookStore, args: argparse.Namespace, stats: RunStats) -> WeComGui:
    base_kwargs: dict[str, Any] = {
        "folder": store.folder,
        "batch_id": stats.batch_id,
        "run_dir": stats.run_dir or store.folder / LOG_ROOT_RELATIVE_PATH / stats.batch_id,
        "wecom_exe": Path(args.wecom_exe) if args.wecom_exe else None,
        "verify_target": not args.skip_target_check,
        "verify_search_box": not args.skip_search_box_check,
        "capture_evidence": not args.no_evidence,
        "low_evidence": args.low_evidence,
        "ocr_script": Path(args.ocr_script).resolve(),
        "search_retries": args.search_retries,
        "search_wait": args.search_wait,
        "chat_wait": args.chat_wait,
        "text_wait": args.text_wait,
        "file_wait": args.file_wait,
        "file_wait_max": args.file_wait_max,
        "file_wait_per_mb": args.file_wait_per_mb,
        "send_settle_timeout": args.send_settle_timeout,
        "send_settle_interval": args.send_settle_interval,
        "between_rows": args.between_rows,
        "paste_method_order": args.paste_method_order,
        "safe_fast": args.safe_fast,
        "trust_clipboard_paste": args.trust_clipboard_paste,
        "fast_input_check": args.fast_input_check,
        "lean_evidence_fast": args.lean_evidence_fast,
        "ultra_fast_dispatch": args.ultra_fast_dispatch,
        "dispatch_only_fast": args.dispatch_only_fast,
        "blind_dispatch_fast": args.blind_dispatch_fast,
        "normalize_window_size": args.normalize_window_size,
        "window_width": args.window_width,
        "window_height": args.window_height,
    }
    if platform.system() == "Darwin":
        base_kwargs["mac_app_name"] = args.mac_app_name
        return MacWeComGui(**base_kwargs)
    return WeComGui(**base_kwargs)


def row_result_payload(item: SendRow, status: str, error: str = "") -> dict[str, Any]:
    payload: dict[str, Any] = {
        "row": item.row,
        "target": item.target,
        "status": status,
        "schedule_status": item.schedule_status,
        "schedule_detail": item.schedule_detail,
        "send_action_time": item.send_action_time,
        "precheck": item.precheck_status,
        "targetcheck": item.targetcheck_status,
        "postcheck": item.postcheck_status,
        "sent_steps": item.sent_steps,
        "target_screenshot": item.target_screenshot,
        "before_screenshot": item.before_screenshot,
        "after_screenshot": item.after_screenshot,
        "step_events": item.step_events,
        "evidence_files": item.evidence_files,
    }
    if item.total_seconds:
        payload["seconds"] = round(item.total_seconds, 3)
    if error:
        payload["error"] = error
    return payload


def mark_rows_blocked_by_wecom_state(
    store: WorkbookStore,
    rows: Iterable[SendRow],
    stats: RunStats,
    mode: str,
    reason: str,
    row_status: str = "发送失败",
    result_status: str = "failed",
) -> list[dict[str, Any]]:
    results: list[dict[str, Any]] = []
    for item in rows:
        item.precheck_status = item.precheck_status or "未执行"
        item.precheck_time = item.precheck_time or now_text()
        item.precheck_detail = item.precheck_detail or "企业微信状态阻断，本行未执行发送前核对"
        item.targetcheck_status = item.targetcheck_status or "未执行"
        item.targetcheck_time = item.targetcheck_time or now_text()
        item.targetcheck_detail = item.targetcheck_detail or "企业微信状态阻断，未打开目标会话"
        item.postcheck_status = item.postcheck_status or "未执行"
        item.postcheck_time = item.postcheck_time or now_text()
        item.postcheck_detail = item.postcheck_detail or reason
        add_step_event(item, "批次阻断", "未发送", reason)
        stats.failed += 1
        store.mark_row(item, row_status, reason, stats, mode)
        results.append(row_result_payload(item, result_status, reason))
    return results


def execute_rows(
    store: WorkbookStore,
    rows: list[SendRow],
    args: argparse.Namespace,
    stats: RunStats,
) -> list[dict[str, Any]]:
    gui = build_gui(store, args, stats)
    results: list[dict[str, Any]] = []
    mode = "fast-grouped" if args.group_targets else "fast-row-order"
    if not args.skip_target_check:
        mode += "+target-ocr"
    if not args.skip_search_box_check:
        mode += "+search-ocr"
    if args.respect_send_time:
        mode += "+schedule"
    if args.blind_dispatch_fast:
        mode += "+blind"
    if args.normalize_window_size:
        mode += "+normalized-window"

    consecutive_ocr_no_text = 0
    ordered = ordered_rows(rows, group_targets=args.group_targets)
    if args.normalize_window_size:
        gui.normalize_window_if_needed(stats)
    for index, item in enumerate(ordered):
        row_start = time.perf_counter()
        try:
            verify_before_send(item)
            planned_send_at = None
            if args.respect_send_time and item.scheduled_at:
                planned_send_at = scheduled_datetime_for_item(item)
                log_line(
                    stats,
                    (
                        f"[{now_text()}] 行{item.row} | {item.target} | 计划发送准备 | "
                        f"计划={planned_send_at.strftime('%Y-%m-%d %H:%M:%S')}"
                    ),
                )
                wait_for_schedule_prepare(
                    planned_send_at,
                    args.schedule_wait_max_minutes,
                    args.schedule_prepare_lead_seconds,
                )
                add_step_event(item, "计划发送准备", "已完成", f"计划={planned_send_at.strftime('%Y-%m-%d %H:%M:%S')}")
            if args.assume_current_target:
                gui.confirm_target(item, gui.activate())
                gui.current_target_key = item.cache_key
                item.search_seconds = 0.0
            else:
                item.search_seconds = gui.open_target(item)
            if planned_send_at is not None:
                wait_until_before_scheduled_time(planned_send_at, args.schedule_final_target_check_lead_seconds)
                gui.confirm_target(item, gui.activate())
                add_step_event(
                    item,
                    "计划发送前目标复核",
                    "已完成",
                    f"计划前{args.schedule_final_target_check_lead_seconds:.1f}s完成目标会话复核",
                    item.target_screenshot,
                )
                wait_until_scheduled_time(planned_send_at)
                add_step_event(item, "计划发送到点", "已到达", f"计划={planned_send_at.strftime('%Y-%m-%d %H:%M:%S')}")
            item.send_seconds, item.sent_steps = gui.send_message(item)
            if planned_send_at is not None:
                item.schedule_status, item.schedule_detail = classify_scheduled_send_action_time(
                    planned_send_at,
                    args.schedule_on_time_tolerance_seconds,
                    item.send_action_time,
                )
                add_step_event(item, "计划发送时间核对", item.schedule_status, item.schedule_detail)
                log_line(stats, f"[{now_text()}] 行{item.row} | {item.target} | {item.schedule_status} | {item.schedule_detail}")
            verify_after_send(item)
            item.total_seconds = time.perf_counter() - row_start
            stats.success += 1
            row_status = "已发送" if not args.resend else "已重发"
            row_error = ""
            if item.schedule_status == "超时已发送":
                stats.late_success += 1
                row_status = "超时已发送"
                row_error = item.schedule_detail
            deleted_evidence = gui.cleanup_row_evidence(item)
            store.mark_row(item, row_status, row_error, stats, mode)
            log_line(
                stats,
                f"[{now_text()}] 行{item.row} | {item.target} | {row_status} | {item.total_seconds:.3f}s",
            )
            results.append(
                {
                    "row": item.row,
                    "target": item.target,
                    "status": "late_success" if item.schedule_status == "超时已发送" else "success",
                    "seconds": round(item.total_seconds, 3),
                    "schedule_status": item.schedule_status,
                    "schedule_detail": item.schedule_detail,
                    "send_action_time": item.send_action_time,
                    "precheck": item.precheck_status,
                    "targetcheck": item.targetcheck_status,
                    "postcheck": item.postcheck_status,
                    "sent_steps": item.sent_steps,
                    "target_screenshot": item.target_screenshot,
                    "before_screenshot": item.before_screenshot,
                    "after_screenshot": item.after_screenshot,
                    "step_events": item.step_events,
                    "evidence_files": item.evidence_files,
                    "low_evidence_deleted_files": deleted_evidence,
                }
            )
            consecutive_ocr_no_text = 0
        except Exception as exc:
            item.total_seconds = time.perf_counter() - row_start
            if not item.postcheck_status:
                item.postcheck_status = "未执行" if item.precheck_status == "未通过" else "未通过"
                item.postcheck_time = now_text()
                item.postcheck_detail = str(exc)
            runtime_skip = is_runtime_skip_error(str(exc)) and not item.sent_steps
            if runtime_skip:
                stats.skipped += 1
                status = "已跳过"
            else:
                stats.failed += 1
                status = "发送失败"
            deleted_evidence = gui.cleanup_row_evidence(item)
            store.mark_row(item, status, str(exc), stats, mode)
            sent_steps = " -> ".join(item.sent_steps) if item.sent_steps else "无"
            log_line(
                stats,
                (
                    f"[{now_text()}] 行{item.row} | {item.target} | {status} | "
                    f"步骤={sent_steps} | 原因={brief_log_text(str(exc))}"
                ),
            )
            results.append(
                {
                    "row": item.row,
                    "target": item.target,
                    "status": "skipped" if runtime_skip else "failed",
                    "error": str(exc),
                    "schedule_status": item.schedule_status,
                    "schedule_detail": item.schedule_detail,
                    "send_action_time": item.send_action_time,
                    "precheck": item.precheck_status,
                    "targetcheck": item.targetcheck_status,
                    "postcheck": item.postcheck_status,
                    "sent_steps": item.sent_steps,
                    "target_screenshot": item.target_screenshot,
                    "before_screenshot": item.before_screenshot,
                    "after_screenshot": item.after_screenshot,
                    "step_events": item.step_events,
                    "evidence_files": item.evidence_files,
                    "low_evidence_deleted_files": deleted_evidence,
                }
            )
            gui.current_target_key = ""
            gui.confirmed_target_key = ""
            if isinstance(exc, WeComBlockedError):
                blocked_reason = f"企业微信状态阻断，本轮已停止，后续对象未发送；首个阻断原因={exc}"
                log_line(stats, f"[{now_text()}] 批次阻断 | {brief_log_text(blocked_reason)}")
                results.extend(mark_rows_blocked_by_wecom_state(store, ordered[index + 1 :], stats, mode, blocked_reason))
                if args.save_each_row:
                    store.save()
                break
            if isinstance(exc, GuiUnavailableError):
                log_line(stats, f"[{now_text()}] 企业微信窗口不可用，已停止本批次，避免继续刷失败记录")
                if args.save_each_row:
                    store.save()
                break
            if isinstance(exc, OcrUnavailableError):
                log_line(stats, f"[{now_text()}] OCR不可用，已停止本批次，避免继续刷失败记录")
                if args.save_each_row:
                    store.save()
                break
            if isinstance(exc, OcrNoTextError):
                consecutive_ocr_no_text += 1
                if args.max_consecutive_ocr_no_text > 0 and consecutive_ocr_no_text >= args.max_consecutive_ocr_no_text:
                    log_line(stats, f"[{now_text()}] OCR连续{consecutive_ocr_no_text}条无文字结果，已停止本批次")
                    if args.save_each_row:
                        store.save()
                    break
            else:
                consecutive_ocr_no_text = 0
            if args.stop_on_error:
                break
        if args.save_each_row:
            store.save()

    return results


def execute_rows_mvp(
    store: WorkbookStore,
    rows: list[SendRow],
    args: argparse.Namespace,
    stats: RunStats,
) -> list[dict[str, Any]]:
    """Minimal send loop: search target -> send -> mark success/failure."""
    gui = build_gui(store, args, stats)
    results: list[dict[str, Any]] = []
    mode = "mvp-simple"
    if args.group_targets:
        mode += "+grouped"

    ordered = ordered_rows(rows, group_targets=args.group_targets)
    for index, item in enumerate(ordered):
        row_start = time.perf_counter()
        try:
            item.precheck_status = "已跳过"
            item.precheck_time = now_text()
            item.precheck_detail = "MVP 模式：跳过发送前核对"
            item.targetcheck_status = "已跳过"
            item.targetcheck_time = now_text()
            item.targetcheck_detail = "MVP 模式：跳过目标 OCR 核对"
            item.search_seconds = gui.open_target(item)

            send_start = time.perf_counter()
            item.send_seconds, item.sent_steps = gui.send_message(item, wait_settle=False)
            time.sleep(max(0.5, float(args.mvp_wait_after_send)))
            item.postcheck_status = "已跳过"
            item.postcheck_time = now_text()
            item.postcheck_detail = f"MVP 模式：点击发送后等待 {max(0.5, float(args.mvp_wait_after_send)):.2f}s"

            item.total_seconds = time.perf_counter() - row_start
            item.send_seconds = max(item.send_seconds, time.perf_counter() - send_start)
            stats.success += 1
            row_status = "已发送" if not args.resend else "已重发"
            store.mark_row(item, row_status, "", stats, mode)
            log_line(
                stats,
                f"[{now_text()}] 行{item.row} | {item.target} | {row_status} | MVP简化流程 | {item.total_seconds:.3f}s",
            )
            results.append(row_result_payload(item, "success"))
        except Exception as exc:
            item.total_seconds = time.perf_counter() - row_start
            if not item.postcheck_status:
                item.postcheck_status = "未执行"
                item.postcheck_time = now_text()
                item.postcheck_detail = f"MVP 模式异常：{exc}"
            stats.failed += 1
            store.mark_row(item, "发送失败", str(exc), stats, mode)
            log_line(
                stats,
                (
                    f"[{now_text()}] 行{item.row} | {item.target} | 发送失败 | "
                    f"MVP简化流程 | 原因={brief_log_text(str(exc))}"
                ),
            )
            results.append(row_result_payload(item, "failed", str(exc)))
            gui.current_target_key = ""
            gui.confirmed_target_key = ""
            if isinstance(exc, WeComBlockedError):
                blocked_reason = f"企业微信状态阻断，本轮已停止，后续对象未发送；首个阻断原因={exc}"
                log_line(stats, f"[{now_text()}] 批次阻断 | {brief_log_text(blocked_reason)}")
                results.extend(mark_rows_blocked_by_wecom_state(store, ordered[index + 1 :], stats, mode, blocked_reason))
                break
            if args.stop_on_error:
                break
        if args.save_each_row:
            store.save()

    return results


def execute_rows_batch_fast_dispatch(
    store: WorkbookStore,
    rows: list[SendRow],
    args: argparse.Namespace,
    stats: RunStats,
) -> list[dict[str, Any]]:
    gui = build_gui(store, args, stats)
    mode = "batch-fast-dispatch"
    if args.group_targets:
        mode += "+grouped"
    if not args.skip_target_check:
        mode += "+target-ocr"
    if not args.skip_search_box_check:
        mode += "+search-ocr"
    if args.respect_send_time:
        mode += "+schedule"
    if args.normalize_window_size:
        mode += "+normalized-window"

    ordered = ordered_rows(rows, group_targets=args.group_targets)
    dispatched: list[tuple[SendRow, float]] = []
    result_by_row: dict[int, dict[str, Any]] = {}
    phase_start = time.perf_counter()

    if args.normalize_window_size:
        gui.normalize_window_if_needed(stats)
    review_note = "跳过统一复核，派发后待人工确认" if args.blind_dispatch_fast else "发送后统一复核"
    log_line(stats, f"[{now_text()}] 批量快速派发启动 | 可发送={len(ordered)} | {review_note}")
    for index, item in enumerate(ordered):
        row_start = time.perf_counter()
        try:
            verify_before_send(item)
            planned_send_at = None
            if args.respect_send_time and item.scheduled_at:
                planned_send_at = scheduled_datetime_for_item(item)
                log_line(
                    stats,
                    (
                        f"[{now_text()}] 行{item.row} | {item.target} | 计划发送准备 | "
                        f"计划={planned_send_at.strftime('%Y-%m-%d %H:%M:%S')}"
                    ),
                )
                wait_for_schedule_prepare(
                    planned_send_at,
                    args.schedule_wait_max_minutes,
                    args.schedule_prepare_lead_seconds,
                )
                add_step_event(item, "计划发送准备", "已完成", f"计划={planned_send_at.strftime('%Y-%m-%d %H:%M:%S')}")
            if args.assume_current_target:
                gui.confirm_target(item, gui.activate())
                gui.current_target_key = item.cache_key
                item.search_seconds = 0.0
            else:
                item.search_seconds = gui.open_target(item)
            if planned_send_at is not None:
                wait_until_before_scheduled_time(planned_send_at, args.schedule_final_target_check_lead_seconds)
                gui.confirm_target(item, gui.activate())
                add_step_event(
                    item,
                    "计划发送前目标复核",
                    "已完成",
                    f"计划前{args.schedule_final_target_check_lead_seconds:.1f}s完成目标会话复核",
                    item.target_screenshot,
                )
                wait_until_scheduled_time(planned_send_at)
                add_step_event(item, "计划发送到点", "已到达", f"计划={planned_send_at.strftime('%Y-%m-%d %H:%M:%S')}")
            item.send_seconds, item.sent_steps = gui.send_message(item, wait_settle=False, dispatch_confirm=True)
            if planned_send_at is not None:
                item.schedule_status, item.schedule_detail = classify_scheduled_send_action_time(
                    planned_send_at,
                    args.schedule_on_time_tolerance_seconds,
                    item.send_action_time,
                )
                add_step_event(item, "计划发送时间核对", item.schedule_status, item.schedule_detail)
                log_line(stats, f"[{now_text()}] 行{item.row} | {item.target} | {item.schedule_status} | {item.schedule_detail}")
            item.postcheck_status = "待人工确认" if args.blind_dispatch_fast else "待统一复核"
            item.postcheck_time = now_text()
            item.postcheck_detail = (
                "blind-dispatch-fast：已执行点击发送，跳过统一复核，需人工确认"
                if args.blind_dispatch_fast
                else "已点击发送，等待批量派发结束后统一复核"
            )
            item.total_seconds = time.perf_counter() - row_start
            dispatched.append((item, row_start))
            if args.blind_dispatch_fast:
                stats.success += 1
                store.mark_row(item, "已派发待确认", "", stats, mode)
                result_by_row[item.row] = row_result_payload(item, "dispatched_pending_review")
            else:
                store.mark_row(item, "发送待复核", "", stats, mode)
            log_line(
                stats,
                (
                    f"[{now_text()}] 行{item.row} | {item.target} | "
                    f"{'已派发待人工确认' if args.blind_dispatch_fast else '已点击发送待复核'} | "
                    f"步骤={' -> '.join(item.sent_steps)} | {item.total_seconds:.3f}s"
                ),
            )
        except Exception as exc:
            item.total_seconds = time.perf_counter() - row_start
            if not item.postcheck_status:
                item.postcheck_status = "未执行" if item.precheck_status == "未通过" else "未通过"
                item.postcheck_time = now_text()
                item.postcheck_detail = str(exc)
            runtime_skip = is_runtime_skip_error(str(exc)) and not item.sent_steps
            if runtime_skip:
                stats.skipped += 1
                row_status = "已跳过"
                result_status = "skipped"
            else:
                stats.failed += 1
                row_status = "发送失败"
                result_status = "failed"
            store.mark_row(item, row_status, str(exc), stats, mode)
            log_line(
                stats,
                (
                    f"[{now_text()}] 行{item.row} | {item.target} | {row_status} | "
                    f"步骤={' -> '.join(item.sent_steps) if item.sent_steps else '无'} | 原因={brief_log_text(str(exc))}"
                ),
            )
            result_by_row[item.row] = row_result_payload(item, result_status, str(exc))
            gui.current_target_key = ""
            gui.confirmed_target_key = ""
            if isinstance(exc, WeComBlockedError):
                blocked_reason = f"企业微信状态阻断，本轮已停止，后续对象未发送；首个阻断原因={exc}"
                log_line(stats, f"[{now_text()}] 批次阻断 | {brief_log_text(blocked_reason)}")
                for payload in mark_rows_blocked_by_wecom_state(store, ordered[index + 1 :], stats, mode, blocked_reason):
                    result_by_row[payload["row"]] = payload
                break
            if args.stop_on_error:
                break
        if args.save_each_row:
            store.save()

    stats.dispatch_completed_at = now_text()
    stats.dispatch_seconds = time.perf_counter() - phase_start
    log_line(
        stats,
        (
            f"[{stats.dispatch_completed_at}] 批量快速派发结束 | 已点击发送={len(dispatched)} | "
            f"发送完成耗时={stats.dispatch_seconds:.3f}s"
            f"{' | 已跳过统一复核，需人工确认' if args.blind_dispatch_fast else ' | 开始统一复核'}"
        ),
    )
    if args.blind_dispatch_fast:
        stats.verification_completed_at = stats.dispatch_completed_at
        stats.verification_seconds = stats.dispatch_seconds
        log_line(
            stats,
            (
                f"[{stats.verification_completed_at}] blind-dispatch-fast 派发完成 | "
                f"已派发待确认={len(dispatched)} | 未执行统一复核"
            ),
        )
        return [result_by_row[row.row] for row in ordered if row.row in result_by_row]

    for verify_index, (item, row_start) in enumerate(dispatched):
        try:
            gui.open_target(item)
            gui.confirm_target(item, gui.activate())
            before_screenshot = Path(item.before_screenshot) if item.before_screenshot else None
            gui.wait_for_send_settle(item, gui.activate(), before_screenshot)
            verify_after_send(item)
            item.total_seconds = time.perf_counter() - row_start
            stats.success += 1
            row_status = "已发送" if not args.resend else "已重发"
            row_error = ""
            result_status = "success"
            if item.schedule_status == "超时已发送":
                stats.late_success += 1
                row_status = "超时已发送"
                row_error = item.schedule_detail
                result_status = "late_success"
            store.mark_row(item, row_status, row_error, stats, mode)
            log_line(stats, f"[{now_text()}] 行{item.row} | {item.target} | 统一复核通过 | {item.total_seconds:.3f}s")
            result_by_row[item.row] = row_result_payload(item, result_status)
        except Exception as exc:
            item.total_seconds = time.perf_counter() - row_start
            item.postcheck_status = "未通过"
            item.postcheck_time = now_text()
            item.postcheck_detail = str(exc)
            stats.failed += 1
            store.mark_row(item, "发送待核对", str(exc), stats, mode)
            log_line(
                stats,
                f"[{now_text()}] 行{item.row} | {item.target} | 统一复核未通过 | 原因={brief_log_text(str(exc))}",
            )
            result_by_row[item.row] = row_result_payload(item, "verification_failed", str(exc))
            gui.current_target_key = ""
            gui.confirmed_target_key = ""
            if isinstance(exc, WeComBlockedError):
                blocked_reason = f"企业微信状态阻断，统一复核已停止，剩余已点击发送对象需人工核对；首个阻断原因={exc}"
                log_line(stats, f"[{now_text()}] 统一复核阻断 | {brief_log_text(blocked_reason)}")
                remaining_dispatched = [remaining for remaining, _start in dispatched[verify_index + 1 :]]
                for payload in mark_rows_blocked_by_wecom_state(
                    store,
                    remaining_dispatched,
                    stats,
                    mode,
                    blocked_reason,
                    row_status="发送待核对",
                    result_status="verification_failed",
                ):
                    result_by_row[payload["row"]] = payload
                break
            if args.stop_on_error:
                break
        if args.save_each_row:
            store.save()

    stats.verification_completed_at = now_text()
    stats.verification_seconds = time.perf_counter() - phase_start
    log_line(
        stats,
        (
            f"[{stats.verification_completed_at}] 统一复核结束 | 已复核={len(result_by_row)} | "
            f"验证完成耗时={stats.verification_seconds:.3f}s"
        ),
    )
    return [result_by_row[row.row] for row in ordered if row.row in result_by_row]


def preview_payload(workbook: Path, rows: list[SendRow], stats: RunStats) -> dict[str, Any]:
    sendable = [row for row in rows if row.should_send]
    skipped = [row for row in rows if not row.should_send]
    return {
        "workbook": str(workbook),
        "batch_id": stats.batch_id,
        "started_at": stats.started_at,
        "ended_at": stats.ended_at,
        "total_seconds": round(stats.total_seconds, 3),
        "run_dir": str(stats.run_dir) if stats.run_dir else "",
        "log_path": str(stats.log_path) if stats.log_path else "",
        "total_rows": len(rows),
        "sendable_rows": len(sendable),
        "skipped_rows": len(skipped),
        "target_count": len({row.cache_key for row in sendable}),
        "rows": [row.as_preview() for row in rows],
    }


def print_preview(payload: dict[str, Any]) -> None:
    print(f"文件: {payload['workbook']}")
    print(f"批次: {payload['batch_id']}")
    print(f"可发送行数: {payload['sendable_rows']} / 总行数: {payload['total_rows']}")
    print(f"去重目标数: {payload['target_count']}")
    print("| 行号 | 渠道 | 对象类型 | 发送对象 | 消息类型 | 图片数 | 文档数 | 是否发送 | 原因 |")
    print("|---:|---|---|---|---|---:|---:|---|---|")
    for row in payload["rows"]:
        print(
            f"| {row['row']} | {row['channel']} | {row['object_type']} | {row['target']} | "
            f"{row['message_type']} | {len(row['image_paths'])} | {len(row['document_paths'])} | "
            f"{'是' if row['should_send'] else '否'} | {row['reason']} |"
        )


def parse_sender_args() -> Any:
    parser = argparse.ArgumentParser(description="Fast Excel-driven Enterprise WeChat sender")
    parser.add_argument("--folder", default=".")
    parser.add_argument("--workbook")
    parser.add_argument("--lesson-workbook", help="Course/reminder workbook used to generate a send workbook in this run directory.")
    parser.add_argument("--target-workbook", help="Target workbook whose 渠道/对象类型/发送对象 columns receive the generated lesson message.")
    parser.add_argument("--lesson", help="Lesson identifier such as 3 or 第3课 when using --lesson-workbook.")
    parser.add_argument("--lesson-content", default="课程提醒", help="Lesson message kind: 课程提醒 or 课后总结.")
    parser.add_argument("--single-target", help="Generate a one-row lesson send workbook for this WeCom target.")
    parser.add_argument("--single-target-type", default="个人", help="Object type for --single-target, usually 个人 or 群聊.")
    parser.add_argument("--single-target-alias", default="", help="OCR confirmation aliases for --single-target.")
    parser.add_argument("--single-channel", default="企业微信", help="Channel for --single-target.")
    parser.add_argument("--run-at", default="", help="Override the lesson workbook send time, for example 11:30 or 2026-05-29 11:30.")
    parser.add_argument("--no-dedupe-targets", action="store_true", help="Do not merge duplicate targets when generating a lesson send workbook.")
    parser.add_argument("--row", type=int, action="append", help="Only process the given Excel row; can be repeated.")
    parser.add_argument("--row-from", type=int, default=0, help="Only process Excel rows from this 1-based row number onward.")
    parser.add_argument("--row-to", type=int, default=0, help="Only process Excel rows up to this 1-based row number.")
    parser.add_argument("--run-dir", default="", help="Use an existing run log folder instead of creating a new one.")
    parser.add_argument("--batch-id", default="", help="Reuse a caller-provided batch id for a multi-step task.")
    parser.add_argument("--execute", action="store_true", help="Actually send messages through WeCom.")
    parser.add_argument("--yes", action="store_true", help="Required with --execute to avoid interactive confirmation.")
    parser.add_argument("--resend", action="store_true", help="Allow rows already marked 已发送/已重发.")
    parser.add_argument("--group-targets", action="store_true", help="Send rows grouped by target to reduce searches.")
    parser.add_argument("--save-each-row", action="store_true", help="Safer but slower Excel write-back after every row.")
    parser.add_argument("--stop-on-error", action="store_true")
    parser.add_argument("--wecom-exe")
    parser.add_argument("--mac-app-name", default="企业微信", help="macOS 下企业微信应用名，例如 企业微信 或 WeCom。")
    parser.add_argument("--skip-target-check", action="store_true", help="Skip OCR confirmation of the opened chat target before sending.")
    parser.add_argument("--assume-current-target", action="store_true", help="Use the currently open chat after title confirmation instead of searching.")
    parser.add_argument("--skip-search-box-check", action="store_true", help="Skip OCR confirmation that the search box received the target text before pressing Enter.")
    parser.add_argument("--no-evidence", action="store_true", help="Do not save after-send screenshots.")
    parser.add_argument("--low-evidence", action="store_true", help="Use screenshots for OCR/checks during each row, then delete temporary evidence files after the row finishes.")
    parser.add_argument("--ocr-script", default=str(OCR_SCRIPT_PATH), help="PowerShell OCR helper used for target confirmation.")
    parser.add_argument("--search-retries", type=int, default=3)
    parser.add_argument("--search-wait", type=float, default=0.22)
    parser.add_argument("--chat-wait", type=float, default=0.22)
    parser.add_argument("--text-wait", type=float, default=0.22)
    parser.add_argument("--file-wait", type=float, default=0.45, help="Minimum wait after sending one file group.")
    parser.add_argument("--file-wait-max", type=float, default=1.25, help="Maximum wait after sending one file group.")
    parser.add_argument("--file-wait-per-mb", type=float, default=0.65, help="Additional wait per MB for file sends.")
    parser.add_argument("--send-settle-timeout", type=float, default=45.0)
    parser.add_argument("--send-settle-interval", type=float, default=0.6)
    parser.add_argument("--between-rows", type=float, default=0.02)
    parser.add_argument(
        "--max-consecutive-ocr-no-text",
        type=int,
        default=3,
        help="Stop the batch after this many consecutive empty OCR target checks; set 0 to keep recording row failures and continue.",
    )
    parser.add_argument(
        "--safe-fast",
        action="store_true",
        help="Keep target checks but use a shorter post-send verification path: input cleared, no red failure marker, and a new outgoing bubble.",
    )
    parser.add_argument(
        "--paste-method-order",
        default=DEFAULT_PASTE_METHOD_ORDER,
        help="Comma-separated text paste attempts. Default is repeated ctrl-v; optional methods: wm-point, wm-window, shift-insert.",
    )
    parser.add_argument(
        "--trust-clipboard-paste",
        action="store_true",
        help="After target confirmation and verified clipboard write, allow partial input OCR matches to send; clear input on unrelated/empty paste checks to avoid drafts.",
    )
    parser.add_argument(
        "--fast-input-check",
        action="store_true",
        help="After target confirmation and verified clipboard paste, use AX/visual input-box checks before falling back to OCR.",
    )
    parser.add_argument(
        "--lean-evidence-fast",
        action="store_true",
        help="Fast mode with less evidence: enable safe-fast/fast-input-check, skip search-box OCR, and skip before-send full-window screenshots.",
    )
    parser.add_argument(
        "--ultra-fast-dispatch",
        action="store_true",
        help="More aggressive fast mode: direct Enter after search, skip current-chat precheck and title OCR; keep title-change and send-failure visual checks.",
    )
    parser.add_argument(
        "--dispatch-only-fast",
        action="store_true",
        help="Fastest batch dispatch mode: after clicking send, short-wait and continue without post-send screenshot/bubble checks.",
    )
    parser.add_argument(
        "--blind-dispatch-fast",
        action="store_true",
        help="Experimental fastest mode: direct Enter, skip target title OCR/change check and input screenshot confirmation; use only for small tests first.",
    )
    parser.add_argument(
        "--normalize-window-size",
        action="store_true",
        help="Before sending, resize the WeCom main window to a fixed size so coordinate-based clicks are more consistent across devices.",
    )
    parser.add_argument("--window-width", type=int, default=1092, help="Target WeCom window width used with --normalize-window-size.")
    parser.add_argument("--window-height", type=int, default=818, help="Target WeCom window height used with --normalize-window-size.")
    parser.add_argument("--respect-send-time", action="store_true", help="Prepare the target before 计划发送时间/定时发送时间, never click send before that time, and record late sends.")
    parser.add_argument("--ignore-send-time", action="store_true", help="Do not wait for lesson/workbook planned send time.")
    parser.add_argument("--batch-fast-dispatch", action="store_true", help="Click-send all rows first after target/input checks, then reopen chats for unified post-send verification.")
    parser.add_argument("--no-auto-batch-fast-dispatch", action="store_true", help="Do not automatically enable batch fast dispatch for multi-row runs.")
    parser.add_argument("--batch-fast-dispatch-threshold", type=int, default=DEFAULT_BATCH_FAST_DISPATCH_THRESHOLD, help="Automatically use batch fast dispatch when a scheduled send group has at least this many targets.")
    parser.add_argument("--batch-dispatch-estimate-per-target-seconds", type=float, default=DEFAULT_BATCH_DISPATCH_ESTIMATE_PER_TARGET_SECONDS, help="Estimated GUI dispatch seconds per target for batch schedule warnings.")
    parser.add_argument("--schedule-wait-max-minutes", type=float, default=360.0)
    parser.add_argument("--schedule-prepare-lead-seconds", type=float, default=None, help="Start target search/check this many seconds before the planned send time. Default is dynamic: short for one target, longer for batch fast dispatch.")
    parser.add_argument("--schedule-final-target-check-lead-seconds", type=float, default=DEFAULT_SCHEDULE_FINAL_TARGET_CHECK_LEAD_SECONDS, help="Re-check the opened chat this many seconds before the planned time so OCR cost does not delay click-send.")
    parser.add_argument("--schedule-on-time-tolerance-seconds", type=float, default=DEFAULT_SCHEDULE_ON_TIME_TOLERANCE_SECONDS, help="Classify as on-time if send starts within this many seconds after the planned time; late sends still proceed.")
    parser.add_argument("--schedule-grace-minutes", type=float, default=1.0, help=argparse.SUPPRESS)
    parser.add_argument("--mvp-simple", action="store_true", help="Use minimal MVP flow: search target -> send -> short wait -> mark sent/failed.")
    parser.add_argument("--mvp-wait-after-send", type=float, default=0.8, help="Seconds to wait after click-send in --mvp-simple mode.")
    parser.add_argument(
        "--test-disclaimer",
        default=DEFAULT_TEST_DISCLAIMER,
        help=f"Append this note to the end of every text message. Example: \"{DEFAULT_TEST_DISCLAIMER}\"",
    )
    parser.add_argument("--quiet-log", action="store_true", help="Disable realtime progress output; still writes run_log.txt.")
    parser.add_argument("--reset-input-locator-cache", action="store_true", help="Delete cached chat input coordinates before this run so they can be relearned after a WeCom UI update.")
    parser.add_argument("--json", action="store_true")
    args = parser.parse_args()
    if args.blind_dispatch_fast:
        args.dispatch_only_fast = True
        args.normalize_window_size = True
        args.chat_wait = min(args.chat_wait, 0.18)
        args.text_wait = min(args.text_wait, 0.08)
        args.between_rows = min(args.between_rows, 0.02)
    if args.dispatch_only_fast:
        args.ultra_fast_dispatch = True
        args.batch_fast_dispatch = True
    if args.ultra_fast_dispatch:
        args.lean_evidence_fast = True
        args.skip_search_box_check = True
    if args.lean_evidence_fast:
        args.safe_fast = True
        args.fast_input_check = True
        args.skip_search_box_check = True
        args.text_wait = min(args.text_wait, 0.12)
        args.send_settle_interval = min(args.send_settle_interval, 0.2)
        args.send_settle_timeout = min(args.send_settle_timeout, 10.0)
    if args.safe_fast:
        args.text_wait = min(args.text_wait, 0.15)
        args.send_settle_interval = min(args.send_settle_interval, 0.25)
        args.send_settle_timeout = min(args.send_settle_timeout, 12.0)
    if args.mvp_simple:
        args.skip_target_check = True
        args.skip_search_box_check = True
        args.no_evidence = True
        args.safe_fast = False
        args.batch_fast_dispatch = False
        args.respect_send_time = False
    return args


def main() -> None:
    args = parse_sender_args()

    program_start = time.perf_counter()
    folder = Path(args.folder).resolve()
    if args.reset_input_locator_cache:
        cache_path = input_locator_cache_path_for_folder(folder)
        try:
            cache_path.unlink()
        except FileNotFoundError:
            pass
        except OSError as exc:
            raise SystemExit(f"输入框定位缓存清理失败：{exc}") from exc
    lesson_mode = bool(args.lesson_workbook or args.target_workbook or args.lesson or args.single_target)
    if lesson_mode and not (args.lesson_workbook and args.lesson):
        raise SystemExit("课程发送模式需要传入 --lesson-workbook --lesson，并通过 --target-workbook 或 --single-target 指定发送对象")
    if lesson_mode and bool(args.target_workbook) == bool(args.single_target):
        raise SystemExit("课程发送模式需要且只能指定一个目标来源：--target-workbook 或 --single-target")
    if lesson_mode and args.workbook:
        raise SystemExit("--workbook 不能与 --lesson-workbook/--target-workbook/--lesson 同时使用")
    workbook = (
        Path(args.lesson_workbook).resolve()
        if lesson_mode
        else Path(args.workbook).resolve()
        if args.workbook
        else find_workbook(folder)
    )
    stats = RunStats(batch_id=args.batch_id) if args.batch_id else RunStats()
    stats.live_log = not args.quiet_log
    init_run_logging(stats, folder, workbook, execute=args.execute, requested_run_dir=args.run_dir)
    log_line(stats, f"[{now_text()}] 运行环境 | {runtime_dependency_status()}")
    if lesson_mode:
        _message_header, content_label = normalize_lesson_content(args.lesson_content)
        target_suffix = "targets"
        if args.single_target:
            target_suffix = safe_filename(args.single_target, "target")
        content_part = "" if content_label == "课程提醒" else f"_{safe_filename(content_label, 'content')}"
        generated_name = f"send_{safe_filename(normalize_lesson_key(args.lesson), 'lesson')}{content_part}_to_{target_suffix}.xlsx"
        generated_path = (stats.run_dir or folder) / generated_name
        if args.single_target:
            build_info = build_lesson_single_target_workbook(
                Path(args.lesson_workbook).resolve(),
                args.lesson,
                generated_path,
                target=args.single_target,
                object_type=args.single_target_type,
                channel=args.single_channel,
                aliases=args.single_target_alias,
                scheduled_at_override=args.run_at,
                lesson_content=args.lesson_content,
            )
        else:
            build_info = build_lesson_target_workbook(
                Path(args.lesson_workbook).resolve(),
                Path(args.target_workbook).resolve(),
                args.lesson,
                generated_path,
                dedupe_targets=not args.no_dedupe_targets,
                scheduled_at_override=args.run_at,
                lesson_content=args.lesson_content,
        )
        workbook = generated_path
        args.respect_send_time = not args.ignore_send_time
        duplicate_text = f" | 重复合并={len(build_info['duplicates'])}" if build_info["duplicates"] else ""
        log_line(
            stats,
            (
                f"[{now_text()}] 课程发送表已生成 | 课程={build_info['lesson']} | 内容={build_info.get('content', '课程提醒')} | "
                f"计划={build_info['scheduled_at'] or '无'} | 可发送行={build_info['send_rows']}"
                f"{duplicate_text} | 文件={generated_path.name}"
            ),
        )
    only_rows = set(args.row or []) or None

    store = WorkbookStore(workbook, folder, ensure_status=args.execute)
    rows = store.build_rows(allow_sent=args.resend, only_rows=only_rows)
    if args.row_from:
        rows = [row for row in rows if row.row >= args.row_from]
    if args.row_to:
        rows = [row for row in rows if row.row <= args.row_to]
    if args.test_disclaimer:
        updated_rows = apply_test_disclaimer(rows, args.test_disclaimer)
        if updated_rows:
            log_line(stats, f"[{now_text()}] 测试尾注已附加 | 行数={updated_rows} | 尾注={cell_text(args.test_disclaimer)}")
    if args.run_at and not lesson_mode:
        planned = parse_scheduled_datetime(args.run_at)
        if planned is None:
            raise SystemExit(f"--run-at 时间无法识别：{args.run_at}")
        planned_text = planned.strftime("%Y-%m-%d %H:%M:%S")
        for row in rows:
            if row.should_send:
                row.scheduled_at = planned_text
        args.respect_send_time = True
        log_line(stats, f"[{now_text()}] 直接工作簿发送时间已覆盖 | 计划={planned_text}")
    sendable_count = sum(1 for row in rows if row.should_send)
    stats.skipped = len(rows) - sendable_count
    log_line(
        stats,
        (
            f"[{now_text()}] Excel读取完成 | 总行={len(rows)} | "
            f"可发送={sendable_count} | 跳过={stats.skipped}"
        ),
    )
    if (
        not args.respect_send_time
        and not args.batch_fast_dispatch
        and not args.no_auto_batch_fast_dispatch
        and not args.mvp_simple
        and sendable_count > 1
    ):
        args.batch_fast_dispatch = True
        log_line(stats, f"[{now_text()}] 自动启用批量快速派发 | 原因=先完成全部发送，再统一验证")
    if args.respect_send_time and not args.batch_fast_dispatch and not args.no_auto_batch_fast_dispatch and should_auto_batch_fast_dispatch(
        rows,
        threshold=max(2, args.batch_fast_dispatch_threshold),
    ):
        args.batch_fast_dispatch = True
        log_line(stats, f"[{now_text()}] 自动启用批量快速派发 | 原因=同一计划时间存在多个发送对象")
    if args.respect_send_time:
        for planned_time, group_rows in sorted(scheduled_groups_for_rows(rows).items()):
            target_count, estimated_seconds = batch_dispatch_estimate(group_rows, args.batch_dispatch_estimate_per_target_seconds)
            window_seconds = max(0.0, args.schedule_on_time_tolerance_seconds)
            planned_text = planned_time.strftime("%Y-%m-%d %H:%M:%S")
            if target_count > 1:
                status = "预计可在准时窗口内完成" if estimated_seconds <= window_seconds else "预计无法全部在准时窗口内完成"
                log_line(
                    stats,
                    (
                        f"[{now_text()}] 批量派发估算 | 计划={planned_text} | 对象={target_count} | "
                        f"估算派发={estimated_seconds:.1f}s | 准时窗口={window_seconds:.1f}s | {status}"
                    ),
                )
    if args.schedule_prepare_lead_seconds is None:
        args.schedule_prepare_lead_seconds = schedule_prepare_lead_seconds_for_rows(
            rows,
            batch_fast_dispatch=args.batch_fast_dispatch,
        )
        if args.respect_send_time and args.schedule_prepare_lead_seconds > 0:
            log_line(stats, f"[{now_text()}] 定时发送动态提前量 | {args.schedule_prepare_lead_seconds:.1f}s")

    if not args.execute:
        stats.ended_at = now_text()
        stats.total_seconds = time.perf_counter() - program_start
        payload = preview_payload(workbook, rows, stats)
        result_path = save_run_result(stats, payload)
        log_line(stats, f"[{stats.ended_at}] 预览完成 | {stats.total_seconds:.3f}s | 结果={result_path.name}")
        if args.json:
            print(json.dumps(payload, ensure_ascii=False, indent=2))
        else:
            print_preview(payload)
        return

    if not args.yes:
        raise SystemExit("真实发送需要同时传入 --execute --yes")

    if not any(row.should_send for row in rows):
        stats.ended_at = now_text()
        stats.total_seconds = time.perf_counter() - program_start
        payload = preview_payload(workbook, rows, stats)
        result_path = save_run_result(stats, payload)
        log_line(stats, f"[{stats.ended_at}] 无可发送行 | {stats.total_seconds:.3f}s | 结果={result_path.name}")
        if args.json:
            print(json.dumps(payload, ensure_ascii=False, indent=2))
        else:
            print_preview(payload)
        return

    if args.mvp_simple:
        results = execute_rows_mvp(store, rows, args, stats)
    elif args.batch_fast_dispatch:
        results = execute_rows_batch_fast_dispatch(store, rows, args, stats)
    else:
        results = execute_rows(store, rows, args, stats)
    stats.ended_at = now_text()
    stats.total_seconds = time.perf_counter() - program_start
    if not stats.verification_completed_at:
        stats.verification_completed_at = stats.ended_at
        stats.verification_seconds = stats.total_seconds
    store.mark_batch_total(rows, stats)
    store.save()
    archived_saved_path, archive_detail = archive_generated_workbook_to_run_dir(stats, store.saved_path)
    if archived_saved_path != store.saved_path:
        store.saved_path = archived_saved_path
        workbook = archived_saved_path
        log_line(stats, f"[{now_text()}] 临时发送表已归档 | {archived_saved_path.name} | {archive_detail}")
    elif archive_detail:
        log_line(stats, f"[{now_text()}] 临时发送表归档未完成 | {archive_detail}")
    evidence_manifest_path = save_evidence_manifest(stats, rows)
    log_line(
        stats,
        (
            f"[{stats.ended_at}] 批次{stats.batch_id} 完成 | 成功={stats.success} 超时已发送={stats.late_success} "
            f"失败={stats.failed} 跳过={stats.skipped} | {stats.total_seconds:.3f}s | 写回={store.writeback_mode}"
        ),
    )
    if stats.dispatch_completed_at:
        log_line(
            stats,
            (
                f"[{now_text()}] 批量耗时汇总 | 发送完成={stats.dispatch_seconds:.3f}s "
                f"({stats.dispatch_completed_at}) | 验证完成={stats.verification_seconds:.3f}s "
                f"({stats.verification_completed_at})"
            ),
        )
    log_line(stats, f"[{now_text()}] 证据清单已保存 | {evidence_manifest_path.name}")

    payload = {
        "workbook": str(workbook),
        "batch_id": stats.batch_id,
        "started_at": stats.started_at,
        "ended_at": stats.ended_at,
        "total_seconds": round(stats.total_seconds, 3),
        "dispatch_completed_at": stats.dispatch_completed_at,
        "dispatch_seconds": round(stats.dispatch_seconds, 3) if stats.dispatch_seconds else None,
        "verification_completed_at": stats.verification_completed_at,
        "verification_seconds": round(stats.verification_seconds, 3),
        "success": stats.success,
        "late_success": stats.late_success,
        "failed": stats.failed,
        "skipped": stats.skipped,
        "writeback_mode": store.writeback_mode,
        "saved_path": str(store.saved_path),
        "run_dir": str(stats.run_dir) if stats.run_dir else "",
        "log_path": str(stats.log_path) if stats.log_path else "",
        "evidence_manifest_path": str(evidence_manifest_path),
        "results": results,
    }
    result_path = save_run_result(stats, payload)
    log_line(stats, f"[{now_text()}] 结果已保存 | {result_path.name}")
    if args.json:
        print(json.dumps(payload, ensure_ascii=False, indent=2))
    else:
        print(f"批次 {stats.batch_id} 完成：成功 {stats.success}（超时已发送 {stats.late_success}），失败 {stats.failed}，跳过 {stats.skipped}，总周期 {stats.total_seconds:.3f}s")
        if stats.dispatch_completed_at:
            print(f"发送完成不含统一验证：{stats.dispatch_seconds:.3f}s；加上统一验证：{stats.verification_seconds:.3f}s")


if __name__ == "__main__":
    main()
